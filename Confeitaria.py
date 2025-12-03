# -*- coding: utf-8 -*-
# Confeitaria Pro – Sistema de Pedidos (PyQt6 + SQLite)
# -----------------------------------------------------
# Requisitos:
#   pip install PyQt6
#   # (opcional)
#   pip install PyQt6-Charts qtawesome matplotlib reportlab bcrypt pyyaml
#   pip install openpyxl
#
# Observações:
# - Projeto single-file para facilitar testes aqui. Em produção, modularizar.
# - Interface com design moderno via QSS e layout com Sidebar + Header + Pages.
# - Banco de dados SQLite local: ./confeitaria.db
# - Módulos: Pedidos, Produtos, Clientes, Relatórios, Configurações
# - CRUD básico para Produtos, Clientes e Pedidos (com cálculo de total)
# - Relatórios com métricas simples (vendas do mês, itens mais vendidos)
# - ✅ Melhorias e correções neste arquivo:
#   * Tema dark/light refinado (QSS)
#   * Toast notifications não intrusivas
#   * Backup automático semanal do banco
#   * Alertas de estoque baixo em Produtos
#   * Tratamento de erros com mensagens amigáveis
#   * Ícones via qtawesome (se disponível)
#   * Fallback robusto para caminho do DB quando __file__ não existe
#   * Helper tipado `show_message` (remove avisos "partially unknown" do Pylance)
#   * `dashboard_cb` declarado em CustomersPage
#   * Anotações (QComboBox/QTableWidget) e casts em `.connect(...)`
#   * openpyxl com get_column_letter (evita MergedCell/column_letter)
#   * Remoção de TYPE_CHECKING e casts desnecessários
#
# Como executar:
#   python Confeitaria.py

from __future__ import annotations

import os
import sys
import sqlite3
import hashlib
import json
import urllib.request
import urllib.error
import ctypes
import threading
from typing import Optional, Any, Callable, TypeVar, Protocol, Sequence, cast
from datetime import datetime, date, timedelta, timezone

T = TypeVar("T")

# ---------------------------------------------------------------------
# Configurações da aplicação
# ---------------------------------------------------------------------
# Local da aplicação (.py ou .exe)
if getattr(sys, "frozen", False):  # executável (PyInstaller)
    APP_DIR = os.path.dirname(sys.executable)
else:                               # script .py
    APP_DIR = os.path.dirname(os.path.abspath(__file__))

from PyQt6.QtCore import Qt, QSize, QTimer, QEasingCurve, QPropertyAnimation, QPoint, QDate, QThread, pyqtSignal
from PyQt6.QtGui import QAction, QIcon, QColor, QFont
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QListWidget, QListWidgetItem, QStackedWidget, QTableWidget,
    QTableWidgetItem, QAbstractItemView, QFormLayout, QLineEdit, QSpinBox,
    QDoubleSpinBox, QComboBox, QDateEdit, QTextEdit, QDialog, QDialogButtonBox,
    QFrame, QGraphicsDropShadowEffect, QHeaderView, QTabWidget, QFileDialog, QMessageBox,
    QMenu, QInputDialog, QColorDialog, QSizePolicy, QProgressDialog, QScrollArea, QCheckBox, QGroupBox, QProgressBar
)

# -----------------------------
# Imports do seu projeto + fallbacks seguros
# -----------------------------
try:
    from ui.dialogs.custom_filedialog import CustomFileDialog  # sua UI
except Exception:
    class CustomFileDialog(QDialog):
        """Fallback simples para salvar/abrir arquivo mantendo API usada no código."""
        def __init__(self, parent: Optional[QWidget] = None, caption: str = "", directory: str = "",
                     filter: str = "", save_mode: bool = False) -> None:
            super().__init__(parent)
            self._selected: Optional[str] = None
            self._caption = caption or "Selecionar arquivo"
            self._directory = directory or os.getcwd()
            self._filter = filter or "Todos (*.*)"
            self._save_mode = save_mode

        def exec(self) -> int:  # type: ignore[override]
            if self._save_mode:
                path, _ = QFileDialog.getSaveFileName(self, self._caption, self._directory, self._filter)
            else:
                path, _ = QFileDialog.getOpenFileName(self, self._caption, self._directory, self._filter)
            self._selected = path or None
            return QDialog.DialogCode.Accepted if self._selected else QDialog.DialogCode.Rejected

        def get_selected_file(self) -> Optional[str]:
            return self._selected

try:
    from core.database import Database
except Exception:
    # Fallback mínimo do seu Database para testes locais
    class Database:
        def __init__(self, path: str) -> None:
            # Conexão com PRAGMAs para melhor concorrência quando usando fallback local
            self.conn = sqlite3.connect(path, timeout=30, check_same_thread=False)
            self.conn.row_factory = sqlite3.Row
            try:
                cur = self.conn.cursor()
                cur.execute("PRAGMA foreign_keys=ON")
                cur.execute("PRAGMA journal_mode=WAL")
                cur.execute("PRAGMA synchronous=NORMAL")
                cur.execute("PRAGMA busy_timeout=30000")
                cur.execute("PRAGMA temp_store=MEMORY")
                self.conn.commit()
            except Exception:
                pass
            self._init_db()

        def _init_db(self) -> None:
            cur = self.conn.cursor()
            cur.execute("""CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE,
                password_hash TEXT,
                role TEXT DEFAULT 'user'
            )""")
            self.conn.commit()

        def query(self, sql: str, params: tuple = ()) -> list[sqlite3.Row]:
            cur = self.conn.cursor()
            cur.execute(sql, params)
            return cur.fetchall()

        def execute(self, sql: str, params: tuple = ()) -> sqlite3.Cursor:
            cur = self.conn.cursor()
            cur.execute(sql, params)
            self.conn.commit()
            return cur

DB = Database  # alias p/ typing local

# ---------------------------------------------------------------------
# NETWORK PATH DETECTION & ASYNC VALIDATION
# ---------------------------------------------------------------------

def is_network_path(path: str) -> bool:
    """Detecta se o caminho é uma unidade de rede (UNC ou mapeada)."""
    if not path:
        return False
    
    # UNC path (\\servidor\compartilhamento)
    if path.startswith('\\\\') or path.startswith('//'):
        return True
    
    # Verificar se é unidade mapeada (Windows)
    if sys.platform == 'win32' and len(path) >= 2 and path[1] == ':':
        try:
            import subprocess
            drive_letter = path[0].upper()
            result = subprocess.run(['net', 'use'], capture_output=True, text=True, timeout=2)
            if result.returncode == 0 and drive_letter in result.stdout:
                return True
        except Exception:
            pass
    
    return False

class AsyncDatabaseValidator(QThread):
    """Thread para validar banco de dados de forma assíncrona (evita travar UI)."""
    finished = pyqtSignal(bool, str, str)  # (is_valid, message, file_path)
    progress = pyqtSignal(str)  # mensagem de progresso
    
    def __init__(self, file_path: str, timeout: int = 15):
        super().__init__()
        self.file_path = file_path
        self.timeout = timeout
        self._stop = False
        
    def stop(self):
        """Sinaliza para parar a validação."""
        self._stop = True
        
    def run(self):
        """Executa validação com timeout."""
        try:
            if self._stop:
                return
                
            self.progress.emit("Verificando arquivo...")
            
            # Timeout usando QTimer (funciona melhor com Qt)
            import threading
            result = [None, None]  # [is_valid, message]
            exception = [None]
            
            def validate_with_timeout():
                try:
                    if self._stop:
                        return
                    from core.config import validate_database_path
                    result[0], result[1] = validate_database_path(self.file_path)
                except Exception as e:
                    exception[0] = e
            
            thread = threading.Thread(target=validate_with_timeout, daemon=True)
            thread.start()
            thread.join(timeout=self.timeout)
            
            if self._stop:
                return
                
            if thread.is_alive():
                # Timeout atingido
                self.finished.emit(
                    False,
                    f"⏱️ Timeout ao validar banco de dados ({self.timeout}s).\n\nO arquivo pode estar em rede lenta ou inacessível.",
                    self.file_path
                )
                return
            
            if exception[0]:
                raise exception[0]
                
            if result[0] is None:
                raise Exception("Validação não retornou resultado")
                
            self.finished.emit(result[0], result[1] or "Validado", self.file_path)
            
        except Exception as e:
            if not self._stop:
                self.finished.emit(False, f"Erro ao validar: {str(e)}", self.file_path)

# ---------------------------------------------------------------------
# Sistema simplificado - sem licenciamento (versão local)
# ---------------------------------------------------------------------

# -----------------------------
# Ícones QtAwesome + fallback
# -----------------------------
try:
    import qtawesome as qta  # type: ignore
    
    # Função helper para uso seguro do QtAwesome
    def safe_qta_icon(icon_name: str, color: str = "#000000") -> "QIcon":
        """Retorna ícone QtAwesome com fallback seguro"""
        try:
            return qta.icon(icon_name, color=color)
        except Exception:
            # Fallback para ícone vazio em caso de erro
            return QIcon()
    
except Exception:
    qta = None
    
    # Fallback quando QtAwesome não está disponível
    def safe_qta_icon(icon_name: str, color: str = "#000000") -> "QIcon":
        """Retorna ícone vazio quando QtAwesome não disponível"""
        return QIcon()

# Senha: usa bcrypt se disponível, senão sha256 como fallback (marcado com prefixo)
try:
    import bcrypt  # type: ignore
except Exception:
    bcrypt = None

def hash_password(password: str) -> str:
    """Retorna o hash da senha usando bcrypt, com fallback para SHA256."""
    if not password:
        return ""
    try:
        if bcrypt:
            h = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            return h.decode('utf-8')
        else:
            # Fallback para SHA256 se bcrypt não estiver disponível
            import hashlib
            return hashlib.sha256(password.encode('utf-8')).hexdigest()
    except Exception as e:
        print(f"Error hashing password with bcrypt, trying SHA256 fallback: {e}")
        try:
            # Último recurso: SHA256
            import hashlib
            return hashlib.sha256(password.encode('utf-8')).hexdigest()
        except Exception as e2:
            print(f"Error hashing password with SHA256: {e2}")
            return ""

def verify_password(password: str, stored_hash: str) -> bool:
    """Verifica se a senha corresponde ao hash armazenado, suportando bcrypt e SHA256."""
    if not password or not stored_hash:
        return False
    
    try:
        # Verifica se é hash bcrypt (começa com $2a$, $2b$, $2y$)
        is_bcrypt = isinstance(stored_hash, str) and stored_hash.startswith('$2')
        
        if bcrypt and is_bcrypt:
            # Verifica com bcrypt
            return bcrypt.checkpw(password.encode('utf-8'), stored_hash.encode('utf-8'))
        else:
            # Verifica com SHA256
            import hashlib
            password_hash = hashlib.sha256(password.encode('utf-8')).hexdigest()
            return password_hash == stored_hash
    except Exception as e:
        print(f"Error verifying password: {e}")
        return False

# Caminho robusto do banco (corrige ambientes sem __file__ e PyInstaller)
def get_base_directory() -> str:
    """
    Determina o diretório base da aplicação de forma robusta.
    Funciona tanto em desenvolvimento quanto em executáveis PyInstaller.
    """
    try:
        # Se estamos executando via PyInstaller
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
            # Para dados do usuário, usa o diretório dos dados de aplicação
            import tempfile
            app_data_dir = os.path.join(os.path.expanduser("~"), "AppData", "Local", "Confeitaria")
            os.makedirs(app_data_dir, exist_ok=True)
            return app_data_dir
        else:
            # Modo desenvolvimento
            return os.path.dirname(os.path.abspath(__file__))
    except NameError:
        # Fallback se __file__ não existir
        app_data_dir = os.path.join(os.path.expanduser("~"), "AppData", "Local", "Confeitaria")
        os.makedirs(app_data_dir, exist_ok=True)
        return app_data_dir

base_dir = get_base_directory()
# Usar função para sempre pegar o caminho correto (configurado ou padrão)
def get_db_path():
    """Retorna o caminho do banco sempre atualizado (configurado ou padrão)"""
    try:
        from core.config import get_database_path
        return get_database_path()
    except Exception:
        # Fallback apenas se core.config não estiver disponível
        return os.path.join(APP_DIR, "confeitaria.db")

DB_PATH = get_db_path()  # Inicialização para compatibilidade
BACKUP_DIR = os.path.join(base_dir, "backups")

# -----------------------------
# Utils de UI: Toast
# -----------------------------
class Toast(QFrame):
    """Pequena notificação flutuante no canto inferior direito."""
    def __init__(self, parent: Optional[QWidget], text: str, duration_ms: int = 2200) -> None:
        super().__init__(parent)
        self.setObjectName("Toast")
        self.setWindowFlags(Qt.WindowType.FramelessWindowHint | Qt.WindowType.ToolTip)
        self.setAttribute(Qt.WidgetAttribute.WA_StyledBackground, True)
        lay = QHBoxLayout(self)
        lbl = QLabel(text)
        lay.addWidget(lbl)
        lay.setContentsMargins(14, 10, 14, 10)
        self.setStyleSheet("""
        #Toast { background: rgba(20,24,36,0.95); color: #fff; border-radius: 10px; border:1px solid #2a3350; }
        #Toast QLabel { color: #ffffff; }
        """)
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(18)
        shadow.setOffset(0, 6)
        self.setGraphicsEffect(shadow)
        self._duration = duration_ms
        self._anim = QPropertyAnimation(self, b"pos", self)
        self._anim.setDuration(280)
        self._anim.setEasingCurve(QEasingCurve.Type.OutCubic)

    def show_near_bottom_right(self) -> None:
        parent = self.parentWidget()
        if not parent:
            self.show()
            return
        geom = parent.frameGeometry()
        self.adjustSize()
        x = geom.width() - self.width() - 24
        y = geom.height() - self.height() - 24
        start = QPoint(x, y + 24)
        end = QPoint(x, y)
        self.move(start)
        self.show()
        self.raise_()
        self._anim.stop()
        self._anim.setStartValue(start)
        self._anim.setEndValue(end)
        self._anim.start()
        cast(Any, QTimer).singleShot(cast(Any, self._duration), cast(Any, self.close))

# -----------------------------
# Helper tipado para mensagens (corrige "partially unknown" do Pylance)
# -----------------------------
class _ShowMessage(Protocol):
    def __call__(
        self,
        parent: Optional[QWidget],
        title: str,
        text: str,
        buttons: Sequence[str] = ("OK",),
        default: int = 0,
        qss: Optional[str] = None,
    ) -> Optional[int]: ...

def show_message(
    parent: Optional[QWidget],
    title: str,
    text: str,
    buttons: Sequence[str] = ("OK",),
    default: int = 0,
    qss: Optional[str] = None,
) -> Optional[int]:
    try:
        from ui.dialogs.custom_messagebox import CustomMessageBox
        # Se qss não foi fornecido, detecta o tema e aplica o estilo apropriado
        if qss is None:
            try:
                from core.config import load_config
                theme = load_config().get("theme", "light")
                if theme == "dark":
                    from core.config import QSS_POPUP_DARK
                    qss = QSS_POPUP_DARK
                else:
                    from core.config import QSS_POPUP_LIGHT
                    qss = QSS_POPUP_LIGHT
            except Exception:
                qss = None
        return cast(_ShowMessage, CustomMessageBox.show_message)(parent, title, text, buttons, default, qss)
    except Exception:
        if len(buttons) == 1:
            QMessageBox.information(parent, title, text)
            return 0
        else:
            r = QMessageBox.question(parent, title, text)
            return 1 if r == QMessageBox.StandardButton.Yes else 0

# -----------------------------
# Database helpers
# -----------------------------
class ExtendedDatabase(Database):
    """Extende core Database com schema do app."""
    def _init_db(self) -> None:
        super()._init_db()
        cur = self.conn.cursor()
        # Clientes
        cur.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                phone TEXT,
                address TEXT,
                birthday TEXT,
                observation TEXT
            )
        """)
        cur.execute("PRAGMA table_info(customers)")
        _cust_cols = [row[1] for row in cur.fetchall()]
        if "observation" not in _cust_cols:
            cur.execute("ALTER TABLE customers ADD COLUMN observation TEXT")
        if "birthday" not in _cust_cols:
            cur.execute("ALTER TABLE customers ADD COLUMN birthday TEXT")
        self.conn.commit()

        # Auditoria / logs de ações (CRUD)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entity TEXT NOT NULL,
                entity_id INTEGER,
                action TEXT NOT NULL,
                user TEXT,
                details TEXT,
                created_at TEXT NOT NULL
            )
        """)
        self.conn.commit()

        # Produtos
        cur.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                size TEXT,
                stock INTEGER NOT NULL DEFAULT 0,
                min_stock INTEGER NOT NULL DEFAULT 0,
                price REAL NOT NULL DEFAULT 0.0
            )
        """)
        
        # Adicionar coluna price se não existir (migração)
        try:
            cur.execute("ALTER TABLE products ADD COLUMN price REAL NOT NULL DEFAULT 0.0")
        except Exception:
            pass  # Coluna já existe

        # Pedidos
        cur.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_number INTEGER UNIQUE,
                customer_id INTEGER NOT NULL,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL,
                delivery_date TEXT NOT NULL,
                total REAL NOT NULL,
                status TEXT NOT NULL,
                label TEXT DEFAULT 'Comum',
                notes TEXT,
                stock_reserved INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id),
                FOREIGN KEY(product_id) REFERENCES products(id)
            )
        """)

        # Migração: adiciona order_number se faltar
        cur.execute("PRAGMA table_info(orders)")
        columns = [row[1] for row in cur.fetchall()]
        if "order_number" not in columns:
            cur.execute("ALTER TABLE orders ADD COLUMN order_number INTEGER")
            self.conn.commit()
        
        # Migração: adiciona stock_reserved se faltar
        cur.execute("PRAGMA table_info(orders)")
        columns = [row[1] for row in cur.fetchall()]
        if "stock_reserved" not in columns:
            # Adiciona coluna e marca pedidos com delivery_date <= hoje como já reservados
            cur.execute("ALTER TABLE orders ADD COLUMN stock_reserved INTEGER NOT NULL DEFAULT 0")
            # Marca pedidos antigos (já entregues ou com data passada) como reservados
            today = datetime.now().strftime("%Y-%m-%d")
            cur.execute("UPDATE orders SET stock_reserved=1 WHERE delivery_date <= ?", (today,))
            self.conn.commit()
            print("✓ Campo stock_reserved adicionado. Pedidos antigos marcados como reservados.")
        
        # Migração: adiciona label (etiqueta) se não existir
        cur.execute("PRAGMA table_info(orders)")
        columns = [row[1] for row in cur.fetchall()]
        if "label" not in columns:
            cur.execute("ALTER TABLE orders ADD COLUMN label TEXT DEFAULT 'Comum'")
            self.conn.commit()
            print("✓ Campo 'label' (etiqueta) adicionado em orders.")
        
        # Migração: adiciona size (tamanho) na tabela products se faltar
        cur.execute("PRAGMA table_info(products)")
        columns = [row[1] for row in cur.fetchall()]
        if "size" not in columns:
            cur.execute("ALTER TABLE products ADD COLUMN size TEXT")
            self.conn.commit()
        
        # Migração: remove requires_min_stock (atribuição) de products se existir
        cur.execute("PRAGMA table_info(products)")
        columns = [row[1] for row in cur.fetchall()]
        if "requires_min_stock" in columns:
            try:
                print("🔄 Removendo campo 'requires_min_stock' (atribuição) da tabela products...")
                cur.execute("DROP TABLE IF EXISTS products_new")
                cur.execute("PRAGMA foreign_keys=OFF")
                cur.execute("BEGIN IMMEDIATE")
                
                # Cria nova tabela sem requires_min_stock
                cur.execute("""
                    CREATE TABLE products_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL UNIQUE,
                        description TEXT,
                        size TEXT,
                        stock INTEGER NOT NULL DEFAULT 0,
                        min_stock INTEGER NOT NULL DEFAULT 0
                    )
                """)
                
                # Copia dados (sem requires_min_stock)
                cur.execute("""
                    INSERT INTO products_new (id, name, description, size, stock, min_stock)
                    SELECT id, name, description, size, stock, min_stock FROM products
                """)
                
                # Remove tabela antiga e renomeia
                cur.execute("DROP TABLE products")
                cur.execute("ALTER TABLE products_new RENAME TO products")
                
                self.conn.commit()
                cur.execute("PRAGMA foreign_keys=ON")
                print("✓ Campo 'requires_min_stock' removido com sucesso!")
                
            except Exception as e:
                self.conn.rollback()
                try:
                    cur.execute("PRAGMA foreign_keys=ON")
                except:
                    pass
                print(f"⚠️ Não foi possível remover campo requires_min_stock: {e}")
        
        # Migração: remove coluna price da tabela products (não controlamos preços)
        cur.execute("PRAGMA table_info(products)")
        columns = [row[1] for row in cur.fetchall()]
        if "price" in columns:
            try:
                # Limpa tabela temporária se existir de tentativa anterior
                cur.execute("DROP TABLE IF EXISTS products_new")
                
                # Desabilita foreign keys temporariamente
                cur.execute("PRAGMA foreign_keys=OFF")
                
                # SQLite não suporta DROP COLUMN diretamente, então recriamos a tabela
                cur.execute("BEGIN IMMEDIATE")
                
                # Cria tabela temporária sem price
                cur.execute("""
                    CREATE TABLE products_new (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT NOT NULL,
                        description TEXT,
                        size TEXT,
                        stock INTEGER NOT NULL DEFAULT 0,
                        min_stock INTEGER NOT NULL DEFAULT 0 INTEGER NOT NULL DEFAULT 0
                    )
                """)
                
                # Copia os dados (sem a coluna price)
                cur.execute("""
                    INSERT INTO products_new (id, name, description, size, stock, min_stock)
                    SELECT id, name, description, size, stock, min_stock, 
                           COALESCE(0) FROM products
                """)
                
                # Remove a tabela antiga e renomeia a nova
                cur.execute("DROP TABLE products")
                cur.execute("ALTER TABLE products_new RENAME TO products")
                
                self.conn.commit()
                
                # Reabilita foreign keys
                cur.execute("PRAGMA foreign_keys=ON")
                
                print("✓ Coluna price removida com sucesso da tabela products")
            except Exception as e:
                self.conn.rollback()
                # Reabilita foreign keys mesmo em caso de erro
                try:
                    cur.execute("PRAGMA foreign_keys=ON")
                except:
                    pass
                # Se falhar, não é crítico - price apenas não será usado
                print(f"Aviso: Não foi possível remover coluna price: {e}")
        
        # Preenche order_number em pedidos antigos
        cur.execute("SELECT id FROM orders WHERE order_number IS NULL OR order_number = '' ORDER BY id")
        rows = cur.fetchall()
        if rows:
            cur.execute("SELECT MAX(order_number) FROM orders")
            maxnum = cur.fetchone()[0] or 0
            for i, row in enumerate(rows, start=1):
                cur.execute("UPDATE orders SET order_number=? WHERE id=?", (maxnum + i, row[0]))
            self.conn.commit()

        # Movimentações de estoque
        cur.execute("""
            CREATE TABLE IF NOT EXISTS stock_movements (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                type TEXT NOT NULL CHECK (type IN ('entrada','saida')),
                quantity INTEGER NOT NULL,
                reason TEXT,
                order_id INTEGER,
                created_at TEXT NOT NULL,
                FOREIGN KEY(product_id) REFERENCES products(id),
                FOREIGN KEY(order_id) REFERENCES orders(id)
            )
        """)
        
        # Etiquetas personalizadas
        cur.execute("""
            CREATE TABLE IF NOT EXISTS labels (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                color TEXT,
                created_at TEXT NOT NULL
            )
        """)
        
        # Tabela de itens da lista de produção (adicionados manualmente)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS production_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_id INTEGER NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 1,
                size TEXT,
                notes TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY (product_id) REFERENCES products (id) ON DELETE CASCADE
            )
        """)
        
        # Tabela de configuração da empresa
        cur.execute("""
            CREATE TABLE IF NOT EXISTS company (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                name TEXT NOT NULL DEFAULT 'Minha Empresa',
                logo_path TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
        """)
        
        # Insere registro padrão se não existir
        existing_company = cur.execute("SELECT 1 FROM company WHERE id=1").fetchone()
        if not existing_company:
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            cur.execute("INSERT INTO company(id, name, logo_path, created_at, updated_at) VALUES (1, 'Confeitaria', NULL, ?, ?)", (now, now))
        
        self.conn.commit()
        
        # Adiciona etiquetas padrão se não existirem
        default_labels = [
            ("Comum", "#6B7280"),
            ("Urgente", "#EF4444"),
            ("Especial", "#F59E0B")
        ]
        for label_name, color in default_labels:
            existing = cur.execute("SELECT 1 FROM labels WHERE name=?", (label_name,)).fetchone()
            if not existing:
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                cur.execute("INSERT INTO labels(name, color, created_at) VALUES (?,?,?)", (label_name, color, now))
        self.conn.commit()

    def audit_log(self, entity: str, entity_id: Optional[int], action: str, details: Optional[str] = None, user: Optional[str] = None) -> None:
        """Registra no banco (e em arquivo logs/actions.log) ações importantes do sistema.

        - entity: 'product'|'customer'|'order' etc
        - entity_id: id do registro (pode ser None)
        - action: 'create'|'update'|'delete'
        - details: texto livre com informações adicionais
        - user: usuário executando a ação; se None, usa self.current_user ou 'unknown'
        """
        try:
            u = user or getattr(self, "current_user", None) or "unknown"
            # ISO para armazenar no banco, humano (dd/MM/aaaa HH:mm:ss) para arquivo texto
            now_iso = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            now_human = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.execute(
                "INSERT INTO audit_logs(entity, entity_id, action, user, details, created_at) VALUES (?,?,?,?,?,?)",
                (entity, entity_id, action, str(u), details, now_iso)
            )
            # também grava em arquivo texto (útil fora do DB)
            try:
                logs_dir = os.path.join(base_dir, "logs")
                os.makedirs(logs_dir, exist_ok=True)
                lf = os.path.join(logs_dir, "actions.log")
                with open(lf, "a", encoding="utf-8") as f:
                    f.write(f"{now_human}\t{u}\t{entity}\t{entity_id}\t{action}\t{details or ''}\n")
            except Exception:
                pass
        except Exception:
            # Não propaga erros de logging para não quebrar a operação principal
            pass

def check_and_reserve_stock(db: Any) -> int:
    """
    Verifica pedidos com data de entrega = hoje que ainda não tiveram estoque reservado.
    Baixa o estoque e marca como reservado.
    
    Returns:
        int: Número de pedidos processados
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        
        # Busca pedidos com data de entrega = hoje e stock_reserved = 0
        rows = db.query("""
            SELECT id, product_id, quantity 
            FROM orders 
            WHERE DATE(delivery_date) = ? 
            AND stock_reserved = 0
            AND product_id IS NOT NULL
        """, (today,))
        
        if not rows:
            return 0
        
        processed = 0
        for row in rows:
            try:
                oid = int(row["id"])
                prod_id = int(row["product_id"])
                qty = int(row["quantity"])
                
                # Baixa estoque
                db.execute("UPDATE products SET stock = stock - ? WHERE id=?", (qty, prod_id))
                
                # Registra movimentação
                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                db.execute(
                    "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                    (prod_id, 'saida', qty, 'Pedido', oid, now)
                )
                
                # Marca como reservado
                db.execute("UPDATE orders SET stock_reserved=1 WHERE id=?", (oid,))
                
                processed += 1
                
            except Exception as e:
                print(f"⚠️ Erro ao reservar estoque do pedido {oid}: {e}")
                continue
        
        print(f"✓ {processed} pedido(s) tiveram estoque reservado automaticamente")
        return processed
        
    except Exception as e:
        print(f"Erro ao verificar reservas de estoque: {e}")
        return 0

def money(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def format_size(size_str: Optional[str]) -> str:
    """Formata tamanho(s) adicionando 'cm' após cada valor.
    
    Exemplos:
        "30" -> "30 cm"
        "15, 20, 25" -> "15 cm, 20 cm, 25 cm"
        "" -> ""
        None -> ""
    """
    if not size_str:
        return ""
    
    # Separa por vírgula, adiciona 'cm' a cada valor e junta novamente
    sizes = [s.strip() for s in size_str.split(",") if s.strip()]
    formatted = [f"{s} cm" if not s.endswith(" cm") and not s.endswith("cm") else s for s in sizes]
    return ", ".join(formatted)

def format_date(date_str: str) -> str:
    """Converte data do formato YYYY-MM-DD para DD/MM/YYYY"""
    if not date_str:
        return ""
    try:
        # Se já estiver no formato DD/MM/YYYY, retorna como está
        if "/" in date_str:
            return date_str
        # Converte de YYYY-MM-DD para DD/MM/YYYY
        d = datetime.strptime(date_str, "%Y-%m-%d")
        return d.strftime("%d/%m/%Y")
    except Exception:
        return date_str

def format_datetime(dt_str: str) -> str:
    """Converte data/hora de ISO (YYYY-MM-DD HH:MM[:SS]) para DD/MM/YYYY HH:MM.

    Mantém como está se já parecer no formato brasileiro.
    """
    if not dt_str:
        return ""
    s = str(dt_str)
    try:
        # Já no padrão brasileiro?
        if "/" in s and (" " in s or len(s) == 10):
            return s
        # Tenta com segundos
        try:
            d = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            # Tenta sem segundos
            try:
                d = datetime.strptime(s, "%Y-%m-%d %H:%M")
            except ValueError:
                # Tenta somente data
                d = datetime.strptime(s, "%Y-%m-%d")
        return d.strftime("%d/%m/%Y %H:%M") if " " in s else d.strftime("%d/%m/%Y")
    except Exception:
        return s

def format_price_br(value: float) -> str:
    """Formata valor monetário no padrão brasileiro.
    
    Exemplos:
        1234.56 -> "R$ 1.234,56"
        1000000.00 -> "R$ 1.000.000,00"
        0.5 -> "R$ 0,50"
    """
    try:
        # Converte para string formatada com 2 casas decimais
        formatted = f"{value:,.2f}"
        # Substitui vírgula por ponto temporário e ponto por vírgula
        formatted = formatted.replace(",", "TEMP").replace(".", ",").replace("TEMP", ".")
        return f"R$ {formatted}"
    except (ValueError, TypeError):
        return "R$ 0,00"

# -----------------------------
# Dialogs (CRUD)
# -----------------------------
class CustomerDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None, data: Optional[sqlite3.Row] = None) -> None:
        super().__init__(parent)
        from core.config import load_config
        
        # Aplica estilo conforme tema atual
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QLineEdit, QComboBox, QSpinBox, QDateEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QTextEdit {
                    background-color: #3c3c3c;
                    border: 2px solid #555555;
                    padding: 8px;
                    border-radius: 4px;
                    color: #ffffff;
                    min-height: 80px;
                }
                QTextEdit:focus {
                    border: 2px solid #0d7377;
                }
                QDateEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QDateEdit::drop-down {
                    border: none;
                    width: 20px;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QLineEdit, QComboBox, QSpinBox, QDateEdit {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QTextEdit {
                    background-color: #ffffff;
                    border: 2px solid #d1d5db;
                    padding: 8px;
                    border-radius: 4px;
                    color: #111827;
                    min-height: 80px;
                }
                QTextEdit:focus {
                    border: 2px solid #0d7377;
                }
                QDateEdit {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QDateEdit::drop-down {
                    border: none;
                    width: 20px;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
            
        self.setWindowTitle("Cliente")
        layout = QFormLayout(self)
        self.name: QLineEdit = QLineEdit()
        self.phone: QLineEdit = QLineEdit()
        self.address: QLineEdit = QLineEdit()
        self.observation: QTextEdit = QTextEdit()
        self.observation.setPlaceholderText("Observações sobre o cliente, preferências, etc.")
        layout.addRow("Nome:", self.name)
        layout.addRow("Telefone:", self.phone)
        layout.addRow("Endereço:", self.address)
        layout.addRow("Observação:", self.observation)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        cast(Any, btns.accepted).connect(self.accept)
        cast(Any, btns.rejected).connect(self.reject)
        layout.addRow(btns)
        if data:
            self.name.setText(data["name"])
            self.phone.setText(data["phone"] or "")
            self.address.setText(data["address"] or "")
            self.observation.setPlainText(data["observation"] or "")

    def get_values(self) -> tuple[str, Optional[str], Optional[str], Optional[str]]:
        return (
            self.name.text().strip(),
            self.phone.text().strip() or None,
            self.address.text().strip() or None,
            self.observation.toPlainText().strip() or None,
        )

class ProductDialog(QDialog):
    def __init__(self, parent: Optional[QWidget] = None, data: Optional[sqlite3.Row] = None) -> None:
        super().__init__(parent)
        from core.config import load_config
        
        # Aplica estilo conforme tema atual
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QLineEdit, QSpinBox, QDoubleSpinBox {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
            
        self.setWindowTitle("Produto")
        layout = QFormLayout(self)
        self.name: QLineEdit = QLineEdit()
        self.description: QLineEdit = QLineEdit()
        
        # Container para os tamanhos dinâmicos
        self.size_inputs: list[QLineEdit] = []
        self.sizes_container = QWidget()
        self.sizes_layout = QVBoxLayout(self.sizes_container)
        self.sizes_layout.setContentsMargins(0, 0, 0, 0)
        self.sizes_layout.setSpacing(5)
        
        # Adiciona primeiro campo de tamanho
        self._add_size_field()
        
        # Botão para adicionar mais tamanhos
        self.btn_add_size = QPushButton("+ Adicionar outro tamanho")
        self.btn_add_size.setObjectName("AddSizeButton")
        cast(Any, self.btn_add_size.clicked).connect(self._add_size_field)
        self.sizes_layout.addWidget(self.btn_add_size)
        
        self.stock: QSpinBox = QSpinBox(); self.stock.setMaximum(1_000_000)
        self.min_stock: QSpinBox = QSpinBox(); self.min_stock.setMaximum(100_000)
        
        # Campo de preço
        self.price: QDoubleSpinBox = QDoubleSpinBox()
        self.price.setMaximum(999999.99)
        self.price.setDecimals(2)
        self.price.setPrefix("R$ ")
        self.price.setSingleStep(1.0)
        
        layout.addRow("Nome:", self.name)
        layout.addRow("Descrição:", self.description)
        layout.addRow("Tamanhos (cm):", self.sizes_container)
        layout.addRow("Preço:", self.price)
        layout.addRow("Estoque:", self.stock)
        layout.addRow("Estoque mín.:", self.min_stock)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        cast(Any, btns.accepted).connect(self.accept)
        cast(Any, btns.rejected).connect(self.reject)
        layout.addRow(btns)
        if data:
            self.name.setText(data["name"])
            self.description.setText(data["description"] or "")
            # Carrega os tamanhos separados por vírgula
            if data["size"]:
                sizes = [s.strip() for s in data["size"].split(",") if s.strip()]
                # Remove o campo inicial vazio e adiciona os tamanhos salvos
                for widget in self.size_inputs[:]:
                    widget.deleteLater()
                    self.size_inputs.remove(widget)
                for size in sizes:
                    self._add_size_field(size)
            # Carregar preço
            try:
                price_value = float(data["price"]) if data["price"] is not None else 0.0
            except (KeyError, TypeError, ValueError):
                price_value = 0.0
            self.price.setValue(price_value)
            self.stock.setValue(int(data["stock"]))
            self.min_stock.setValue(int(data["min_stock"]))
    
    def _add_size_field(self, initial_value: str = "") -> None:
        """Adiciona um novo campo de tamanho à lista."""
        # Cria um container horizontal para o campo + botão de remover
        field_container = QWidget()
        field_layout = QHBoxLayout(field_container)
        field_layout.setContentsMargins(0, 0, 0, 0)
        field_layout.setSpacing(5)
        
        # Cria o campo de texto
        size_field = QLineEdit()
        size_field.setPlaceholderText("Ex: 15 cm" if not self.size_inputs else "Ex: 20 cm")
        if initial_value:
            size_field.setText(initial_value)
        field_layout.addWidget(size_field)
        
        # Adiciona o campo à lista
        self.size_inputs.append(size_field)
        
        # Adiciona botão de remover (exceto para o primeiro campo)
        if len(self.size_inputs) > 1:
            btn_remove = QPushButton("×")
            btn_remove.setFixedWidth(30)
            btn_remove.setStyleSheet("font-size: 18px; font-weight: bold;")
            cast(Any, btn_remove.clicked).connect(lambda: self._remove_size_field(field_container, size_field))
            field_layout.addWidget(btn_remove)
        
        # Insere o container antes do botão "+ Adicionar outro tamanho"
        insert_index = self.sizes_layout.count() - 1  # -1 para ficar antes do botão
        self.sizes_layout.insertWidget(insert_index, field_container)
    
    def _remove_size_field(self, container: QWidget, field: QLineEdit) -> None:
        """Remove um campo de tamanho da lista."""
        if field in self.size_inputs:
            self.size_inputs.remove(field)
        container.deleteLater()

    def get_values(self) -> tuple[str, Optional[str], Optional[str], int, int, float]:
        # Coleta todos os tamanhos não vazios e junta com vírgula
        sizes = [field.text().strip() for field in self.size_inputs if field.text().strip()]
        size_str = ", ".join(sizes) if sizes else None
        
        return (
            self.name.text().strip(),
            self.description.text().strip() or None,
            size_str,
            int(self.stock.value()),
            int(self.min_stock.value()),
            float(self.price.value()),
        )

class MultiProductOrderDialog(QDialog):
    """
    Diálogo para criação de pedido com MÚLTIPLOS produtos.
    Permite adicionar vários produtos para o mesmo cliente em um único pedido.
    """
    def __init__(self, db: Database, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        from core.config import load_config
        
        self.db = db
        self.products_list: list[dict] = []  # Lista de produtos adicionados
        
        self.setWindowTitle("Novo Pedido - Múltiplos Produtos")
        self.resize(800, 600)
        
        # Aplica tema
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog { background-color: #2b2b2b; color: #ffffff; }
                QLabel { color: #ffffff; }
                QLineEdit, QComboBox, QSpinBox, QDateEdit, QTextEdit {
                    background-color: #3c3c3c; border: 1px solid #555555;
                    padding: 5px; border-radius: 3px; color: #ffffff;
                }
                QTableWidget {
                    background-color: #3c3c3c; color: #ffffff;
                    gridline-color: #555555; border: 1px solid #555555;
                }
                QTableWidget::item { padding: 5px; }
                QHeaderView::section {
                    background-color: #2b2b2b; color: #ffffff;
                    padding: 5px; border: 1px solid #555555;
                }
                QPushButton {
                    background: #1a2031; color: #ffffff; padding: 8px 14px;
                    border: 1px solid #2c3550; border-radius: 10px;
                }
                QPushButton:hover { background: #252c45; border: 1px solid #3d4a70; }
            """)
        else:
            self.setStyleSheet("""
                QDialog { background-color: #ffffff; color: #111827; }
                QLabel { color: #111827; }
                QLineEdit, QComboBox, QSpinBox, QDateEdit, QTextEdit {
                    background-color: #ffffff; border: 1px solid #d1d5db;
                    padding: 5px; border-radius: 3px; color: #111827;
                }
                QTableWidget {
                    background-color: #ffffff; color: #111827;
                    gridline-color: #e5e7eb; border: 1px solid #d1d5db;
                }
                QTableWidget::item { padding: 5px; }
                QHeaderView::section {
                    background-color: #f9fafb; color: #111827;
                    padding: 5px; border: 1px solid #e5e7eb;
                }
                QPushButton {
                    background: #e5e7eb; color: #111827; padding: 8px 14px;
                    border: 1px solid #d1d5db; border-radius: 10px;
                }
                QPushButton:hover { background: #dbeafe; border: 1px solid #bfdbfe; }
            """)
        
        main_layout = QVBoxLayout(self)
        
        # === SEÇÃO 1: Informações Gerais do Pedido ===
        info_group = QGroupBox("Informações do Pedido")
        info_layout = QFormLayout()
        
        # Cliente
        self.customer = QComboBox()
        self._load_customers()
        info_layout.addRow("Cliente:", self.customer)
        
        # Data de entrega
        self.delivery = QDateEdit()
        self.delivery.setCalendarPopup(True)
        self.delivery.setDisplayFormat("dd/MM/yyyy")
        self.delivery.setDate(QDate.currentDate())
        
        # Aplica estilo ao calendário
        if theme_cfg == "light":
            self.delivery.setStyleSheet("""
                QDateEdit {
                    background: #ffffff;
                    color: #111827;
                    border: 1px solid #d1d5db;
                    border-radius: 8px;
                    padding: 6px;
                }
                QDateEdit::drop-down {
                    background: #f9fafb;
                    border-left: 1px solid #d1d5db;
                }
            """)
            if (cw := self.delivery.calendarWidget()) is not None:
                cw.setStyleSheet("""
                    QCalendarWidget {
                        background-color: #ffffff;
                        color: #111827;
                        border: 1px solid #e5e7eb;
                        border-radius: 8px;
                    }
                    QCalendarWidget QWidget {
                        background: #ffffff;
                        color: #111827;
                    }
                    QCalendarWidget QAbstractItemView {
                        background: #ffffff;
                        color: #111827;
                        selection-background-color: #e8eefc;
                        selection-color: #1b2240;
                        gridline-color: #e5e7eb;
                    }
                    QCalendarWidget QAbstractItemView::item:selected {
                        background: #e8eefc;
                        color: #1b2240;
                        border-radius: 4px;
                    }
                    QCalendarWidget QAbstractItemView::item:hover {
                        background: #dbeafe;
                        color: #111827;
                    }
                    QCalendarWidget QTableView {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                    }
                    QCalendarWidget QToolButton {
                        background-color: #f3f4f6;
                        background: #f3f4f6;
                        color: #1f2937;
                        border: 1px solid #e5e7eb;
                        border-radius: 6px;
                        padding: 4px 8px;
                    }
                    QCalendarWidget QToolButton:hover {
                        background-color: #e5e7eb;
                        background: #e5e7eb;
                    }
                    QCalendarWidget QSpinBox {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                        border: 1px solid #d1d5db;
                        border-radius: 6px;
                        padding: 2px 6px;
                    }
                    QCalendarWidget QHeaderView::section {
                        background-color: #f9fafb;
                        background: #f9fafb;
                        color: #1f2937;
                        border: 1px solid #e5e7eb;
                        padding: 4px;
                    }
                """)
        else:
            self.delivery.setStyleSheet("""
                QDateEdit {
                    background: #0f1422;
                    color: #ffffff;
                    border: 1px solid #2c3550;
                    border-radius: 8px;
                    padding: 6px;
                }
                QDateEdit::drop-down {
                    background: #1a1f2e;
                    border-left: 1px solid #2c3550;
                }
            """)
            if (cw := self.delivery.calendarWidget()) is not None:
                cw.setStyleSheet("""
                    QCalendarWidget {
                        background-color: #1a1f2e;
                        color: #ffffff;
                        border: 1px solid #2c3550;
                        border-radius: 8px;
                    }
                    QCalendarWidget QWidget {
                        background: #1a1f2e;
                        color: #ffffff;
                    }
                    QCalendarWidget QAbstractItemView {
                        background: #1a1f2e;
                        color: #ffffff;
                        selection-background-color: #2a2f43;
                        selection-color: #ffffff;
                        gridline-color: #2c3550;
                    }
                    QCalendarWidget QAbstractItemView::item:selected {
                        background: #2a2f43;
                        color: #ffffff;
                        border-radius: 4px;
                    }
                    QCalendarWidget QAbstractItemView::item:hover {
                        background: #1e3a5f;
                        color: #ffffff;
                    }
                """)
        
        info_layout.addRow("Data Entrega:", self.delivery)
        
        # Status
        self.status = QComboBox()
        self.status.addItems(["Aguardando Pagamento", "Pagamento na Retirada", "Pago"])
        info_layout.addRow("Status:", self.status)
        
        # Etiqueta
        self.label = QComboBox()
        self._load_labels()
        info_layout.addRow("Etiqueta:", self.label)
        
        # Observações gerais
        self.notes = QTextEdit()
        self.notes.setPlaceholderText("Observações gerais do pedido...")
        self.notes.setMaximumHeight(60)
        info_layout.addRow("Observações:", self.notes)
        
        info_group.setLayout(info_layout)
        main_layout.addWidget(info_group)
        
        # === SEÇÃO 2: Adicionar Produtos ===
        product_group = QGroupBox("Adicionar Produtos")
        product_layout = QHBoxLayout()
        
        product_layout.addWidget(QLabel("Produto:"))
        self.product = QComboBox()
        self._load_products()
        product_layout.addWidget(self.product, 2)
        
        product_layout.addWidget(QLabel("Qtd:"))
        self.quantity = QSpinBox()
        self.quantity.setMinimum(1)
        self.quantity.setMaximum(1000000)
        self.quantity.setValue(1)
        product_layout.addWidget(self.quantity, 1)
        
        self.btn_add_product = QPushButton("➕ Adicionar")
        cast(Any, self.btn_add_product.clicked).connect(self._add_product_to_list)
        product_layout.addWidget(self.btn_add_product)
        
        product_group.setLayout(product_layout)
        main_layout.addWidget(product_group)
        
        # === SEÇÃO 3: Tabela de Produtos Adicionados ===
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Produto", "Tamanho", "Quantidade", "Ação"])
        self.table.horizontalHeader().setStretchLastSection(False)
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        main_layout.addWidget(self.table)
        
        # Informação
        info_label = QLabel("💡 Adicione todos os produtos desejados antes de salvar o pedido")
        info_label.setStyleSheet("color: #6b7280; font-style: italic; padding: 5px;")
        main_layout.addWidget(info_label)
        
        # === BOTÕES FINAIS ===
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        cast(Any, btns.accepted).connect(self._save_order)
        cast(Any, btns.rejected).connect(self.reject)
        main_layout.addWidget(btns)
    
    def _load_customers(self) -> None:
        self.customer.clear()
        rows = self.db.query("SELECT id, name FROM customers ORDER BY name")
        for r in rows:
            self.customer.addItem(r["name"], r["id"])
    
    def _load_products(self) -> None:
        self.product.clear()
        rows = self.db.query("SELECT id, name FROM products ORDER BY name")
        for r in rows:
            self.product.addItem(r["name"], r["id"])
    
    def _load_labels(self) -> None:
        self.label.clear()
        rows = self.db.query("SELECT name FROM labels ORDER BY name")
        for r in rows:
            self.label.addItem(r["name"])
        comum_index = self.label.findText("Comum")
        if comum_index >= 0:
            self.label.setCurrentIndex(comum_index)
    
    def _add_product_to_list(self) -> None:
        """Adiciona o produto selecionado à lista"""
        prod_id = self.product.currentData()
        prod_name = self.product.currentText()
        qty = self.quantity.value()
        
        if not prod_id:
            QMessageBox.warning(self, "Aviso", "Selecione um produto!")
            return
        
        if qty <= 0:
            QMessageBox.warning(self, "Aviso", "Quantidade deve ser maior que 0!")
            return
        
        # Busca informações do produto
        prod_data = self.db.query("SELECT size, stock FROM products WHERE id=?", (prod_id,))
        if not prod_data:
            QMessageBox.warning(self, "Erro", "Produto não encontrado!")
            return
        
        size = prod_data[0]["size"] or ""
        
        # Adiciona à lista (sem validação de estoque - permite estoque negativo)
        self.products_list.append({
            "product_id": prod_id,
            "product_name": prod_name,
            "size": size,
            "quantity": qty
        })
        
        self._refresh_table()
        
        # Reseta quantidade
        self.quantity.setValue(1)
    
    def _refresh_table(self) -> None:
        """Atualiza a tabela com os produtos adicionados"""
        self.table.setRowCount(len(self.products_list))
        
        for row, item in enumerate(self.products_list):
            # Define altura mínima da linha para acomodar o botão
            self.table.setRowHeight(row, 50)
            
            # Produto
            self.table.setItem(row, 0, QTableWidgetItem(item["product_name"]))
            
            # Tamanho
            size_formatted = format_size(item["size"]) if item["size"] else "-"
            self.table.setItem(row, 1, QTableWidgetItem(size_formatted))
            
            # Quantidade
            self.table.setItem(row, 2, QTableWidgetItem(str(item["quantity"])))
            
            # Botão remover - apenas o ícone sem fundo
            btn_remove = QPushButton("🗑️")
            btn_remove.setToolTip("Remover este produto")
            btn_remove.setFixedSize(40, 35)
            btn_remove.setStyleSheet("""
                QPushButton {
                    background-color: transparent !important;
                    background: transparent !important;
                    border: none !important;
                    padding: 6px !important;
                    font-size: 20px !important;
                }
                QPushButton:hover {
                    background-color: rgba(239, 68, 68, 0.1) !important;
                    background: rgba(239, 68, 68, 0.1) !important;
                    border-radius: 6px !important;
                }
                QPushButton:pressed {
                    background-color: rgba(239, 68, 68, 0.2) !important;
                    background: rgba(239, 68, 68, 0.2) !important;
                }
            """)
            btn_remove.setProperty("row", row)
            cast(Any, btn_remove.clicked).connect(lambda checked, r=row: self._remove_product(r))
            
            # Container para centralizar o botão
            container = QWidget()
            container_layout = QHBoxLayout(container)
            container_layout.addWidget(btn_remove)
            container_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            container_layout.setContentsMargins(5, 5, 5, 5)
            self.table.setCellWidget(row, 3, container)
    
    def _remove_product(self, row: int) -> None:
        """Remove um produto da lista"""
        if 0 <= row < len(self.products_list):
            del self.products_list[row]
            self._refresh_table()
    
    def _save_order(self) -> None:
        """Salva o pedido com todos os produtos"""
        if not self.products_list:
            QMessageBox.warning(self, "Aviso", "Adicione pelo menos um produto ao pedido!")
            return
        
        customer_id = self.customer.currentData()
        if not customer_id:
            QMessageBox.warning(self, "Aviso", "Selecione um cliente!")
            return
        
        # Aceita e salva
        self.accept()
    
    def get_order_data(self) -> dict:
        """Retorna os dados do pedido completo"""
        return {
            "customer_id": self.customer.currentData(),
            "delivery_date": self.delivery.date().toPyDate().isoformat(),
            "status": self.status.currentText(),
            "label": self.label.currentText(),
            "notes": self.notes.toPlainText().strip() or None,
            "products": self.products_list
        }

class OrderDialog(QDialog):
    """
    Diálogo de criação/edição de pedidos.
    - Permite adicionar MÚLTIPLOS produtos por cliente
    - Carrega clientes e produtos do banco.
    - Valida quantidade > 0.
    - Valida a data (campo com máscara dd/MM/yyyy).
    - Total sempre é 0.0 (não controlamos preços).
    """
    def __init__(self, db: Database, parent: Optional[QWidget] = None, data: Optional[sqlite3.Row] = None) -> None:
        super().__init__(parent)
        from PyQt6.QtCore import QRegularExpression
        from PyQt6.QtGui import QRegularExpressionValidator
        from core.config import load_config
        
        # Aplica estilo conforme tema atual
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QLineEdit, QComboBox, QSpinBox, QDateEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QTextEdit {
                    background-color: #3c3c3c;
                    border: 2px solid #555555;
                    padding: 8px;
                    border-radius: 4px;
                    color: #ffffff;
                    min-height: 80px;
                }
                QTextEdit:focus {
                    border: 2px solid #0d7377;
                }
                QDateEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QDateEdit::drop-down {
                    border: none;
                    width: 20px;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QLineEdit, QComboBox, QSpinBox, QDateEdit {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QTextEdit {
                    background-color: #ffffff;
                    border: 2px solid #d1d5db;
                    padding: 8px;
                    border-radius: 4px;
                    color: #111827;
                    min-height: 80px;
                }
                QTextEdit:focus {
                    border: 2px solid #0d7377;
                }
                QDateEdit {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QDateEdit::drop-down {
                    border: none;
                    width: 20px;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)

        self.db = db
        self.setWindowTitle("Pedido")

        layout = QFormLayout(self)

        # Cliente
        self.customer: QComboBox = QComboBox()
        self._load_customers()

        # Produto
        self.product: QComboBox = QComboBox()
        
        # Label para mostrar o tamanho do produto selecionado (criar ANTES de carregar produtos)
        self.product_size_label: QLabel = QLabel()
        self.product_size_label.setStyleSheet("color: #6b7280; font-style: italic;")
        
        # Agora carrega os produtos (que vai chamar _on_product_changed)
        self._load_products()
        cast(Any, self.product.currentIndexChanged).connect(self._on_product_changed)

        # Quantidade
        self.quantity: QSpinBox = QSpinBox()
        self.quantity.setMaximum(1_000_000)
        self.quantity.setValue(1)

        # Data de entrega
        self.delivery: QDateEdit = QDateEdit()
        self.delivery.setCalendarPopup(True)
        self.delivery.setDisplayFormat("dd/MM/yyyy")
        self.delivery.setReadOnly(False)
        # Validador dd/MM/yyyy
        date_regex = QRegularExpression(r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$")
        cast(Any, self.delivery.lineEdit()).setValidator(QRegularExpressionValidator(date_regex))
        # Se for novo pedido, já deixa hoje
        self.delivery.setDate(QDate.currentDate())
        # Aplica estilo ao calendário do QDateEdit para corrigir fundo da seleção
        try:
            cal = self.delivery.calendarWidget()
            if cal is not None:
                from core.config import load_config
                theme_cfg = load_config().get("theme", "light")
                if theme_cfg == "dark":
                    cal_style = """
                    QCalendarWidget {
                        background-color: #1a1f2e;
                        background: #1a1f2e;
                        color: #ffffff;
                        border: 1px solid #2c3550;
                        border-radius: 8px;
                    }
                    QCalendarWidget QWidget {
                        background-color: #1a1f2e;
                        background: #1a1f2e;
                        color: #ffffff;
                    }
                    QCalendarWidget QAbstractItemView {
                        background-color: #1a1f2e;
                        background: #1a1f2e;
                        color: #ffffff;
                        selection-background-color: #2a2f43;
                        selection-color: #ffffff;
                        gridline-color: #2c3550;
                    }
                    /* Seleção explícita dos dias (evita fundo preto) */
                    QCalendarWidget QAbstractItemView::item:selected {
                        background-color: #2a2f43;
                        color: #ffffff;
                        border-radius: 4px;
                    }
                    QCalendarWidget QAbstractItemView::item:hover {
                        background-color: #1e3a5f;
                        color: #ffffff;
                    }
                    QCalendarWidget QToolButton {
                        background-color: #252c45;
                        background: #252c45;
                        color: #ffffff;
                        border: 1px solid #2c3550;
                        border-radius: 6px;
                        padding: 6px;
                    }
                    QCalendarWidget QToolButton:hover {
                        background-color: #333b5e;
                        background: #333b5e;
                    }
                    QCalendarWidget QSpinBox {
                        background-color: #1a1f2e;
                        background: #1a1f2e;
                        color: #ffffff;
                        border: 1px solid #2c3550;
                        border-radius: 6px;
                        padding: 2px 6px;
                    }
                    QCalendarWidget QHeaderView::section {
                        background-color: #252c45;
                        background: #252c45;
                        color: #ffffff;
                        border: 1px solid #2c3550;
                        padding: 4px;
                    }
                    """
                else:
                    cal_style = """
                    QCalendarWidget {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                        border: 1px solid #e5e7eb;
                        border-radius: 8px;
                    }
                    QCalendarWidget QWidget {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                    }
                    QCalendarWidget QAbstractItemView {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                        selection-background-color: #e8eefc;
                        selection-color: #1b2240;
                        gridline-color: #e5e7eb;
                    }
                    /* Seleção explícita dos dias (evita fundo preto) */
                    QCalendarWidget QAbstractItemView::item:selected {
                        background-color: #e8eefc;
                        color: #1b2240;
                        border-radius: 4px;
                    }
                    QCalendarWidget QAbstractItemView::item:hover {
                        background-color: #dbeafe;
                        color: #111827;
                    }
                    QCalendarWidget QTableView {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                    }
                    QCalendarWidget QToolButton {
                        background-color: #f3f4f6;
                        background: #f3f4f6;
                        color: #1f2937;
                        border: 1px solid #e5e7eb;
                        border-radius: 6px;
                        padding: 4px 8px;
                    }
                    QCalendarWidget QToolButton:hover {
                        background-color: #e5e7eb;
                        background: #e5e7eb;
                    }
                    QCalendarWidget QSpinBox {
                        background-color: #ffffff;
                        background: #ffffff;
                        color: #111827;
                        border: 1px solid #d1d5db;
                        border-radius: 6px;
                        padding: 2px 6px;
                    }
                    QCalendarWidget QHeaderView::section {
                        background-color: #f9fafb;
                        background: #f9fafb;
                        color: #1f2937;
                        border: 1px solid #e5e7eb;
                        padding: 4px;
                    }
                    """
                cal.setStyleSheet(cal_style)
        except Exception:
            pass

        # Status (financeiro)
        self.status: QComboBox = QComboBox()
        self.status.addItems([
            "Aguardando Pagamento",
            "Pagamento na Retirada",
            "Pago"
        ])

        # Etiqueta
        self.label: QComboBox = QComboBox()
        self._load_labels()

        # Observações
        self.notes: QTextEdit = QTextEdit()
        self.notes.setPlaceholderText("Observações do pedido, sabores, decoração, etc.")

        # Adiciona campos no form
        layout.addRow("Cliente:", self.customer)
        layout.addRow("Produto:", self.product)
        layout.addRow("", self.product_size_label)
        layout.addRow("Quantidade:", self.quantity)
        layout.addRow("Entrega:", self.delivery)
        layout.addRow("Status:", self.status)
        layout.addRow("Etiqueta:", self.label)
        layout.addRow("Obs:", self.notes)

        # Botões
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        cast(Any, btns.accepted).connect(self.accept)
        cast(Any, btns.rejected).connect(self.reject)
        layout.addRow(btns)

        # Se veio com registro para edição, pré-carrega
        if data:
            self._select_combobox_by_data(self.customer, data["customer_id"])
            self._select_combobox_by_data(self.product, data["product_id"])
            self._on_product_changed()  # Atualiza o label de tamanho após selecionar o produto
            self.quantity.setValue(int(data["quantity"]))
            # Data
            try:
                d = datetime.strptime(data["delivery_date"], "%Y-%m-%d").date()
                self.delivery.setDate(QDate(d.year, d.month, d.day))
            except Exception:
                pass
            # Status
            if data["status"]:
                existing = [self.status.itemText(i) for i in range(self.status.count())]
                if data["status"] not in existing:
                    self.status.addItem(data["status"])
                self.status.setCurrentText(data["status"])
            # Observações
            self.notes.setPlainText(data["notes"] or "")

    # Helpers
    def _select_combobox_by_data(self, combo: QComboBox, id_value: int) -> None:
        for i in range(combo.count()):
            if combo.itemData(i) == id_value:
                combo.setCurrentIndex(i)
                return

    def _load_customers(self) -> None:
        self.customer.clear()
        rows = self.db.query("SELECT id, name FROM customers ORDER BY name")
        for r in rows:
            self.customer.addItem(r["name"], r["id"])

    def _load_products(self) -> None:
        self.product.clear()
        rows = self.db.query("SELECT id, name, size FROM products ORDER BY name")
        for r in rows:
            self.product.addItem(r['name'], r["id"])
        # Atualiza o label de tamanho após carregar
        self._on_product_changed()
    
    def _load_labels(self) -> None:
        """Carrega as etiquetas disponíveis no combobox"""
        self.label.clear()
        rows = self.db.query("SELECT name FROM labels ORDER BY name")
        for r in rows:
            self.label.addItem(r["name"])
        # Define "Comum" como padrão se existir
        comum_index = self.label.findText("Comum")
        if comum_index >= 0:
            self.label.setCurrentIndex(comum_index)
    
    def _on_product_changed(self) -> None:
        """Atualiza o label com o tamanho do produto selecionado"""
        prod_id = self.product.currentData()
        if prod_id:
            try:
                rows = self.db.query("SELECT size FROM products WHERE id=?", (prod_id,))
                if rows:
                    size_value = rows[0]["size"]
                    if size_value:
                        formatted_size = format_size(size_value)
                        self.product_size_label.setText(f"📏 Tamanhos disponíveis: {formatted_size}")
                        self.product_size_label.setVisible(True)
                    else:
                        self.product_size_label.setText("")
                        self.product_size_label.setVisible(False)
                else:
                    self.product_size_label.setText("")
                    self.product_size_label.setVisible(False)
            except Exception as e:
                # Se der erro (coluna não existe ainda), esconde o label
                self.product_size_label.setText("")
                self.product_size_label.setVisible(False)
        else:
            self.product_size_label.setText("")
            self.product_size_label.setVisible(False)

    # Saída do diálogo
    def get_values(self) -> tuple[int, int, int, str, float, str, str, Optional[str]]:
        """
        Retorna: (customer_id, product_id, quantity, delivery_iso, total, status, notes)
        - Faz validação de quantidade
        - Total sempre será 0.0 (não controlamos preços)
        - Checa estoque disponível
        """
        cust_id = self.customer.currentData()
        prod_id = self.product.currentData()
        qty = int(self.quantity.value())
        if qty <= 0:
            raise ValueError("Quantidade deve ser maior que 0")

        # Data no formato ISO (yyyy-MM-dd)
        d_iso = self.delivery.date().toPyDate().isoformat()

        status = self.status.currentText()
        label = self.label.currentText()
        notes = self.notes.toPlainText().strip() or None
        
        # Produtos podem ficar com estoque negativo
        # Os produtos com estoque negativo ou zero aparecerão automaticamente na lista de produção

        total = 0.0  # Não controlamos preços
        return cust_id, prod_id, qty, d_iso, total, status, label, notes


class LogsDialog(QDialog):
    """Dialog simples para visualizar audit_logs por entidade/id."""
    def __init__(self, db: Database, entity: str, entity_id: Optional[int], parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        # Aplica estilo escuro apenas se o tema ativo for escuro
        try:
            from core.config import apply_dark_popup_style
            apply_dark_popup_style(self)
        except Exception:
            pass
        self.setWindowTitle(f"Logs - {entity} #{entity_id}")
        self.resize(700, 400)
        self.db = db
        v = QVBoxLayout(self)
        self.tbl = QTableWidget(0, 4)
        self.tbl.setHorizontalHeaderLabels(["Data", "Usuário", "Ação", "Detalhes"])
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        if h := self.tbl.horizontalHeader():
            h.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            h.setStretchLastSection(True)
        v.addWidget(self.tbl)
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        cast(Any, btns.rejected).connect(self.reject)
        cast(Any, btns.accepted).connect(self.accept)
        v.addWidget(btns)

        rows = self.db.query("SELECT created_at, user, action, details FROM audit_logs WHERE entity=? AND entity_id=? ORDER BY created_at DESC", (entity, entity_id))
        if not rows:
            self.tbl.setRowCount(0)
            # mostra mensagem simples
            self.tbl.insertRow(0)
            self.tbl.setItem(0, 0, QTableWidgetItem("—"))
            self.tbl.setItem(0, 1, QTableWidgetItem("—"))
            self.tbl.setItem(0, 2, QTableWidgetItem("—"))
            self.tbl.setItem(0, 3, QTableWidgetItem("Nenhum log encontrado."))
        else:
            for r in rows:
                row = self.tbl.rowCount(); self.tbl.insertRow(row)
                # Formata data/hora no padrão brasileiro (DD/MM/AAAA HH:MM)
                self.tbl.setItem(row, 0, QTableWidgetItem(format_datetime(str(r["created_at"]))))
                self.tbl.setItem(row, 1, QTableWidgetItem(str(r["user"] or "")))
                self.tbl.setItem(row, 2, QTableWidgetItem(str(r["action"])))
                details_raw = str(r["details"] or "")
                # Se for update, esperamos um resumo com '; ' separando mudanças — mostra em linhas separadas
                if str(r["action"]).lower() == "update" and details_raw:
                    details_text = details_raw.replace("; ", "\n")
                else:
                    details_text = details_raw
                item = QTableWidgetItem(details_text)
                item.setToolTip(details_raw)
                self.tbl.setItem(row, 3, item)
        self.tbl.resizeColumnsToContents()


class UserDialog(QDialog):
    """Diálogo para criar/editar usuário (username + password + role)."""
    def __init__(self, parent: Optional[QWidget] = None, data: Optional[sqlite3.Row] = None) -> None:
        super().__init__(parent)
        
        from core.config import load_config
        
        # Aplica estilo conforme tema
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QLineEdit, QComboBox {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QComboBox::drop-down {
                    border: none;
                    background-color: #555555;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 6px solid #ffffff;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QLineEdit, QComboBox {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QComboBox::drop-down {
                    border: none;
                    background-color: #e5e7eb;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 4px solid transparent;
                    border-right: 4px solid transparent;
                    border-top: 6px solid #111827;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
            
        self.setWindowTitle("Usuário")
        self.resize(360, 280)
        layout = QFormLayout(self)
        
        # Flag para indicar se é edição ou criação
        self._editing = data is not None
        
        self.username: QLineEdit = QLineEdit()
        self.password: QLineEdit = QLineEdit()
        self.password.setEchoMode(QLineEdit.EchoMode.Password)
        self.password_confirm: QLineEdit = QLineEdit()
        self.password_confirm.setEchoMode(QLineEdit.EchoMode.Password)
        self.role: QComboBox = QComboBox()
        self.role.addItems(["admin", "common"])
        
        layout.addRow("Nome de usuário:", self.username)
        layout.addRow("Senha:", self.password)
        layout.addRow("Confirmar senha:", self.password_confirm)
        layout.addRow("Tipo:", self.role)
        
        # Password placeholder text
        if data:
            self.password.setPlaceholderText("Deixe em branco para manter a senha atual")
            self.password_confirm.setPlaceholderText("Confirme a nova senha se alterando")
        else:
            self.password.setPlaceholderText("Digite uma senha (mínimo 4 caracteres)")
            self.password_confirm.setPlaceholderText("Confirme a senha")
        
        # Info label
        info_label = QLabel("⚠️ A senha deve ter pelo menos 4 caracteres")
        info_label.setStyleSheet("color: #666; font-size: 11px;")
        layout.addRow(info_label)
        
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel)
        cast(Any, btns.accepted).connect(self.validate_and_accept)
        cast(Any, btns.rejected).connect(self.reject)
        layout.addRow(btns)
        
        if data:
            self.username.setText(data["username"] or "")
            try:
                self.role.setCurrentText(data["role"] or "common")
            except Exception:
                pass

    def validate_and_accept(self) -> None:
        """Valida os dados antes de aceitar o diálogo"""
        username = self.username.text().strip()
        password = self.password.text().strip()
        password_confirm = self.password_confirm.text().strip()
        
        # Validações
        if not username:
            self.show_error("Nome de usuário é obrigatório")
            return
        
        if len(username) < 3:
            self.show_error("Nome de usuário deve ter pelo menos 3 caracteres")
            return
        
        # Se é novo usuário, senha é obrigatória
        if not self._editing and not password:
            self.show_error("Senha é obrigatória para novos usuários")
            return
        
        # Se senha foi fornecida, precisa validar
        if password:
            if len(password) < 4:
                self.show_error("A senha deve ter pelo menos 4 caracteres")
                return
            
            if password != password_confirm:
                self.show_error("As senhas não coincidem")
                return
        
        # Se chegou até aqui, está tudo OK
        self.accept()
    
    def show_error(self, message: str) -> None:
        """Mostra uma mensagem de erro"""
        from PyQt6.QtWidgets import QMessageBox
        QMessageBox.warning(self, "Erro", message)

    def get_values(self) -> tuple[str, str, str]:
        """Returns (username, password, role). Password may be empty if unchanged."""
        return (
            self.username.text().strip(),
            self.password.text().strip(),
            self.role.currentText()
        )


class UsersDialog(QDialog):
    """Diálogo simples para listar/criar/editar/excluir usuários.

    Exibirá apenas para administradores (visibilidade controlada externamente).
    """
    def __init__(self, db: Database, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        
        from core.config import load_config
        
        # Aplica estilo conforme tema
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QTableWidget {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    color: #ffffff;
                }
                QTableWidget::item:selected {
                    background-color: #1a2031;
                    color: #ffffff;
                }
                QTableWidget::item:hover {
                    background-color: #1e3a5f;
                }
                QHeaderView::section {
                    background-color: #2b2b2b;
                    color: #ffffff;
                    border: 1px solid #555555;
                    padding: 4px;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QTableWidget {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    color: #111827;
                }
                QTableWidget::item:selected {
                    background-color: #dbeafe;
                    color: #111827;
                }
                QTableWidget::item:hover {
                    background-color: #dbeafe;
                }
                QHeaderView::section {
                    background-color: #f9fafb;
                    color: #111827;
                    border: 1px solid #d1d5db;
                    padding: 4px;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
            
        self.db = db
        self.setWindowTitle("Usuários do Sistema")
        self.resize(480, 320)
        v = QVBoxLayout(self)
        self.tbl = QTableWidget(0, 2)
        self.tbl.setHorizontalHeaderLabels(["Nome de usuário", "Tipo"])
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        if h := self.tbl.horizontalHeader():
            h.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            h.setStretchLastSection(True)
        v.addWidget(self.tbl)

        h = QHBoxLayout()
        self.btn_add = QPushButton("Adicionar")
        self.btn_edit = QPushButton("Editar")
        self.btn_del = QPushButton("Excluir")
        h.addWidget(self.btn_add); h.addWidget(self.btn_edit); h.addWidget(self.btn_del); h.addStretch(1)
        v.addLayout(h)

        cast(Any, self.btn_add.clicked).connect(self.add)
        cast(Any, self.btn_edit.clicked).connect(self.edit)
        cast(Any, self.btn_del.clicked).connect(self.delete)

        self.refresh()

    def refresh(self) -> None:
        rows = self.db.query("SELECT id, username, role FROM users ORDER BY username")
        self.tbl.setRowCount(0)
        for r in rows:
            row = self.tbl.rowCount(); self.tbl.insertRow(row)
            self.tbl.setItem(row, 0, QTableWidgetItem(str(r["username"])))
            self.tbl.setItem(row, 1, QTableWidgetItem(str(r["role"] or "common")))
            self.tbl.setVerticalHeaderItem(row, QTableWidgetItem(str(r["id"])))
        self.tbl.resizeColumnsToContents()

    def current_id(self) -> Optional[int]:
        idx = self.tbl.currentRow()
        if idx < 0:
            return None
        vh = self.tbl.verticalHeaderItem(idx)
        return int(vh.text()) if vh else None

    def add(self) -> None:
        d = UserDialog(self)
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                username, password, role = d.get_values()
                if not username:
                    raise ValueError("Nome de usuário é obrigatório")
                if not password:
                    raise ValueError("Senha é obrigatória para novos usuários")
                
                # Debug: verificar se a senha não está vazia
                print(f"Debug: Criando usuário '{username}' com senha de {len(password)} caracteres")
                
                password_hash = hash_password(password)
                if not password_hash:
                    raise ValueError("Erro ao gerar hash da senha. Verifique se a senha foi inserida corretamente.")
                
                print(f"Debug: Hash gerado: {password_hash[:50]}...")  # Log parcial do hash para debug
                
                self.db.execute("INSERT INTO users(username, password_hash, role) VALUES (?,?,?)", (username, password_hash, role))
                print(f"Debug: Usuário '{username}' inserido no banco com sucesso")
                self.refresh()
            except Exception as e:
                print(f"Debug: Erro ao adicionar usuário: {e}")
                show_message(self, "Erro", str(e), ("OK",))

    def edit(self) -> None:
        uid = self.current_id()
        if not uid:
            return
        row = self.db.query("SELECT * FROM users WHERE id=?", (uid,))
        if not row:
            return
        d = UserDialog(self, row[0])  # row[0] já faz com que _editing seja True
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                username, password, role = d.get_values()
                if not username:
                    raise ValueError("Nome de usuário é obrigatório")
                    
                # Debug: verificar se a senha foi alterada
                if password:
                    print(f"Debug: Atualizando senha para usuário '{username}' (nova senha tem {len(password)} caracteres)")
                    password_hash = hash_password(password)
                    if not password_hash:
                        raise ValueError("Erro ao gerar hash da nova senha. Verifique se a senha foi inserida corretamente.")
                    print(f"Debug: Novo hash gerado: {password_hash[:50]}...")
                    self.db.execute("UPDATE users SET username=?, password_hash=?, role=? WHERE id=?", 
                                  (username, password_hash, role, uid))
                    
                    # Verificação extra: confirma que foi salvo
                    check = self.db.query("SELECT password_hash FROM users WHERE id=?", (uid,))
                    if check and check[0]['password_hash'] == password_hash:
                        print(f"Debug: ✅ Senha atualizada e verificada no banco")
                    else:
                        print(f"Debug: ❌ ERRO: Senha não foi persistida corretamente!")
                        raise ValueError("Erro: a senha não foi salva corretamente no banco de dados")
                else:
                    print(f"Debug: Atualizando usuário '{username}' sem alterar a senha")
                    self.db.execute("UPDATE users SET username=?, role=? WHERE id=?", 
                                  (username, role, uid))
                print(f"Debug: Usuário '{username}' atualizado com sucesso")
                self.refresh()
            except Exception as e:
                print(f"Debug: Erro ao editar usuário: {e}")
                show_message(self, "Erro", str(e), ("OK",))

    def delete(self) -> None:
        uid = self.current_id()
        if not uid:
            return
        res = show_message(self, "Confirmação", "Excluir este usuário?", ("Não", "Sim"), default=1)
        if res == 1:
            try:
                self.db.execute("DELETE FROM users WHERE id=?", (uid,))
                self.refresh()
            except Exception as e:
                show_message(self, "Erro", str(e), ("OK",))


class LabelsDialog(QDialog):
    """Diálogo para gerenciar etiquetas de pedidos."""
    def __init__(self, db: Database, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        from core.config import load_config
        
        # Aplica estilo conforme tema atual
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QTableWidget {
                    background: #0f1422;
                    alternate-background-color: #0b1020;
                    color: #ffffff;
                    gridline-color: #232323;
                    border: 1px solid #2c3550;
                    border-radius: 4px;
                }
                QTableWidget::item {
                    padding: 8px;
                    color: #ffffff;
                }
                QTableWidget::item:selected {
                    background: #2a2f43;
                    color: #ffffff;
                }
                QHeaderView::section {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 6px;
                    border: none;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QTableWidget {
                    background: #ffffff;
                    alternate-background-color: #f8fafc;
                    color: #111827;
                    gridline-color: #d1d5db;
                    border: 1px solid #e5e7eb;
                    border-radius: 4px;
                }
                QTableWidget::item {
                    padding: 8px;
                    color: #111827;
                }
                QTableWidget::item:selected {
                    background: #e8eefc;
                    color: #1b2240;
                }
                QHeaderView::section {
                    background: #f3f4f6;
                    color: #1f2937;
                    padding: 6px;
                    border: none;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
            
        self.db = db
        self.setWindowTitle("Gerenciar Etiquetas")
        self.resize(500, 400)
        
        v = QVBoxLayout(self)
        
        # Tabela de etiquetas
        self.tbl = QTableWidget(0, 2)
        self.tbl.setHorizontalHeaderLabels(["Nome", "Cor"])
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        if h := self.tbl.horizontalHeader():
            h.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            h.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        v.addWidget(self.tbl)
        
        # Botões de ação
        h_layout = QHBoxLayout()
        self.btn_add = QPushButton("+ Adicionar")
        self.btn_edit = QPushButton("Editar")
        self.btn_del = QPushButton("Excluir")
        h_layout.addWidget(self.btn_add)
        h_layout.addWidget(self.btn_edit)
        h_layout.addWidget(self.btn_del)
        h_layout.addStretch(1)
        v.addLayout(h_layout)
        
        # Botão fechar
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        cast(Any, btns.rejected).connect(self.reject)
        v.addWidget(btns)
        
        cast(Any, self.btn_add.clicked).connect(self.add)
        cast(Any, self.btn_edit.clicked).connect(self.edit)
        cast(Any, self.btn_del.clicked).connect(self.delete)
        
        self.refresh()
    
    def refresh(self) -> None:
        rows = self.db.query("SELECT id, name, color FROM labels ORDER BY name")
        self.tbl.setRowCount(0)
        for r in rows:
            row_idx = self.tbl.rowCount()
            self.tbl.insertRow(row_idx)
            self.tbl.setItem(row_idx, 0, QTableWidgetItem(str(r["name"])))
            
            # Mostra cor com um widget visual
            color_text = str(r["color"] or "#6B7280")
            color_item = QTableWidgetItem(color_text)
            color_item.setBackground(QColor(color_text))
            self.tbl.setItem(row_idx, 1, color_item)
            
            self.tbl.setVerticalHeaderItem(row_idx, QTableWidgetItem(str(r["id"])))
        self.tbl.resizeColumnsToContents()
    
    def current_id(self) -> Optional[int]:
        idx = self.tbl.currentRow()
        if idx < 0:
            return None
        vh = self.tbl.verticalHeaderItem(idx)
        return int(vh.text()) if vh else None
    
    def add(self) -> None:
        from core.config import load_config
        
        # Criar QInputDialog com estilo correto
        input_dialog = QInputDialog(self)
        input_dialog.setWindowTitle("Nova Etiqueta")
        input_dialog.setLabelText("Nome da etiqueta:")
        input_dialog.setTextValue("")
        
        # Aplicar estilo conforme tema
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            input_dialog.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QLineEdit {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            input_dialog.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QLineEdit {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
        
        ok = input_dialog.exec()
        name = input_dialog.textValue() if ok else ""
        
        if ok and name.strip():
            name = name.strip()
            # Verifica duplicata
            existing = self.db.query("SELECT 1 FROM labels WHERE name=?", (name,))
            if existing:
                show_message(self, "Erro", "Já existe uma etiqueta com este nome.", ("OK",))
                return
            
            # Permite escolher cor usando método estático
            initial_color = QColor("#6B7280")
            color = QColorDialog.getColor(initial_color, self, "Escolher cor da etiqueta")
            
            if color.isValid():
                color_hex = color.name()
            else:
                # Se cancelar, usa cor padrão
                color_hex = "#6B7280"
            
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            try:
                self.db.execute("INSERT INTO labels(name, color, created_at) VALUES (?,?,?)", (name, color_hex, now))
                self.refresh()
            except Exception as e:
                show_message(self, "Erro", f"Erro ao adicionar etiqueta: {e}", ("OK",))
    
    def edit(self) -> None:
        lid = self.current_id()
        if not lid:
            show_message(self, "Aviso", "Selecione uma etiqueta.", ("OK",))
            return
        
        row = self.db.query("SELECT * FROM labels WHERE id=?", (lid,))
        if not row:
            return
        
        old_name = str(row[0]["name"])
        old_color = str(row[0]["color"] or "#6B7280")
        
        # Pergunta o que deseja editar
        # Retorna índice: 0=Nome, 1=Cor, 2=Ambos, 3=Cancelar
        choice = show_message(
            self, 
            "Editar Etiqueta", 
            f"Etiqueta: {old_name}\n\nO que deseja editar?", 
            ("Nome", "Cor", "Ambos", "Cancelar")
        )
        
        if choice == 3 or choice is None:  # Cancelar
            return
        
        new_name = old_name
        new_color = old_color
        
        # Edita o nome se solicitado (choice 0=Nome ou 2=Ambos)
        if choice in (0, 2):
            from core.config import load_config
            
            # Criar QInputDialog com estilo correto
            input_dialog = QInputDialog(self)
            input_dialog.setWindowTitle("Editar Nome")
            input_dialog.setLabelText("Nome da etiqueta:")
            input_dialog.setTextValue(old_name)
            
            # Aplicar estilo conforme tema
            theme_cfg = load_config().get("theme", "light")
            if theme_cfg == "dark":
                input_dialog.setStyleSheet("""
                    QDialog {
                        background-color: #2b2b2b;
                        color: #ffffff;
                    }
                    QLabel {
                        color: #ffffff;
                    }
                    QLineEdit {
                        background-color: #3c3c3c;
                        border: 1px solid #555555;
                        padding: 5px;
                        border-radius: 3px;
                        color: #ffffff;
                    }
                    QPushButton {
                        background: #1a2031;
                        color: #ffffff;
                        padding: 8px 14px;
                        border: 1px solid #2c3550 !important;
                        border-radius: 10px;
                    }
                    QPushButton:hover {
                        background: #252c45;
                        border: 1px solid #3d4a70 !important;
                    }
                    QPushButton:pressed {
                        background: #333b5e;
                        border: 1px solid #4a5480 !important;
                    }
                """)
            else:
                input_dialog.setStyleSheet("""
                    QDialog {
                        background-color: #ffffff;
                        color: #111827;
                    }
                    QLabel {
                        color: #111827;
                    }
                    QLineEdit {
                        background-color: #ffffff;
                        border: 1px solid #d1d5db;
                        padding: 5px;
                        border-radius: 3px;
                        color: #111827;
                    }
                    QPushButton {
                        background: #e5e7eb;
                        color: #111827;
                        padding: 8px 14px;
                        border: 1px solid #d1d5db !important;
                        border-radius: 10px;
                    }
                    QPushButton:hover {
                        background: #dbeafe;
                        border: 1px solid #bfdbfe !important;
                    }
                    QPushButton:pressed {
                        background: #c7d2fe;
                        border: 1px solid #a5b4fc !important;
                    }
                """)
            
            ok = input_dialog.exec()
            name_input = input_dialog.textValue() if ok else ""
            
            if not ok:
                return
            
            new_name = name_input.strip()
            if not new_name:
                show_message(self, "Erro", "O nome não pode estar vazio.", ("OK",))
                return
            
            if new_name != old_name:
                # Verifica duplicata
                existing = self.db.query("SELECT 1 FROM labels WHERE name=? AND id!=?", (new_name, lid))
                if existing:
                    show_message(self, "Erro", "Já existe uma etiqueta com este nome.", ("OK",))
                    return
        
        # Edita a cor se solicitado (choice 1=Cor ou 2=Ambos)
        if choice in (1, 2):
            initial_color = QColor(old_color)
            color = QColorDialog.getColor(initial_color, self, "Escolher cor da etiqueta")
            
            if color.isValid():
                new_color = color.name()
            else:
                # Se cancelar o seletor de cor, mantém a cor antiga
                new_color = old_color
        
        try:
            self.db.execute("UPDATE labels SET name=?, color=? WHERE id=?", (new_name, new_color, lid))
            # Atualiza todos os pedidos que usam esta etiqueta
            if new_name != old_name:
                self.db.execute("UPDATE orders SET label=? WHERE label=?", (new_name, old_name))
            self.refresh()
        except Exception as e:
            show_message(self, "Erro", f"Erro ao editar etiqueta: {e}", ("OK",))
    
    def delete(self) -> None:
        lid = self.current_id()
        if not lid:
            show_message(self, "Aviso", "Selecione uma etiqueta.", ("OK",))
            return
        
        row = self.db.query("SELECT name FROM labels WHERE id=?", (lid,))
        if not row:
            return
        
        label_name = str(row[0]["name"])
        
        # Verifica se há pedidos usando esta etiqueta
        orders_count = self.db.query("SELECT COUNT(*) as c FROM orders WHERE label=?", (label_name,))
        count = int(orders_count[0]["c"]) if orders_count else 0
        
        msg = f"Excluir a etiqueta '{label_name}'?"
        if count > 0:
            msg += f"\n\n{count} pedido(s) usam esta etiqueta e serão alterados para 'Comum'."
        
        res = show_message(self, "Confirmação", msg, ("Não", "Sim"), default=1)
        if res == 1:
            try:
                # Atualiza pedidos para 'Comum'
                if count > 0:
                    self.db.execute("UPDATE orders SET label='Comum' WHERE label=?", (label_name,))
                # Exclui a etiqueta
                self.db.execute("DELETE FROM labels WHERE id=?", (lid,))
                self.refresh()
            except Exception as e:
                show_message(self, "Erro", f"Erro ao excluir: {e}", ("OK",))


# -----------------------------
# Pages
# -----------------------------
class BasePage(QWidget):
    def __init__(self, title: str, subtitle: str = "") -> None:
        super().__init__()
        self.v = QVBoxLayout(self)
        head = QWidget()
        hl = QHBoxLayout(head)
        t = QLabel(f"<h2 style='margin:0'>{title}</h2>")
        s = QLabel(subtitle)
        s.setObjectName("subtitle")
        hl.addWidget(t)
        hl.addStretch(1)
        hl.addWidget(s)
        self.v.addWidget(head)
        self.body = QWidget()
        self.v.addWidget(self.body)
        self.v.setContentsMargins(16, 16, 16, 16)

class ProductsPage(BasePage):
    def __init__(self, db: DB, toast_cb: Optional[Callable[[str], None]] = None) -> None:
        super().__init__("Produtos", "Gerencie o catálogo")
        self.toast_cb = toast_cb
        self.db = db
        bl = QVBoxLayout(self.body)
        
        actions = QHBoxLayout()
        self.btn_add: QPushButton = QPushButton("+ Novo")
        self.btn_del: QPushButton = QPushButton("Excluir")
        self.btn_edit: QPushButton = QPushButton("Editar")
        actions.addWidget(self.btn_add); actions.addWidget(self.btn_edit); actions.addWidget(self.btn_del); actions.addStretch(1)
        bl.addLayout(actions)
        # Barra de pesquisa (nome/descrição)
        search_box = QHBoxLayout()
        search_box.addWidget(QLabel("Pesquisar:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Nome ou descrição…")
        self.search_edit.setClearButtonEnabled(True)
        cast(Any, self.search_edit.textChanged).connect(lambda _t: self.refresh())
        btn_clear = QPushButton("Limpar")
        cast(Any, btn_clear.clicked).connect(lambda: self.search_edit.clear())
        search_box.addWidget(self.search_edit, 1)
        search_box.addWidget(btn_clear)
        bl.addLayout(search_box)
        # adiciona coluna extra para o botão de visualização de logs
        self.table: QTableWidget = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Nome", "Descrição", "Tamanho", "Preço", "Estoque", "Logs"])
        self.table.setAlternatingRowColors(True)
        header = self.table.horizontalHeader()
        if header:
            header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            # não esticar a última coluna; manter Logs compacto
            header.setStretchLastSection(False)
            # fixa largura da coluna Logs para manter alinhamento consistente
            try:
                logs_col = 5
                header.setSectionResizeMode(logs_col, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(logs_col, 96)
                # Centraliza o título
                item = self.table.horizontalHeaderItem(logs_col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            except Exception:
                pass
        # Configura header vertical para altura automática das linhas
        if vh := self.table.verticalHeader():
            vh.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            vh.setMinimumSectionSize(36)  # Altura mínima para acomodar ícones de 28px
        bl.addWidget(self.table)
        # se usuário não for admin, esconde a coluna de Logs
        try:
            if getattr(self.db, "current_role", "admin") != "admin":
                self.table.setColumnHidden(4, True)
        except Exception:
            pass
        cast(Any, self.btn_add.clicked).connect(self.add)
        cast(Any, self.btn_edit.clicked).connect(self.edit)
        cast(Any, self.btn_del.clicked).connect(self.delete)
        self.refresh()
        self._warn_low_stock()
        
        # Timer para atualização automática a cada 5 segundos
        self.auto_refresh_timer = QTimer(self)
        self.auto_refresh_timer.timeout.connect(self.refresh)
        self.auto_refresh_timer.start(5000)  # 5000ms = 5 segundos

    def _warn_low_stock(self) -> None:
        # Avisa sobre produtos com estoque baixo
        rows = self.db.query("SELECT name, stock, min_stock FROM products WHERE stock <= min_stock AND min_stock > 0 ORDER BY name")
        if rows and self.toast_cb:
            names = ", ".join([f"{r['name']}({r['stock']})" for r in rows])
            self.toast_cb(f"⚠️ Estoque baixo: {names}")

    def refresh(self) -> None:
        # Filtro por texto (nome/descrição)
        try:
            term = self.search_edit.text().strip()
        except Exception:
            term = ""
        if term:
            like = f"%{term}%"
            rows = self.db.query(
                "SELECT * FROM products WHERE name LIKE ? OR description LIKE ? ORDER BY name",
                (like, like)
            )
        else:
            rows = self.db.query("SELECT * FROM products ORDER BY name")
        self.table.setRowCount(0)
        
        for r in rows:
            row = self.table.rowCount(); self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(r["name"]))
            self.table.setItem(row, 1, QTableWidgetItem(r["description"] or ""))
            self.table.setItem(row, 2, QTableWidgetItem(format_size(r["size"])))
            # Preço formatado no padrão brasileiro
            try:
                price_value = float(r["price"]) if r["price"] is not None else 0.0
            except (KeyError, TypeError, ValueError):
                price_value = 0.0
            price_item = QTableWidgetItem(format_price_br(price_value))
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.table.setItem(row, 3, price_item)
            self.table.setItem(row, 4, QTableWidgetItem(str(r["stock"])))
            self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(r["id"])))
            
            # botão de logs (ícone) com container centralizado - coluna 5
            btn = QPushButton(); btn.setObjectName("IconButton")
            btn.setToolTip("Ver logs deste produto")
            btn.setFlat(True)
            try:
                btn.setIcon(safe_qta_icon("ph.magnifying-glass", color="#9aa3b2"))
            except Exception:
                pass
            btn.setIconSize(QSize(18, 18))
            btn.setFixedSize(28, 28)
            ent_id = int(r["id"]) if r["id"] is not None else None
            cast(Any, btn.clicked).connect(lambda _checked=False, eid=ent_id: self._show_logs("product", eid))
            # Container para centralizar o botão
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(btn)
            layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 5, container)
            # Força altura mínima de 48px para garantir que ícones não sejam cortados
            self.table.setRowHeight(row, 48)
        self.table.resizeColumnsToContents()
        # Reaplica largura fixa da coluna Logs após autoajuste
        try:
            self.table.setColumnWidth(5, 96)
        except Exception:
            pass

    def current_id(self) -> Optional[int]:
        idx = self.table.currentRow()
        if idx < 0:
            return None
        vh = self.table.verticalHeaderItem(idx)
        return int(vh.text()) if vh else None
    
    def _show_logs(self, entity: str, entity_id: Optional[int]) -> None:
        dlg = LogsDialog(self.db, entity, entity_id, self)
        dlg.exec()

    def add(self) -> None:
        d = ProductDialog(self)
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                name, desc, size, stock, min_stock, price = d.get_values()
                if not name:
                    raise ValueError("Nome é obrigatório.")
                
                # Verifica se tem múltiplos tamanhos separados por vírgula
                sizes = [s.strip() for s in size.split(",") if s.strip()] if size else []
                
                if len(sizes) > 1:
                    # Criar um produto para cada tamanho
                    created_count = 0
                    for sz in sizes:
                        product_name = f"{name} {sz}cm"
                        cur = self.db.execute(
                            "INSERT INTO products(name, description, size, stock, min_stock, price) VALUES (?,?,?,?,?,?)",
                            (product_name, desc, sz, stock, min_stock, price)
                        )
                        pid_new = int(cur.lastrowid) if cur and cur.lastrowid is not None else None
                        if pid_new and stock and int(stock) > 0:
                            self.db.execute(
                                "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                                (pid_new, 'entrada', int(stock), 'Cadastro inicial', None, datetime.now().strftime("%Y-%m-%d %H:%M"))
                            )
                        try:
                            self.db.audit_log("product", pid_new, "create", details=f"name={product_name},stock={stock},price={price}")
                        except Exception:
                            pass
                        created_count += 1
                    
                    self.refresh()
                    if self.toast_cb:
                        self.toast_cb(f"{created_count} produtos criados com sucesso")
                    self._warn_low_stock()
                else:
                    # Criar apenas um produto (comportamento normal)
                    cur = self.db.execute(
                        "INSERT INTO products(name, description, size, stock, min_stock, price) VALUES (?,?,?,?,?,?)",
                        (name, desc, size, stock, min_stock, price)
                    )
                    pid_new = int(cur.lastrowid) if cur and cur.lastrowid is not None else None
                    if pid_new and stock and int(stock) > 0:
                        self.db.execute(
                            "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                            (pid_new, 'entrada', int(stock), 'Cadastro inicial', None, datetime.now().strftime("%Y-%m-%d %H:%M"))
                        )
                    self.refresh()
                    if self.toast_cb:
                        self.toast_cb("Produto criado com sucesso")
                    self._warn_low_stock()
                    try:
                        self.db.audit_log("product", pid_new, "create", details=f"name={name},stock={stock}")
                    except Exception:
                        pass
            except Exception as e:
                show_message(self, "Erro ao salvar", str(e), ("OK",))

    def edit(self) -> None:
        pid = self.current_id()
        if not pid:
            return
        row = self.db.query("SELECT * FROM products WHERE id=?", (pid,))
        if not row:
            return
        d = ProductDialog(self, row[0])
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                name, desc, size, stock, min_stock, price = d.get_values()
                if not name:
                    raise ValueError("Nome é obrigatório.")
                
                # Verifica se tem múltiplos tamanhos separados por vírgula
                sizes = [s.strip() for s in size.split(",") if s.strip()] if size else []
                
                if len(sizes) > 1:
                    # Atualiza o produto atual com o primeiro tamanho
                    first_size = sizes[0]
                    product_name = f"{name} {first_size}cm"
                    self.db.execute(
                        "UPDATE products SET name=?, description=?, size=?, stock=?, min_stock=?, price=? WHERE id=?",
                        (product_name, desc, first_size, stock, min_stock, price, pid)
                    )
                    
                    # Cria novos produtos para os tamanhos restantes
                    created_count = 0
                    for sz in sizes[1:]:
                        new_product_name = f"{name} {sz}cm"
                        cur = self.db.execute(
                            "INSERT INTO products(name, description, size, stock, min_stock, price) VALUES (?,?,?,?,?,?)",
                            (new_product_name, desc, sz, stock, min_stock, price)
                        )
                        new_pid = int(cur.lastrowid) if cur and cur.lastrowid is not None else None
                        if new_pid and stock and int(stock) > 0:
                            self.db.execute(
                                "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                                (new_pid, 'entrada', int(stock), 'Cadastro inicial', None, datetime.now().strftime("%Y-%m-%d %H:%M"))
                            )
                        try:
                            self.db.audit_log("product", new_pid, "create", details=f"name={new_product_name},stock={stock},price={price}")
                        except Exception:
                            pass
                        created_count += 1
                    
                    self.refresh()
                    self._warn_low_stock()
                    if self.toast_cb:
                        self.toast_cb(f"Produto atualizado e {created_count} novo(s) produto(s) criado(s)")
                else:
                    # Comportamento normal - apenas atualiza
                    self.db.execute(
                        "UPDATE products SET name=?, description=?, size=?, stock=?, min_stock=?, price=? WHERE id=?",
                        (name, desc, size, stock, min_stock, price, pid)
                    )
                    self.refresh()
                    self._warn_low_stock()
                    if self.toast_cb:
                        self.toast_cb("Produto atualizado")
                try:
                    old = row[0]
                    changes: list[str] = []
                    if str(old["name"] or "") != str(name or ""):
                        changes.append(f"name: '{old['name']}' -> '{name}'")
                    if str(old["description"] or "") != str(desc or ""):
                        changes.append(f"description: '{old['description'] or ''}' -> '{desc or ''}'")
                    try:
                        old_stock = int(old["stock"] or 0)
                    except Exception:
                        old_stock = 0
                    if int(stock) != old_stock:
                        changes.append(f"stock: {old_stock} -> {stock}")
                    try:
                        old_min = int(old["min_stock"] or 0)
                    except Exception:
                        old_min = 0
                    if int(min_stock) != old_min:
                        changes.append(f"min_stock: {old_min} -> {min_stock}")
                    details = "; ".join(changes) if changes else None
                    self.db.audit_log("product", pid, "update", details=details)
                except Exception:
                    pass
            except Exception as e:
                show_message(self, "Erro ao atualizar", str(e), ("OK",))

    def delete(self) -> None:
        pid = self.current_id()
        if not pid:
            return
        res = show_message(self, "Confirmação", "Excluir este produto?", ("Não", "Sim"), default=1)
        if res == 1:
            try:
                deps = self.db.query("SELECT COUNT(*) AS c FROM orders WHERE product_id=?", (pid,))
                if deps and deps[0]["c"] > 0:
                    show_message(self, "Aviso",
                                 "Existe(m) pedido(s) usando este produto. Exclua/ajuste os pedidos primeiro.",
                                 ("OK",))
                    return
                
                # Deleta registros relacionados antes de deletar o produto
                self.db.execute("DELETE FROM stock_movements WHERE product_id=?", (pid,))
                self.db.execute("DELETE FROM production_items WHERE product_id=?", (pid,))
                self.db.execute("DELETE FROM products WHERE id=?", (pid,))
                
                self.refresh()
                self._warn_low_stock()
                if self.toast_cb:
                    self.toast_cb("Produto excluído")
                try:
                    self.db.audit_log("product", pid, "delete", details=None)
                except Exception:
                    pass
            except Exception as e:
                show_message(self, "Erro ao excluir", str(e), ("OK",))

class ReadyStockPage(BasePage):
    """Lista de produtos com estoque disponível (Pronta Entrega)."""
    def __init__(self, db: DB, toast_cb: Optional[Callable[[str], None]] = None) -> None:
        super().__init__("Pronta entrega", "Itens disponíveis no estoque")
        self.db = db
        self.toast_cb = toast_cb
        bl = QVBoxLayout(self.body)
        
        # Botões de ação
        actions = QHBoxLayout()
        self.btn_edit: QPushButton = QPushButton("Editar")
        self.btn_edit.clicked.connect(self.edit_product)
        actions.addWidget(self.btn_edit)
        actions.addStretch(1)
        bl.addLayout(actions)
        
        # Barra de pesquisa simples
        search_box = QHBoxLayout()
        search_box.addWidget(QLabel("Pesquisar:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Nome ou descrição…")
        self.search_edit.setClearButtonEnabled(True)
        cast(Any, self.search_edit.textChanged).connect(lambda _t: self.refresh())
        btn_refresh = QPushButton("Atualizar")
        cast(Any, btn_refresh.clicked).connect(self.refresh)
        search_box.addWidget(self.search_edit, 1)
        search_box.addWidget(btn_refresh)
        bl.addLayout(search_box)

        # Tabela
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Nome", "Descrição", "Tamanho", "Estoque"])
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self.edit_product)
        if header := self.table.horizontalHeader():
            header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            header.setStretchLastSection(False)
        bl.addWidget(self.table)
        self.refresh()
        
        # Timer para atualização automática a cada 5 segundos
        self.auto_refresh_timer = QTimer(self)
        self.auto_refresh_timer.timeout.connect(self.refresh)
        self.auto_refresh_timer.start(5000)  # 5000ms = 5 segundos

    def refresh(self) -> None:
        try:
            try:
                term = self.search_edit.text().strip()
            except Exception:
                term = ""
            if term:
                like = f"%{term}%"
                rows = self.db.query(
                    """
                    SELECT id, name, description, size, stock, min_stock
                    FROM products
                    WHERE stock > 0 AND (name LIKE ? OR description LIKE ?)
                    ORDER BY name
                    """,
                    (like, like)
                )
            else:
                rows = self.db.query(
                    "SELECT id, name, description, size, stock, min_stock FROM products WHERE stock > 0 ORDER BY name"
                )
            self.table.setRowCount(0)
            for r in rows:
                row = self.table.rowCount(); self.table.insertRow(row)
                self.table.setItem(row, 0, QTableWidgetItem(r["name"]))
                self.table.setItem(row, 1, QTableWidgetItem(r["description"] or ""))
                self.table.setItem(row, 2, QTableWidgetItem(format_size(r["size"])))
                self.table.setItem(row, 3, QTableWidgetItem(str(r["stock"])) )
                self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(r["id"])) )
        except Exception as e:
            print(f"Erro no refresh da página Estoque: {e}")
            # Não propaga o erro para não travar o aplicativo

    def edit_product(self) -> None:
        """Edita o produto selecionado"""
        current_row = self.table.currentRow()
        if current_row < 0:
            if self.toast_cb:
                self.toast_cb("Selecione um produto para editar")
            return
        
        # Obter ID do produto
        try:
            product_id = int(self.table.verticalHeaderItem(current_row).text())
        except (ValueError, AttributeError):
            if self.toast_cb:
                self.toast_cb("Erro ao obter ID do produto")
            return
        
        # Buscar dados completos do produto
        product_data = self.db.query("SELECT * FROM products WHERE id = ?", (product_id,))
        if not product_data:
            if self.toast_cb:
                self.toast_cb("Produto não encontrado")
            return
        
        product = product_data[0]
        
        # Abrir diálogo de edição
        dialog = ProductDialog(self, product)
        dialog.setWindowTitle("Editar Produto")
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Atualizar produto no banco
            try:
                name, description, size, stock, min_stock = dialog.get_values()
                
                if not name:
                    if self.toast_cb:
                        self.toast_cb("Nome do produto é obrigatório")
                    return
                
                self.db.execute(
                    """
                    UPDATE products 
                    SET name = ?, description = ?, size = ?, stock = ?, min_stock = ?
                    WHERE id = ?
                    """,
                    (name, description, size, stock, min_stock, product_id)
                )
                
                if self.toast_cb:
                    self.toast_cb("Produto atualizado com sucesso!")
                
                self.refresh()
                
            except Exception as e:
                if self.toast_cb:
                    self.toast_cb(f"Erro ao atualizar produto: {str(e)}")


class ProductionPage(BasePage):
    """Página de Produção - Lista de produção editável com atualização automática"""
    def __init__(self, db: DB, toast_cb: Optional[Callable[[str], None]] = None) -> None:
        super().__init__("Produção", "Lista de produção e controle de estoque")
        self.db = db
        self.toast_cb = toast_cb
        
        # Cache para detectar mudanças
        self._last_production_hash = ""
        self._last_stock_hash = ""
        
        # Timer para auto-refresh (a cada 2 segundos)
        self.refresh_timer = QTimer()
        self.refresh_timer.timeout.connect(self._auto_refresh)
        self.refresh_timer.start(2000)  # 2000ms = 2 segundos

        bl = QVBoxLayout(self.body)
        
        # Barra de ações
        actions_box = QHBoxLayout()
        
        # Botão para adicionar item manualmente
        self.btn_add_item = QPushButton("+ Adicionar Item")
        cast(Any, self.btn_add_item.clicked).connect(self._add_production_item)
        actions_box.addWidget(self.btn_add_item)
        
        # Botão para importar pedidos pendentes
        self.btn_import_orders = QPushButton("📋 Importar de Pedidos")
        cast(Any, self.btn_import_orders.clicked).connect(self._import_from_orders)
        actions_box.addWidget(self.btn_import_orders)
        
        # Botão para limpar lista
        self.btn_clear = QPushButton("🗑️ Limpar Lista")
        cast(Any, self.btn_clear.clicked).connect(self._clear_list)
        actions_box.addWidget(self.btn_clear)
        
        actions_box.addStretch(1)
        
        bl.addLayout(actions_box)

        # Container com duas tabelas lado a lado
        tables_container = QHBoxLayout()
        
        # === LADO ESQUERDO: Lista de Produção ===
        left_container = QWidget()
        left_layout = QVBoxLayout(left_container)
        left_layout.setContentsMargins(0, 0, 5, 0)
        
        # Cabeçalho com título, data e botão
        header_box = QHBoxLayout()
        left_title = QLabel("📝 Lista de Produção")
        left_title.setStyleSheet("font-weight: bold; font-size: 14px; padding: 8px;")
        header_box.addWidget(left_title)
        
        # Data picker
        self.production_date = QDateEdit()
        self.production_date.setDate(QDate.currentDate())
        self.production_date.setCalendarPopup(True)
        self.production_date.setDisplayFormat("dd/MM/yyyy")
        
        # Estilo explícito do QDateEdit e do calendário conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.production_date.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.production_date.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
            if (prod_cal := self.production_date.calendarWidget()) is not None:
                if _theme == "dark":
                    prod_cal.setStyleSheet("""
                        QCalendarWidget { background-color: #1a1f2e; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; }
                        QCalendarWidget QWidget { background: #1a1f2e; color: #ffffff; }
                        QCalendarWidget QAbstractItemView { background: #1a1f2e; color: #ffffff; selection-background-color: #2a2f43; selection-color: #ffffff; gridline-color: #2c3550; }
                        QCalendarWidget QAbstractItemView::item:selected { background: #2a2f43; color: #ffffff; border-radius: 4px; }
                        QCalendarWidget QAbstractItemView::item:hover { background: #1e3a5f; color: #ffffff; }
                        QCalendarWidget QTableView { background-color: #1a1f2e; background: #1a1f2e; color: #ffffff; }
                        QCalendarWidget QToolButton { background-color: #252c45; background: #252c45; color: #ffffff; border: 1px solid #2c3550; border-radius: 6px; padding: 6px; }
                        QCalendarWidget QToolButton:hover { background-color: #333b5e; background: #333b5e; }
                        QCalendarWidget QSpinBox { background-color: #1a1f2e; background: #1a1f2e; color: #ffffff; border: 1px solid #2c3550; border-radius: 6px; padding: 2px 6px; }
                        QCalendarWidget QHeaderView::section { background-color: #252c45; background: #252c45; color: #ffffff; border: 1px solid #2c3550; padding: 4px; }
                    """)
                else:
                    prod_cal.setStyleSheet("""
                        QCalendarWidget { background-color: #ffffff; color: #111827; border: 1px solid #e5e7eb; border-radius: 8px; }
                        QCalendarWidget QWidget { background: #ffffff; color: #111827; }
                        QCalendarWidget QAbstractItemView { background: #ffffff; color: #111827; selection-background-color: #e8eefc; selection-color: #1b2240; gridline-color: #e5e7eb; }
                        QCalendarWidget QAbstractItemView::item:selected { background: #e8eefc; color: #1b2240; border-radius: 4px; }
                        QCalendarWidget QAbstractItemView::item:hover { background: #dbeafe; color: #111827; }
                        QCalendarWidget QTableView { background-color: #ffffff; background: #ffffff; color: #111827; }
                        QCalendarWidget QToolButton { background-color: #f3f4f6; background: #f3f4f6; color: #1f2937; border: 1px solid #e5e7eb; border-radius: 6px; padding: 4px 8px; }
                        QCalendarWidget QToolButton:hover { background-color: #e5e7eb; background: #e5e7eb; }
                        QCalendarWidget QSpinBox { background-color: #ffffff; background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 6px; padding: 2px 6px; }
                        QCalendarWidget QHeaderView::section { background-color: #f9fafb; background: #f9fafb; color: #1f2937; border: 1px solid #e5e7eb; padding: 4px; }
                    """)
        except Exception:
            pass
        
        cast(Any, self.production_date.dateChanged).connect(self._on_date_changed)
        header_box.addWidget(self.production_date)
        
        # Botão "Disparar para Cozinha"
        self.btn_dispatch = QPushButton("🔥 Disparar para Cozinha")
        self.btn_dispatch.setStyleSheet("""
            QPushButton {
                background: #16a34a;
                color: white;
                font-weight: bold;
                padding: 8px 16px;
                border-radius: 6px;
            }
            QPushButton:hover {
                background: #15803d;
            }
        """)
        cast(Any, self.btn_dispatch.clicked).connect(self._dispatch_to_kitchen)
        header_box.addWidget(self.btn_dispatch)
        
        header_box.addStretch(1)
        left_layout.addLayout(header_box)
        
        self.production_table = QTableWidget(0, 4)
        self.production_table.setHorizontalHeaderLabels(["Produto", "Qtd", "Tamanho", "Obs"])
        self.production_table.setAlternatingRowColors(True)
        self.production_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        if header := self.production_table.horizontalHeader():
            header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Produto estica
            header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)  # Qtd
            header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)  # Tamanho
            header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Obs estica
        
        # Conectar sinal de edição de célula
        cast(Any, self.production_table.itemChanged).connect(self._on_production_item_changed)
        
        left_layout.addWidget(self.production_table)
        
        # Botão para remover item selecionado
        remove_btn_box = QHBoxLayout()
        remove_btn_box.addStretch(1)
        self.btn_remove_item = QPushButton("- Remover Selecionado")
        cast(Any, self.btn_remove_item.clicked).connect(self._remove_selected_item)
        remove_btn_box.addWidget(self.btn_remove_item)
        left_layout.addLayout(remove_btn_box)
        
        tables_container.addWidget(left_container, 1)
        
        # === LADO DIREITO: Estoque Atual ===
        right_container = QWidget()
        right_layout = QVBoxLayout(right_container)
        right_layout.setContentsMargins(5, 0, 0, 0)
        
        right_title = QLabel("📦 Estoque Atual")
        right_title.setStyleSheet("font-weight: bold; font-size: 14px; padding: 8px;")
        right_layout.addWidget(right_title)
        
        # Barra de pesquisa para estoque
        search_stock_box = QHBoxLayout()
        search_stock_box.addWidget(QLabel("Pesquisar:"))
        self.search_stock_edit = QLineEdit()
        self.search_stock_edit.setPlaceholderText("Produto...")
        self.search_stock_edit.setClearButtonEnabled(True)
        cast(Any, self.search_stock_edit.textChanged).connect(lambda _t: self._refresh_stock_table())
        search_stock_box.addWidget(self.search_stock_edit, 1)
        right_layout.addLayout(search_stock_box)
        
        self.stock_table = QTableWidget(0, 2)
        self.stock_table.setHorizontalHeaderLabels(["Qtd", "Encomendas"])
        self.stock_table.setAlternatingRowColors(True)
        self.stock_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        if header2 := self.stock_table.horizontalHeader():
            header2.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        # Adiciona coluna vertical com nome dos produtos
        self.stock_table.verticalHeader().setVisible(True)
        right_layout.addWidget(self.stock_table)
        
        tables_container.addWidget(right_container, 1)
        
        bl.addLayout(tables_container)
        
        # Carrega dados iniciais
        self.refresh()

    def refresh(self) -> None:
        """Atualiza ambas as tabelas (produção e estoque)"""
        self._refresh_production_table()
        self._refresh_stock_table()
    
    def _auto_refresh(self) -> None:
        """Auto-refresh com detecção de mudanças para evitar flickering"""
        # Gera hash do estado atual da produção
        production_rows = self.db.query("""
            SELECT pi.id, p.name, pi.quantity, pi.size
            FROM production_items pi
            JOIN products p ON p.id = pi.product_id
            ORDER BY pi.created_at DESC
        """)
        
        # Gera hash do estado atual do estoque com encomendas da data selecionada
        selected_date = self.production_date.date().toString("yyyy-MM-dd")
        stock_rows = self.db.query("""
            SELECT p.id, p.name, p.stock, 
                   COALESCE(SUM(o.quantity), 0) as total_orders
            FROM products p
            LEFT JOIN orders o ON o.product_id = p.id 
                              AND o.status != 'Entregue' 
                              AND DATE(o.delivery_date) = ?
            GROUP BY p.id, p.name, p.stock
            ORDER BY 
              CASE WHEN p.stock <= 0 THEN 0 WHEN p.stock < 5 THEN 1 ELSE 2 END,
              p.name
        """, (selected_date,))
        
        # Cria hashes simples baseados nos dados
        production_data = str([(r["id"], r["name"], r["quantity"], r["size"]) for r in production_rows])
        stock_data = str([(r["id"], r["name"], r["stock"], r["total_orders"]) for r in stock_rows])
        
        # Só atualiza se houve mudança
        if production_data != self._last_production_hash:
            self._last_production_hash = production_data
            self._refresh_production_table()
        
        if stock_data != self._last_stock_hash:
            self._last_stock_hash = stock_data
            self._refresh_stock_table()
    
    def _refresh_production_table(self) -> None:
        """Atualiza a tabela de produção (lado esquerdo) - carrega do banco"""
        # Desconecta temporariamente o sinal para evitar trigger durante refresh
        try:
            cast(Any, self.production_table.itemChanged).disconnect(self._on_production_item_changed)
        except Exception:
            pass
        
        # Busca todos os itens da lista de produção salvos
        rows = self.db.query("""
            SELECT pi.id, p.id as product_id, p.name, pi.quantity, pi.size, pi.notes
            FROM production_items pi
            JOIN products p ON p.id = pi.product_id
            ORDER BY pi.created_at DESC
        """)
        
        self.production_table.setRowCount(0)
        for r in rows:
            row = self.production_table.rowCount()
            self.production_table.insertRow(row)
            
            # Coluna 0: Nome do produto (não editável)
            name_item = QTableWidgetItem(str(r["name"]))
            name_item.setData(Qt.ItemDataRole.UserRole, r["id"])  # Armazena ID do production_item
            name_item.setData(Qt.ItemDataRole.UserRole + 1, r["product_id"])  # Armazena product_id
            name_item.setFlags(name_item.flags() & ~Qt.ItemFlag.ItemIsEditable)  # Não editável
            self.production_table.setItem(row, 0, name_item)
            
            # Coluna 1: Quantidade (editável)
            qty_item = QTableWidgetItem(str(r["quantity"]))
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.production_table.setItem(row, 1, qty_item)
            
            # Coluna 2: Tamanho (editável)
            size_text = format_size(str(r["size"])) if r["size"] else ""
            size_item = QTableWidgetItem(size_text)
            size_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            size_item.setData(Qt.ItemDataRole.UserRole, str(r["size"]))  # Armazena tamanho bruto
            self.production_table.setItem(row, 2, size_item)
            
            # Coluna 3: Observações (editável)
            notes_text = str(r["notes"]) if r["notes"] else ""
            notes_item = QTableWidgetItem(notes_text)
            self.production_table.setItem(row, 3, notes_item)
        
        # Reconecta o sinal
        cast(Any, self.production_table.itemChanged).connect(self._on_production_item_changed)
    
    def _on_production_item_changed(self, item: QTableWidgetItem) -> None:
        """Chamado quando uma célula da tabela de produção é editada"""
        row = item.row()
        col = item.column()
        
        # Ignora edições na coluna de nome (não deveria acontecer)
        if col == 0:
            return
        
        # Pega o ID do item da primeira coluna
        name_item = self.production_table.item(row, 0)
        if not name_item:
            return
        
        item_id = name_item.data(Qt.ItemDataRole.UserRole)
        product_id = name_item.data(Qt.ItemDataRole.UserRole + 1)  # product_id armazenado
        if not item_id:
            return
        
        try:
            if col == 1:  # Quantidade
                new_qty = int(item.text().strip())
                if new_qty <= 0:
                    raise ValueError("Quantidade deve ser maior que zero")
                
                # Busca a quantidade antiga para calcular a diferença
                old_qty_row = self.db.query(
                    "SELECT quantity FROM production_items WHERE id = ?",
                    (item_id,)
                )
                old_qty = int(old_qty_row[0]["quantity"]) if old_qty_row else 0
                
                # Atualiza a quantidade na lista de produção
                self.db.execute(
                    "UPDATE production_items SET quantity = ? WHERE id = ?",
                    (new_qty, item_id)
                )
                
                # Atualiza o estoque do produto com a diferença
                quantity_change = new_qty - old_qty
                if quantity_change != 0 and product_id:
                    self.db.execute(
                        "UPDATE products SET stock = stock + ? WHERE id = ?",
                        (quantity_change, product_id)
                    )
                    
                    if self.toast_cb:
                        self.toast_cb(f"Quantidade atualizada: {new_qty} (estoque {quantity_change:+d})")
                    
                    # Atualiza a tabela de estoque
                    self._refresh_stock_table()
                else:
                    if self.toast_cb:
                        self.toast_cb(f"Quantidade atualizada: {new_qty}")
            
            elif col == 2:  # Tamanho
                new_size = item.text().strip()
                
                self.db.execute(
                    "UPDATE production_items SET size = ? WHERE id = ?",
                    (new_size, item_id)
                )
                
                # Atualiza o display com formatação
                item.setText(format_size(new_size) if new_size else "")
                item.setData(Qt.ItemDataRole.UserRole, new_size)
                
                if self.toast_cb:
                    self.toast_cb(f"Tamanho atualizado: {format_size(new_size) if new_size else 'N/A'}")
            
            elif col == 3:  # Observações
                new_notes = item.text().strip()
                
                self.db.execute(
                    "UPDATE production_items SET notes = ? WHERE id = ?",
                    (new_notes, item_id)
                )
                
                if self.toast_cb:
                    self.toast_cb("Observação atualizada")
        
        except ValueError as e:
            if self.toast_cb:
                self.toast_cb(f"Erro: {str(e)}")
            # Reverte para o valor anterior
            self._refresh_production_table()
        except Exception as e:
            if self.toast_cb:
                self.toast_cb(f"Erro ao atualizar: {str(e)}")
            self._refresh_production_table()
    
    def _refresh_stock_table(self) -> None:
        """Atualiza a tabela de estoque (lado direito) com encomendas da data selecionada"""
        try:
            term = self.search_stock_edit.text().strip()
        except Exception:
            term = ""
        
        # Obtém a data selecionada no calendário de produção
        selected_date = self.production_date.date().toString("yyyy-MM-dd")
        
        # Busca produtos com total de encomendas da data selecionada
        if term:
            like = f"%{term}%"
            rows = self.db.query(
                """SELECT p.id, p.name, p.stock, 
                          COALESCE(SUM(o.quantity), 0) as total_orders
                   FROM products p
                   LEFT JOIN orders o ON o.product_id = p.id 
                                      AND o.status != 'Entregue'
                                      AND DATE(o.delivery_date) = ?
                   WHERE p.name LIKE ?
                   GROUP BY p.id, p.name, p.stock
                   ORDER BY 
                     CASE WHEN p.stock <= 0 THEN 0 WHEN p.stock < 5 THEN 1 ELSE 2 END,
                     p.name""", 
                (selected_date, like)
            )
        else:
            rows = self.db.query(
                """SELECT p.id, p.name, p.stock, 
                          COALESCE(SUM(o.quantity), 0) as total_orders
                   FROM products p
                   LEFT JOIN orders o ON o.product_id = p.id 
                                      AND o.status != 'Entregue'
                                      AND DATE(o.delivery_date) = ?
                   GROUP BY p.id, p.name, p.stock
                   ORDER BY 
                     CASE WHEN p.stock <= 0 THEN 0 WHEN p.stock < 5 THEN 1 ELSE 2 END,
                     p.name""",
                (selected_date,)
            )
        
        self.stock_table.setRowCount(0)
        for r in rows:
            row = self.stock_table.rowCount()
            self.stock_table.insertRow(row)
            
            # Nome do produto no header vertical
            self.stock_table.setVerticalHeaderItem(row, QTableWidgetItem(str(r["name"])))
            
            # Qtd (estoque atual)
            stock_item = QTableWidgetItem(str(r["stock"]))
            stock_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            # Destaca em vermelho claro se estoque negativo ou zero
            if r["stock"] <= 0:
                stock_item.setBackground(QColor(255, 102, 102))  # Vermelho claro
                stock_item.setForeground(QColor(255, 255, 255))  # Texto branco
                stock_item.setFont(QFont("Arial", 10, QFont.Weight.Bold))
            # Destaca em amarelo se estoque baixo (menos de 5)
            elif r["stock"] < 5:
                stock_item.setBackground(QColor(255, 193, 7))  # Amarelo
                stock_item.setForeground(QColor(0, 0, 0))  # Texto preto
            self.stock_table.setItem(row, 0, stock_item)
            
            # Encomendas (total de pedidos não entregues)
            orders_item = QTableWidgetItem(str(int(r["total_orders"])))
            orders_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.stock_table.setItem(row, 1, orders_item)
    
    def _add_production_item(self) -> None:
        """Adiciona um item manualmente à lista de produção"""
        dialog = ProductionItemDialog(self.db, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            product_id, product_name, quantity, sizes, notes = dialog.get_values()
            
            # Salva no banco de dados
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            self.db.execute("""
                INSERT INTO production_items (product_id, quantity, size, notes, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (product_id, quantity, sizes, notes, now))
            
            # Atualiza a tabela
            self._refresh_production_table()
            
            if self.toast_cb:
                self.toast_cb(f"Item adicionado: {product_name}")
    
    def _import_from_orders(self) -> None:
        """Importa pedidos pendentes/em produção para a lista baseado na data selecionada"""
        # Usa a data selecionada no campo production_date
        selected_date = self.production_date.date().toString("yyyy-MM-dd")
        
        # Debug: mostra quais pedidos existem na data
        debug_rows = self.db.query(
            """
            SELECT o.id, p.name, o.quantity, o.delivery_date, o.status, o.notes
            FROM orders o
            LEFT JOIN products p ON p.id = o.product_id
            WHERE DATE(o.delivery_date) = ?
            ORDER BY p.name
            """,
            (selected_date,)
        )
        
        if debug_rows:
            print(f"\n📋 Pedidos encontrados para {selected_date}:")
            for dr in debug_rows:
                product_name = dr['name'] if dr['name'] else "Pedidos em lote"
                notes_info = f" (LOTE)" if dr['notes'] and dr['notes'].startswith('LOTE:') else ""
                print(f"  - ID:{dr['id']} | {product_name}{notes_info} | Qtd:{dr['quantity']} | Status:{dr['status']}")
        else:
            print(f"\n📋 Nenhum pedido encontrado para {selected_date}")
        
        rows = self.db.query(
            """
            SELECT 
                o.id as order_id,
                c.name as customer_name,
                CASE 
                    WHEN o.notes LIKE 'LOTE:%' THEN 'Pedidos em lote'
                    ELSE p.name
                END as produto,
                CASE 
                    WHEN o.notes LIKE 'LOTE:%' THEN 'Lote'
                    ELSE p.size
                END as size,
                o.quantity as total,
                o.product_id,
                o.notes
            FROM orders o
            LEFT JOIN products p ON p.id = o.product_id
            LEFT JOIN customers c ON c.id = o.customer_id
            WHERE DATE(o.delivery_date) = ?
            AND o.status IN ('Pendente', 'Em produção', 'Aguardando Pagamento', 'Pago')
            ORDER BY c.name, o.notes
            """,
            (selected_date,)
        )
        
        if not rows:
            if self.toast_cb:
                self.toast_cb(f"Nenhum pedido encontrado para {selected_date}")
            return
        
        # Limpa a lista atual (do banco também)
        self.db.execute("DELETE FROM production_items")
        
        # Adiciona os itens ao banco
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        for r in rows:
            # Para pedidos em lote ou normais
            customer_info = r["customer_name"] if r["customer_name"] else "Cliente desconhecido"
            notes = f"{customer_info} - Importado de pedidos"
            
            self.db.execute("""
                INSERT INTO production_items (product_id, quantity, size, notes, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (r["product_id"], r["total"], r["size"], notes, now))
        
        # Atualiza a tabela
        self._refresh_production_table()
        
        if self.toast_cb:
            self.toast_cb(f"Importados {len(rows)} item(ns) de pedidos pendentes")
    
    def _remove_selected_item(self) -> None:
        """Remove o item selecionado da lista de produção"""
        selected = self.production_table.currentRow()
        if selected >= 0:
            product_name = self.production_table.item(selected, 0).text() if self.production_table.item(selected, 0) else ""
            
            # Busca o ID do item no banco
            item_id = self.production_table.item(selected, 0).data(Qt.ItemDataRole.UserRole)
            
            if item_id:
                # Remove do banco
                self.db.execute("DELETE FROM production_items WHERE id = ?", (item_id,))
                
                # Atualiza a tabela
                self._refresh_production_table()
                
                if self.toast_cb:
                    self.toast_cb(f"Item removido: {product_name}")
            else:
                # Fallback: remove apenas da UI (não deveria acontecer)
                self.production_table.removeRow(selected)
                if self.toast_cb:
                    self.toast_cb(f"Item removido (apenas da UI): {product_name}")
        else:
            if self.toast_cb:
                self.toast_cb("Selecione um item para remover")
    
    def showEvent(self, event) -> None:
        """Inicia o timer quando a página é exibida"""
        super().showEvent(event)
        if hasattr(self, 'refresh_timer'):
            self.refresh_timer.start(2000)
            self.refresh()  # Refresh imediato ao mostrar
    
    def hideEvent(self, event) -> None:
        """Para o timer quando a página é ocultada"""
        super().hideEvent(event)
        if hasattr(self, 'refresh_timer'):
            self.refresh_timer.stop()
    
    
    def _on_date_changed(self, date: QDate) -> None:
        """Callback quando a data é alterada"""
        # Pode adicionar lógica para filtrar por data no futuro
        pass
    
    def _dispatch_to_kitchen(self) -> None:
        """Dispara a lista de produção para a cozinha (painel web)"""
        selected_date = self.production_date.date().toPyDate().isoformat()
        
        # Salva a lista atual com a data selecionada
        production_rows = self.db.query("""
            SELECT p.name, pi.quantity, pi.size, pi.notes
            FROM production_items pi
            JOIN products p ON p.id = pi.product_id
            ORDER BY pi.created_at DESC
        """)
        
        if not production_rows:
            if self.toast_cb:
                self.toast_cb("⚠️ Lista de produção vazia!")
            return
        
        # Salva no arquivo JSON para o painel web
        try:
            import json
            data = {
                "date": selected_date,
                "items": [
                    {
                        "produto": row["name"],
                        "quantidade": row["quantity"],
                        "tamanho": row["size"] or "N/A",
                        "obs": row["notes"] or ""
                    }
                    for row in production_rows
                ]
            }
            
            # Salva no diretório web
            web_dir = os.path.join(os.path.dirname(__file__), "web")
            os.makedirs(web_dir, exist_ok=True)
            
            json_path = os.path.join(web_dir, "production_list.json")
            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            if self.toast_cb:
                self.toast_cb(f"🔥 Lista disparada para cozinha! ({len(production_rows)} itens)")
        
        except Exception as e:
            show_message(self, "Erro", f"Erro ao disparar lista: {e}", ("OK",))
    
    def _clear_list(self) -> None:
        """Limpa toda a lista de produção"""
        if self.production_table.rowCount() == 0:
            if self.toast_cb:
                self.toast_cb("Lista já está vazia")
            return
        
        reply = show_message(
            self,
            "Limpar Lista",
            f"Tem certeza que deseja limpar todos os {self.production_table.rowCount()} item(ns) da lista de produção?",
            ("Cancelar", "Limpar"),
            default=0
        )
        
        if reply == 1:
            # Remove do banco
            self.db.execute("DELETE FROM production_items")
            
            # Atualiza a tabela
            self._refresh_production_table()
            
            if self.toast_cb:
                self.toast_cb("Lista de produção limpa")


class ProductionItemDialog(QDialog):
    """Diálogo para adicionar item manualmente à lista de produção"""
    def __init__(self, db: DB, parent: Optional[QWidget] = None) -> None:
        super().__init__(parent)
        from core.config import load_config
        
        # Aplica estilo conforme tema atual
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "dark":
            self.setStyleSheet("""
                QDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                }
                QLineEdit, QComboBox, QSpinBox, QListWidget {
                    background-color: #3c3c3c;
                    border: 1px solid #555555;
                    padding: 5px;
                    border-radius: 3px;
                    color: #ffffff;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550 !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70 !important;
                }
                QPushButton:pressed {
                    background: #333b5e;
                    border: 1px solid #4a5480 !important;
                }
            """)
        else:
            self.setStyleSheet("""
                QDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                }
                QLineEdit, QComboBox, QSpinBox, QListWidget {
                    background-color: #ffffff;
                    border: 1px solid #d1d5db;
                    padding: 5px;
                    border-radius: 3px;
                    color: #111827;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db !important;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe !important;
                }
                QPushButton:pressed {
                    background: #c7d2fe;
                    border: 1px solid #a5b4fc !important;
                }
            """)
            
        self.db = db
        self.setWindowTitle("Adicionar Item à Produção")
        self.resize(450, 280)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Título
        title_label = QLabel("➕ Adicionar Item à Lista de Produção")
        title_label.setStyleSheet("font-size: 16px; font-weight: bold; margin-bottom: 10px;")
        layout.addWidget(title_label)
        
        # Form layout para campos
        form_layout = QFormLayout()
        form_layout.setSpacing(12)
        form_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        
        # Produto (ComboBox)
        self.product = QComboBox()
        self.product.setMinimumHeight(35)
        self._load_products()
        cast(Any, self.product.currentIndexChanged).connect(self._on_product_changed)
        form_layout.addRow("Produto:", self.product)
        
        # Quantidade
        self.quantity = QSpinBox()
        self.quantity.setMinimum(1)
        self.quantity.setMaximum(1_000_000)
        self.quantity.setValue(1)
        self.quantity.setMinimumHeight(35)
        self.quantity.setSuffix(" un")
        form_layout.addRow("Quantidade:", self.quantity)
        
        # Tamanho (ComboBox único - mais intuitivo)
        self.size_combo = QComboBox()
        self.size_combo.setMinimumHeight(35)
        self.size_combo.setEditable(True)  # Permite edição manual
        self.size_combo.setPlaceholderText("Selecione ou digite o tamanho...")
        form_layout.addRow("Tamanho:", self.size_combo)
        
        # Observações (QLineEdit para texto livre)
        self.notes_edit = QLineEdit()
        self.notes_edit.setMinimumHeight(35)
        self.notes_edit.setPlaceholderText("Digite uma observação (opcional)...")
        form_layout.addRow("Obs:", self.notes_edit)
        
        layout.addLayout(form_layout)
        
        # Carrega tamanhos do primeiro produto
        self._on_product_changed()
        
        # Espaçador
        layout.addStretch(1)
        
        # Botões
        btns_layout = QHBoxLayout()
        btns_layout.addStretch(1)
        
        btn_cancel = QPushButton("✖ Cancelar")
        btn_cancel.setMinimumSize(120, 40)
        cast(Any, btn_cancel.clicked).connect(self.reject)
        
        btn_save = QPushButton("✔ Adicionar")
        btn_save.setMinimumSize(120, 40)
        btn_save.setDefault(True)
        cast(Any, btn_save.clicked).connect(self.accept)
        
        btns_layout.addWidget(btn_cancel)
        btns_layout.addWidget(btn_save)
        
        layout.addLayout(btns_layout)
    
    def _load_products(self) -> None:
        """Carrega lista de produtos no combobox"""
        self.product.clear()
        rows = self.db.query("SELECT id, name FROM products ORDER BY name")
        for r in rows:
            self.product.addItem(r["name"], r["id"])
    
    def _on_product_changed(self) -> None:
        """Atualiza lista de tamanhos quando produto muda"""
        self.size_combo.clear()
        
        prod_id = self.product.currentData()
        if not prod_id:
            return
        
        # Busca os tamanhos do produto
        rows = self.db.query("SELECT size FROM products WHERE id=?", (prod_id,))
        if not rows or not rows[0]["size"]:
            self.size_combo.addItem("Tamanho único")
            return
        
        # Separa tamanhos por vírgula e adiciona ao combo
        size_str = rows[0]["size"]
        sizes = [s.strip() for s in size_str.split(",") if s.strip()]
        
        for size in sizes:
            formatted_size = format_size(size)
            self.size_combo.addItem(formatted_size, size)  # Exibe formatado, guarda original
        
        # Seleciona o primeiro por padrão
        if self.size_combo.count() > 0:
            self.size_combo.setCurrentIndex(0)
    
    def get_values(self) -> tuple[int, str, int, str, str]:
        """Retorna (product_id, nome_produto, quantidade, tamanho_selecionado, observações)"""
        product_name = self.product.currentText()
        
        # Busca o ID do produto selecionado
        product_id = 0
        rows = self.db.query("SELECT id FROM products WHERE name = ?", (product_name,))
        if rows:
            product_id = rows[0]["id"]
        
        qty = self.quantity.value()
        
        # Pega o tamanho selecionado ou digitado
        # Tenta pegar o valor original (userData), senão pega o texto exibido
        size = self.size_combo.currentData()
        if size is None or size == "":
            size = self.size_combo.currentText().strip()
        
        # Pega as observações
        notes = self.notes_edit.text().strip()
        
        return product_id, product_name, qty, size, notes


# ========================================
# Database Dialog - Modal para gerenciamento completo do banco
# ========================================
class DatabaseDialog(QDialog):
    """Diálogo modal para gerenciamento completo do banco de dados"""
    
    def __init__(self, parent: Optional[QWidget] = None, toast_cb: Optional[Callable[[str], None]] = None, backup_cb: Optional[Callable[[], None]] = None) -> None:
        super().__init__(parent)
        self.toast_cb = toast_cb
        self.backup_cb = backup_cb
        self.setWindowTitle("Gerenciamento de Banco de Dados")
        self.setMinimumSize(700, 600)
        self.setModal(True)
        
        # Detecta o tema atual
        try:
            from core.config import load_config
            theme = load_config().get("theme", "light")
            is_dark = (theme == "dark")
        except Exception:
            is_dark = False
        
        # Define cores baseadas no tema
        if is_dark:
            bg_main = "#1e1e1e"
            bg_group = "rgba(60, 60, 60, 0.3)"
            text_main = "#f3f4f6"
            text_secondary = "#9ca3af"
            btn_bg = "#4b5563"
            btn_hover = "#6b7280"
            btn_pressed = "#374151"
            btn_primary = "#3b82f6"
            btn_primary_hover = "#2563eb"
            warning_bg = "rgba(245, 158, 11, 0.15)"
            warning_text = "#fbbf24"
        else:
            bg_main = "#ffffff"
            bg_group = "rgba(100, 100, 100, 0.08)"
            text_main = "#1f2937"
            text_secondary = "#6b7280"
            btn_bg = "#e5e7eb"
            btn_hover = "#d1d5db"
            btn_pressed = "#c7d2fe"
            btn_primary = "#3b82f6"
            btn_primary_hover = "#2563eb"
            warning_bg = "rgba(245, 158, 11, 0.1)"
            warning_text = "#d97706"
        
        # Aplica estilo ao diálogo
        self.setStyleSheet(f"""
            QDialog {{
                background: {bg_main};
            }}
            QLabel {{
                color: {text_main};
            }}
        """)
        
        # Layout principal
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(16)
        main_layout.setContentsMargins(24, 24, 24, 24)
        
        # Título
        title = QLabel("🗄️ Banco de Dados")
        title.setStyleSheet(f"font-size: 20px; font-weight: bold; margin-bottom: 8px; color: {text_main};")
        main_layout.addWidget(title)
        
        # Área de scroll para o conteúdo
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        scroll.setStyleSheet(f"QScrollArea {{ background: transparent; border: none; }}")
        
        content = QWidget()
        content.setStyleSheet(f"QWidget {{ background: transparent; }}")
        content_layout = QVBoxLayout(content)
        content_layout.setSpacing(20)
        
        # === Seção: Caminho Atual ===
        path_group = QFrame()
        path_group.setObjectName("SettingsGroup")
        path_group.setStyleSheet(f"""
            QFrame#SettingsGroup {{
                background: {bg_group};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        path_layout = QVBoxLayout(path_group)
        
        path_title = QLabel("<b>Caminho Atual</b>")
        path_title.setStyleSheet(f"font-size: 14px; color: {text_main};")
        path_layout.addWidget(path_title)
        
        self.db_path_label = QLabel()
        self.db_path_label.setWordWrap(True)
        self.db_path_label.setStyleSheet(f"color: {text_secondary}; padding: 8px; font-family: monospace; font-size: 12px;")
        path_layout.addWidget(self.db_path_label)
        
        content_layout.addWidget(path_group)
        
        # === Seção: Ações ===
        actions_group = QFrame()
        actions_group.setObjectName("SettingsGroup")
        actions_group.setStyleSheet(f"""
            QFrame#SettingsGroup {{
                background: {bg_group};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        actions_layout = QVBoxLayout(actions_group)
        
        actions_title = QLabel("<b>Ações</b>")
        actions_title.setStyleSheet(f"font-size: 14px; margin-bottom: 8px; color: {text_main};")
        actions_layout.addWidget(actions_title)
        
        # Botões de ação
        buttons_layout = QVBoxLayout()
        buttons_layout.setSpacing(12)
        
        self.btn_select_db = QPushButton("📂 Selecionar Banco de Dados")
        self.btn_create_new_db = QPushButton("➕ Criar Novo Banco")
        self.btn_connect_network = QPushButton("🌐 Conectar a Rede")
        
        btn_style = f"""
            QPushButton {{
                background: {btn_bg};
                color: {text_main if not is_dark else 'white'};
                border: none;
                border-radius: 8px;
                padding: 10px 16px;
                font-size: 14px;
                font-weight: 500;
                text-align: left;
            }}
            QPushButton:hover {{
                background: {btn_hover};
            }}
            QPushButton:pressed {{
                background: {btn_pressed};
            }}
        """
        
        for btn in [self.btn_select_db, self.btn_create_new_db, self.btn_connect_network]:
            btn.setMinimumHeight(40)
            btn.setStyleSheet(btn_style)
            buttons_layout.addWidget(btn)
        
        actions_layout.addLayout(buttons_layout)
        content_layout.addWidget(actions_group)
        
        # === Seção: Status ===
        status_group = QFrame()
        status_group.setObjectName("SettingsGroup")
        status_group.setStyleSheet(f"""
            QFrame#SettingsGroup {{
                background: {bg_group};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        status_layout = QVBoxLayout(status_group)
        
        status_title = QLabel("<b>Status do Banco</b>")
        status_title.setStyleSheet(f"font-size: 14px; margin-bottom: 8px; color: {text_main};")
        status_layout.addWidget(status_title)
        
        self.db_status_label = QLabel()
        self.db_status_label.setWordWrap(True)
        status_bg = "rgba(0,0,0,0.1)" if not is_dark else "rgba(255,255,255,0.05)"
        self.db_status_label.setStyleSheet(f"color: {text_secondary}; padding: 8px; background: {status_bg}; border-radius: 4px; font-size: 12px;")
        status_layout.addWidget(self.db_status_label)
        
        # Botões de status
        status_buttons = QHBoxLayout()
        self.btn_check_status = QPushButton("🔍 Verificar Status")
        self.btn_test_shared = QPushButton("🔗 Testar Modo Compartilhado")
        
        status_btn_style = f"""
            QPushButton {{
                background: {btn_bg};
                color: {text_main if not is_dark else 'white'};
                border: none;
                border-radius: 8px;
                padding: 8px 14px;
                font-size: 13px;
            }}
            QPushButton:hover {{
                background: {btn_hover};
            }}
            QPushButton:pressed {{
                background: {btn_pressed};
            }}
        """
        
        for btn in [self.btn_check_status, self.btn_test_shared]:
            btn.setMinimumHeight(36)
            btn.setStyleSheet(status_btn_style)
            status_buttons.addWidget(btn)
        
        status_layout.addLayout(status_buttons)
        content_layout.addWidget(status_group)
        
        # === Seção: Backup ===
        backup_group = QFrame()
        backup_group.setObjectName("SettingsGroup")
        backup_group.setStyleSheet(f"""
            QFrame#SettingsGroup {{
                background: {bg_group};
                border-radius: 8px;
                padding: 12px;
            }}
        """)
        backup_layout = QVBoxLayout(backup_group)
        
        backup_title = QLabel("<b>Backup</b>")
        backup_title.setStyleSheet(f"font-size: 14px; margin-bottom: 8px; color: {text_main};")
        backup_layout.addWidget(backup_title)
        
        # Botão de automação
        self.btn_config_auto_backup = QPushButton("⚙️ Configurar Automação do Backup")
        self.btn_config_auto_backup.setMinimumHeight(40)
        btn_config_bg = "#8b5cf6" if not is_dark else "#7c3aed"
        btn_config_hover = "#7c3aed" if not is_dark else "#6d28d9"
        self.btn_config_auto_backup.setStyleSheet(f"""
            QPushButton {{
                background: {btn_config_bg};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 18px;
                font-size: 14px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: {btn_config_hover};
            }}
            QPushButton:pressed {{
                background: #6d28d9;
            }}
        """)
        backup_layout.addWidget(self.btn_config_auto_backup)
        
        self.btn_do_backup = QPushButton("💾 Fazer Backup Agora")
        self.btn_do_backup.setMinimumHeight(40)
        self.btn_do_backup.setStyleSheet(f"""
            QPushButton {{
                background: {btn_primary};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 18px;
                font-size: 14px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: {btn_primary_hover};
            }}
            QPushButton:pressed {{
                background: #1d4ed8;
            }}
        """)
        backup_layout.addWidget(self.btn_do_backup)
        
        # Botão para backup na nuvem
        self.btn_cloud_backup = QPushButton("☁️ Fazer Backup na Nuvem Agora")
        self.btn_cloud_backup.setMinimumHeight(40)
        btn_cloud_bg = "#10b981" if not is_dark else "#059669"
        btn_cloud_hover = "#059669" if not is_dark else "#047857"
        self.btn_cloud_backup.setStyleSheet(f"""
            QPushButton {{
                background: {btn_cloud_bg};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 18px;
                font-size: 14px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: {btn_cloud_hover};
            }}
            QPushButton:pressed {{
                background: #047857;
            }}
        """)
        backup_layout.addWidget(self.btn_cloud_backup)
        
        # Botão para restaurar backup
        self.btn_restore_backup = QPushButton("🔄 Restaurar Backup")
        self.btn_restore_backup.setMinimumHeight(40)
        btn_restore_bg = "#f59e0b" if not is_dark else "#d97706"
        btn_restore_hover = "#d97706" if not is_dark else "#b45309"
        self.btn_restore_backup.setStyleSheet(f"""
            QPushButton {{
                background: {btn_restore_bg};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 10px 18px;
                font-size: 14px;
                font-weight: 500;
            }}
            QPushButton:hover {{
                background: {btn_restore_hover};
            }}
            QPushButton:pressed {{
                background: #b45309;
            }}
        """)
        backup_layout.addWidget(self.btn_restore_backup)
        content_layout.addWidget(backup_group)
        
        # Aviso
        warning = QLabel("⚠️ Após alterar o banco de dados, reinicie o sistema para aplicar as mudanças.")
        warning.setWordWrap(True)
        warning.setStyleSheet(f"color: {warning_text}; padding: 12px; font-size: 13px; background: {warning_bg}; border-radius: 6px;")
        content_layout.addWidget(warning)
        
        content_layout.addStretch()
        scroll.setWidget(content)
        main_layout.addWidget(scroll)
        
        # Botão Fechar
        close_button = QPushButton("Fechar")
        close_button.setMinimumHeight(36)
        close_btn_bg = "#6b7280" if is_dark else "#9ca3af"
        close_btn_hover = "#4b5563" if is_dark else "#6b7280"
        close_button.setStyleSheet(f"""
            QPushButton {{
                background: {close_btn_bg};
                color: white;
                border: none;
                border-radius: 8px;
                padding: 8px 24px;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background: {close_btn_hover};
            }}
        """)
        cast(Any, close_button.clicked).connect(self.accept)
        main_layout.addWidget(close_button, alignment=Qt.AlignmentFlag.AlignRight)
        
        # Conectar eventos
        cast(Any, self.btn_select_db.clicked).connect(self.select_database)
        cast(Any, self.btn_create_new_db.clicked).connect(self.create_new_database)
        cast(Any, self.btn_connect_network.clicked).connect(self.connect_to_network)
        cast(Any, self.btn_check_status.clicked).connect(self.refresh_db_status)
        cast(Any, self.btn_test_shared.clicked).connect(self.test_shared_mode)
        cast(Any, self.btn_config_auto_backup.clicked).connect(self.configure_auto_backup)
        cast(Any, self.btn_do_backup.clicked).connect(lambda: self.backup_cb() if self.backup_cb else None)
        cast(Any, self.btn_cloud_backup.clicked).connect(self.do_cloud_backup_now)
        cast(Any, self.btn_restore_backup.clicked).connect(self.restore_backup)
        
        # Atualizar labels
        self.update_db_path_label()
        self.refresh_db_status()
    
    def update_db_path_label(self) -> None:
        """Atualiza o label com o caminho do banco atual"""
        try:
            from core.config import get_database_path
            path = get_database_path()
            self.db_path_label.setText(path)
        except Exception as e:
            self.db_path_label.setText(f"Erro: {e}")
    
    def select_database(self) -> None:
        """Seleciona um banco de dados existente"""
        choice = show_message(
            self,
            "Selecionar Banco de Dados",
            "Escolha o tipo de localização do banco:\n\n"
            "💻 LOCAL: Para arquivos no seu computador (C:\\, D:\\, documentos, etc.)\n"
            "   • Usa diálogo Qt (mais rápido)\n"
            "   • Ideal para bancos em pastas locais\n\n"
            "🌐 REDE: Para arquivos em pastas compartilhadas de rede (Z:\\, \\\\servidor\\pasta)\n"
            "   • Usa Explorer do Windows (melhor para rede)\n"
            "   • Permite navegação em UNC paths e unidades mapeadas\n"
            "   • Validação assíncrona (não trava)",
            ("Cancelar", "💻 Local", "🌐 Rede"),
            0
        )
        
        if choice == 0:
            return
        
        use_native = (choice == 2)
        
        try:
            from core.config import get_database_path, set_database_path, validate_database_path
            
            if use_native:
                # Usa Explorer do Windows (melhor para rede)
                if self.toast_cb:
                    self.toast_cb("Abrindo Explorer do Windows...")
                
                # Abre diálogo nativo do Windows
                import subprocess
                import tempfile
                
                # Cria um script VBS temporário para seleção de arquivo
                vbs_code = '''
Set objDialog = CreateObject("SAFRdialog.FileSave")
objDialog.FileName = "confeitaria.db"
objDialog.FileType = "Banco de Dados SQLite (*.db)"
objDialog.Title = "Selecionar Banco de Dados"
If objDialog.Show Then
    WScript.Echo objDialog.FileName
End If
'''
                # Usa QFileDialog com opção nativa
                file_path, _ = QFileDialog.getOpenFileName(
                    self,
                    "Selecionar Banco de Dados - Confeitaria",
                    os.path.expanduser("~"),
                    "Banco de Dados SQLite (*.db);;Todos os arquivos (*.*)",
                    options=QFileDialog.Option.DontUseNativeDialog if not use_native else QFileDialog.Option(0)
                )
            else:
                # Usa diálogo Qt padrão (mais rápido para local)
                current_db = get_database_path()
                start_dir = os.path.dirname(current_db) if current_db and os.path.exists(current_db) else os.path.expanduser("~")
                
                file_path, _ = QFileDialog.getOpenFileName(
                    self,
                    "Selecionar Banco de Dados - Confeitaria",
                    start_dir,
                    "Banco de Dados SQLite (*.db);;Todos os arquivos (*.*)"
                )
            
            if not file_path:
                return
            
            # Valida o banco de dados de forma assíncrona
            progress = QProgressDialog(
                "Validando banco de dados...\n\nPor favor aguarde.",
                "Cancelar",
                0, 0,
                self
            )
            progress.setWindowTitle("Confeitaria - Validação")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.show()
            QApplication.processEvents()
            
            # Cria thread de validação
            validator = AsyncDatabaseValidator(file_path, timeout=10)
            
            def on_complete(is_valid: bool, message: str, validated_path: str):
                progress.close()
                
                if not is_valid:
                    show_message(
                        self,
                        "Erro de Validação",
                        f"O arquivo selecionado não é um banco de dados válido:\n\n{message}",
                        ("OK",)
                    )
                    return
                
                # Salva a configuração
                if set_database_path(validated_path):
                    self.update_db_path_label()
                    self.refresh_db_status()
                    
                    restart = show_message(
                        self,
                        "Configuração Atualizada",
                        f"Banco de dados configurado com sucesso!\n\n📁 {validated_path}\n\n🔄 É necessário reiniciar o aplicativo para aplicar as mudanças.\n\nDeseja reiniciar agora?",
                        ("Mais Tarde", "Reiniciar Agora"),
                        default=1
                    )
                    
                    if restart == 1:
                        if self.toast_cb:
                            self.toast_cb("Reiniciando aplicativo...")
                        QTimer.singleShot(1000, lambda: self._restart_application())
                    else:
                        if self.toast_cb:
                            self.toast_cb("⚠️ Reinicie o aplicativo para aplicar as mudanças!")
                else:
                    show_message(
                        self,
                        "Erro",
                        "Não foi possível salvar a configuração de banco.\n\nVerifique as permissões do sistema.",
                        ("OK",)
                    )
            
            def on_progress(msg: str):
                progress.setLabelText(f"Validando banco de dados...\n\n{msg}")
                QApplication.processEvents()
            
            def on_cancel():
                validator.stop()
                validator.quit()
                validator.wait(1000)
                progress.close()
                if self.toast_cb:
                    self.toast_cb("❌ Validação cancelada")
            
            cast(Any, validator.finished).connect(on_complete)
            cast(Any, validator.progress).connect(on_progress)
            cast(Any, progress.canceled).connect(on_cancel)
            
            validator.start()
            
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao selecionar banco de dados:\n\n{e}",
                ("OK",)
            )
    
    def _restart_application(self) -> None:
        """Reinicia o aplicativo"""
        try:
            import subprocess
            import sys
            
            # Fecha o servidor web para evitar erro de thread
            try:
                if hasattr(self, 'web_thread') and self.web_thread:
                    print("Finalizando servidor web...")
                    # Não espera, apenas fecha
            except Exception:
                pass
            
            # Pega o executável atual
            if getattr(sys, 'frozen', False):
                # Executável PyInstaller
                executable = sys.executable
                script = None
            else:
                # Modo desenvolvimento
                executable = sys.executable
                script = sys.argv[0]
            
            # Inicia novo processo
            if getattr(sys, 'frozen', False):
                # PyInstaller: apenas executa o .exe
                subprocess.Popen([executable], 
                               creationflags=subprocess.CREATE_NEW_CONSOLE if sys.platform == 'win32' else 0)
            elif script:
                # Desenvolvimento: python script.py
                subprocess.Popen([executable, script],
                               creationflags=subprocess.CREATE_NEW_CONSOLE if sys.platform == 'win32' else 0)
            
            print("Novo processo iniciado. Fechando aplicação atual...")
            
            # Fecha a aplicação atual com delay
            QTimer.singleShot(500, lambda: QApplication.instance().quit())
            
        except Exception as e:
            print(f"Erro ao reiniciar: {e}")
            show_message(
                self,
                "Erro ao Reiniciar",
                f"Não foi possível reiniciar automaticamente:\n\n{e}\n\nPor favor, reinicie manualmente o aplicativo.",
                ("OK",)
            )
    
    def create_new_database(self) -> None:
        """Cria um novo banco de dados"""
        try:
            from core.config import set_database_path, validate_database_path, get_app_data_directory
            
            # Sugere o diretório de dados da aplicação
            try:
                start_dir = get_app_data_directory()
            except Exception:
                start_dir = os.path.expanduser("~")
            
            # Pede ao usuário para escolher onde salvar o novo banco
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Criar Novo Banco de Dados - Confeitaria",
                os.path.join(start_dir, "confeitaria_novo.db"),
                "Banco de Dados SQLite (*.db);;Todos os arquivos (*.*)"
            )
            
            if not file_path:
                return
            
            # Garante que termina com .db
            if not file_path.lower().endswith('.db'):
                file_path += '.db'
            
            # Verifica se o arquivo já existe
            if os.path.exists(file_path):
                overwrite = show_message(
                    self,
                    "Arquivo Existente",
                    f"O arquivo já existe:\n\n{file_path}\n\nDeseja sobrescrever?",
                    ("Cancelar", "Sobrescrever"),
                    default=0
                )
                if overwrite == 0:
                    return
                # Remove o arquivo antigo
                try:
                    os.remove(file_path)
                except Exception as e:
                    show_message(
                        self,
                        "Erro",
                        f"Não foi possível remover o arquivo existente:\n\n{e}",
                        ("OK",)
                    )
                    return
            
            # Cria o novo banco
            try:
                import sqlite3
                conn = sqlite3.connect(file_path)
                conn.close()
                
                # Valida
                is_valid, message = validate_database_path(file_path)
                if not is_valid:
                    show_message(
                        self,
                        "Erro",
                        f"Falha ao criar banco válido:\n\n{message}",
                        ("OK",)
                    )
                    try:
                        os.remove(file_path)
                    except Exception:
                        pass
                    return
                
                # Salva nas configurações
                if set_database_path(file_path):
                    self.update_db_path_label()
                    self.refresh_db_status()
                    
                    restart = show_message(
                        self,
                        "Banco Criado",
                        f"Novo banco de dados criado com sucesso!\n\n📁 {file_path}\n\n"
                        f"🔄 É necessário reiniciar o aplicativo para usar o novo banco.\n\n"
                        f"Deseja reiniciar agora?",
                        ("Mais Tarde", "Reiniciar Agora"),
                        default=1
                    )
                    
                    if restart == 1:
                        if self.toast_cb:
                            self.toast_cb("Reiniciando aplicativo...")
                        QTimer.singleShot(1000, lambda: self._restart_application())
                    else:
                        if self.toast_cb:
                            self.toast_cb("⚠️ Reinicie o aplicativo para usar o novo banco!")
                else:
                    show_message(
                        self,
                        "Erro",
                        "Não foi possível salvar a configuração.\n\nVerifique as permissões do sistema.",
                        ("OK",)
                    )
            except Exception as e:
                show_message(
                    self,
                    "Erro",
                    f"Erro ao criar banco de dados:\n\n{e}",
                    ("OK",)
                )
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception:
                    pass
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao criar novo banco:\n\n{e}",
                ("OK",)
            )
    
    def connect_to_network(self) -> None:
        """Conecta a um banco em rede"""
        try:
            from core.config import set_database_path, validate_database_path
            
            # Input manual do caminho UNC
            network_path, ok = QInputDialog.getText(
                self,
                "Conectar a Banco em Rede",
                "Digite o caminho UNC completo do banco de dados:\n\n"
                "Exemplo: \\\\NOME_PC\\Compartilhamento\\confeitaria.db\n\n"
                "⚠️ IMPORTANTE:\n"
                "• Use \\\\ no início (caminho UNC)\n"
                "• O compartilhamento deve estar acessível\n"
                "• Você precisa ter permissão de leitura/escrita\n\n"
                "Caminho UNC:",
                QLineEdit.EchoMode.Normal,
                "\\\\"
            )
            
            if not ok or not network_path or not network_path.strip():
                return
            
            network_path = network_path.strip()
            
            # Valida formato UNC
            if not network_path.startswith("\\\\"):
                show_message(
                    self,
                    "Formato Inválido",
                    "O caminho deve começar com \\\\ (formato UNC).\n\n"
                    "Exemplo correto:\n\\\\NOME_PC\\Compartilhamento\\confeitaria.db",
                    ("OK",)
                )
                return
            
            # Verifica se o arquivo existe
            if not os.path.exists(network_path):
                create = show_message(
                    self,
                    "Arquivo Não Encontrado",
                    f"O arquivo não foi encontrado:\n\n{network_path}\n\n"
                    f"Possíveis causas:\n"
                    f"• Caminho digitado incorretamente\n"
                    f"• Compartilhamento não acessível\n"
                    f"• Sem permissão de acesso\n\n"
                    f"Deseja tentar criar um novo banco neste local?",
                    ("Cancelar", "Criar Novo"),
                    default=0
                )
                
                if create == 0:
                    return
                
                # Tenta criar
                try:
                    import sqlite3
                    conn = sqlite3.connect(network_path)
                    conn.close()
                except Exception as e:
                    show_message(
                        self,
                        "Erro ao Criar",
                        f"Não foi possível criar o banco:\n\n{e}\n\n"
                        f"Verifique:\n"
                        f"• Se o compartilhamento existe\n"
                        f"• Se você tem permissão de escrita\n"
                        f"• Se o caminho está correto",
                        ("OK",)
                    )
                    return
            
            # Validação assíncrona com timeout
            progress = QProgressDialog(
                "Validando banco de dados em rede...\n\nPor favor aguarde.",
                "Cancelar",
                0, 0,
                self
            )
            progress.setWindowTitle("Confeitaria - Validação de Rede")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.show()
            QApplication.processEvents()
            
            # Cria thread de validação (timeout de 20s para rede)
            validator = AsyncDatabaseValidator(network_path, timeout=20)
            
            def on_network_complete(is_valid: bool, message: str, validated_path: str):
                progress.close()
                
                if not is_valid:
                    show_message(
                        self,
                        "Erro de Validação",
                        f"O banco de dados não é válido:\n\n{message}\n\n"
                        f"Certifique-se de que o servidor foi configurado corretamente.",
                        ("OK",)
                    )
                    return
                
                # Salva a configuração
                if set_database_path(validated_path):
                    self.update_db_path_label()
                    self.refresh_db_status()
                    
                    restart = show_message(
                        self,
                        "Conexão Configurada",
                        f"✅ Conectado com sucesso!\n\n"
                        f"📁 Caminho: {validated_path}\n"
                        f"🌐 Origem: Rede (UNC)\n\n"
                        f"🔄 É necessário reiniciar o aplicativo para usar o banco de rede.\n\n"
                        f"Deseja reiniciar agora?",
                        ("Mais Tarde", "Reiniciar Agora"),
                        default=1
                    )
                    
                    if restart == 1:
                        if self.toast_cb:
                            self.toast_cb("Reiniciando aplicativo...")
                        QTimer.singleShot(1000, lambda: self._restart_application())
                    else:
                        if self.toast_cb:
                            self.toast_cb("⚠️ Reinicie o aplicativo para usar o banco de rede!")
                else:
                    show_message(
                        self,
                        "Erro",
                        "Não foi possível salvar a configuração de rede.\n\n"
                        "Verifique as permissões do sistema.",
                        ("OK",)
                    )
            
            def on_network_progress(msg: str):
                progress.setLabelText(f"Validando banco de dados em rede...\n\n{msg}")
                QApplication.processEvents()
            
            def on_network_cancel():
                validator.stop()
                validator.quit()
                validator.wait(1000)
                progress.close()
                if self.toast_cb:
                    self.toast_cb("❌ Validação cancelada")
            
            cast(Any, validator.finished).connect(on_network_complete)
            cast(Any, validator.progress).connect(on_network_progress)
            cast(Any, progress.canceled).connect(on_network_cancel)
            
            validator.start()
                
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao conectar à rede:\n\n{e}",
                ("OK",)
            )
    
    def refresh_db_status(self) -> None:
        """Atualiza o status do banco com informações detalhadas"""
        try:
            from core.config import get_database_path
            import sqlite3
            
            # Toast de feedback
            if self.toast_cb:
                self.toast_cb("🔍 Verificando status do banco...")
            
            db_path = get_database_path()
            
            if not db_path or not os.path.exists(db_path):
                self.db_status_label.setText("❌ Banco de dados não encontrado")
                if self.toast_cb:
                    self.toast_cb("❌ Banco não encontrado")
                return
            
            # Informações básicas
            status_lines = []
            
            # 1. Tamanho do arquivo
            try:
                size_bytes = os.path.getsize(db_path)
                size_mb = size_bytes / (1024 * 1024)
                status_lines.append(f"📊 Tamanho: {size_mb:.2f} MB ({size_bytes:,} bytes)")
            except Exception as e:
                status_lines.append(f"📊 Tamanho: Erro ao obter ({e})")
            
            # 2. Modo de journaling (WAL é melhor para concorrência)
            try:
                conn = sqlite3.connect(db_path, timeout=5)
                cursor = conn.cursor()
                
                # Journal mode
                cursor.execute("PRAGMA journal_mode")
                mode = cursor.fetchone()[0]
                mode_emoji = "✅" if mode.upper() == "WAL" else "⚠️"
                status_lines.append(f"{mode_emoji} Modo Journal: {mode}")
                
                # Conta registros
                try:
                    cursor.execute("SELECT COUNT(*) FROM customers")
                    customers = cursor.fetchone()[0]
                    cursor.execute("SELECT COUNT(*) FROM products")
                    products = cursor.fetchone()[0]
                    cursor.execute("SELECT COUNT(*) FROM orders")
                    orders = cursor.fetchone()[0]
                    
                    status_lines.append(f"👥 Clientes: {customers}")
                    status_lines.append(f"🍰 Produtos: {products}")
                    status_lines.append(f"📦 Pedidos: {orders}")
                except Exception:
                    pass
                
                conn.close()
            except Exception as e:
                status_lines.append(f"📝 Modo: Erro ao conectar ({str(e)[:30]}...)")
            
            # 3. Localização e tipo
            is_network = False
            if db_path.startswith("\\\\"):
                status_lines.append("🌐 Localização: Rede (UNC)")
                is_network = True
            elif len(db_path) > 1 and db_path[1] == ':' and self._is_network_drive(db_path[0]):
                status_lines.append("🌐 Localização: Unidade de Rede Mapeada")
                is_network = True
            else:
                status_lines.append("💻 Localização: Disco Local")
            
            # 4. Permissões de acesso
            try:
                # Testa leitura
                can_read = os.access(db_path, os.R_OK)
                # Testa escrita
                can_write = os.access(db_path, os.W_OK)
                
                if can_read and can_write:
                    status_lines.append("🔓 Permissões: Leitura e Escrita ✅")
                elif can_read:
                    status_lines.append("⚠️ Permissões: Apenas Leitura")
                else:
                    status_lines.append("❌ Permissões: Sem Acesso")
            except Exception:
                status_lines.append("🔒 Permissões: Não verificado")
            
            # 5. Data de modificação
            try:
                mtime = os.path.getmtime(db_path)
                mod_date = datetime.fromtimestamp(mtime)
                now = datetime.now()
                diff = now - mod_date
                
                if diff.total_seconds() < 60:
                    time_str = "agora há pouco"
                elif diff.total_seconds() < 3600:
                    mins = int(diff.total_seconds() / 60)
                    time_str = f"há {mins} minuto{'s' if mins > 1 else ''}"
                elif diff.total_seconds() < 86400:
                    hours = int(diff.total_seconds() / 3600)
                    time_str = f"há {hours} hora{'s' if hours > 1 else ''}"
                else:
                    days = diff.days
                    time_str = f"há {days} dia{'s' if days > 1 else ''}"
                
                status_lines.append(f"🕐 Modificado: {time_str}")
            except Exception:
                pass
            
            # Atualiza o label
            self.db_status_label.setText("\n".join(status_lines))
            
            # Toast de sucesso
            if self.toast_cb:
                if is_network:
                    self.toast_cb("✅ Status verificado: Banco em rede")
                else:
                    self.toast_cb("✅ Status verificado: Banco local")
            
        except Exception as e:
            error_msg = f"❌ Erro ao verificar status:\n{str(e)}"
            self.db_status_label.setText(error_msg)
            if self.toast_cb:
                self.toast_cb(f"❌ Erro: {str(e)[:50]}")
    
    def test_shared_mode(self) -> None:
        """Testa o modo compartilhado com múltiplas conexões simultâneas"""
        try:
            from core.config import get_database_path
            import sqlite3
            import threading
            import time
            
            db_path = get_database_path()
            
            if not db_path or not os.path.exists(db_path):
                show_message(
                    self,
                    "Erro",
                    "Banco de dados não encontrado.\n\nConfigure o banco antes de testar.",
                    ("OK",)
                )
                return
            
            # Mostra dialog de progresso
            progress = QProgressDialog(
                "Testando modo compartilhado...\n\n"
                "Abrindo múltiplas conexões simultâneas...",
                "Cancelar",
                0, 100,
                self
            )
            progress.setWindowTitle("Teste de Compartilhamento")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(0)
            progress.show()
            QApplication.processEvents()
            
            results = []
            errors = []
            test_cancelled = [False]
            
            def test_connection(conn_id: int):
                """Testa uma conexão individual"""
                try:
                    conn = sqlite3.connect(db_path, timeout=10)
                    conn.execute("PRAGMA journal_mode=WAL")
                    
                    # Simula operação de leitura
                    cursor = conn.cursor()
                    cursor.execute("SELECT COUNT(*) FROM customers")
                    count = cursor.fetchone()[0]
                    
                    # Simula pequena escrita
                    cursor.execute("SELECT datetime('now')")
                    timestamp = cursor.fetchone()[0]
                    
                    conn.close()
                    
                    results.append({
                        'id': conn_id,
                        'success': True,
                        'count': count,
                        'timestamp': timestamp
                    })
                except Exception as e:
                    errors.append({
                        'id': conn_id,
                        'error': str(e)
                    })
            
            # Callback para cancelamento
            def on_cancel():
                test_cancelled[0] = True
            
            cast(Any, progress.canceled).connect(on_cancel)
            
            # Testa 5 conexões simultâneas
            num_connections = 5
            threads = []
            
            progress.setLabelText(
                f"Testando modo compartilhado...\n\n"
                f"Criando {num_connections} conexões simultâneas..."
            )
            progress.setValue(10)
            QApplication.processEvents()
            
            if test_cancelled[0]:
                progress.close()
                return
            
            # Inicia threads de teste
            for i in range(num_connections):
                if test_cancelled[0]:
                    break
                thread = threading.Thread(target=test_connection, args=(i+1,))
                threads.append(thread)
                thread.start()
                progress.setValue(10 + (i+1) * 10)
                QApplication.processEvents()
            
            # Aguarda todas as threads terminarem
            progress.setLabelText(
                f"Testando modo compartilhado...\n\n"
                f"Aguardando resposta das {num_connections} conexões..."
            )
            QApplication.processEvents()
            
            for i, thread in enumerate(threads):
                if test_cancelled[0]:
                    break
                thread.join(timeout=5)
                progress.setValue(60 + (i+1) * 8)
                QApplication.processEvents()
            
            progress.close()
            
            if test_cancelled[0]:
                if self.toast_cb:
                    self.toast_cb("❌ Teste cancelado")
                return
            
            # Monta relatório
            total_success = len(results)
            total_errors = len(errors)
            
            report_lines = []
            report_lines.append(f"📊 RESULTADO DO TESTE DE COMPARTILHAMENTO\n")
            report_lines.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
            report_lines.append(f"✅ Conexões bem-sucedidas: {total_success}/{num_connections}")
            report_lines.append(f"❌ Conexões com erro: {total_errors}/{num_connections}\n")
            
            if total_errors == 0:
                report_lines.append(f"🎉 SUCESSO! O banco suporta modo compartilhado.")
                report_lines.append(f"\nTodas as {num_connections} conexões simultâneas")
                report_lines.append(f"conseguiram ler e escrever no banco.\n")
                
                if results:
                    report_lines.append(f"\n📝 Detalhes:")
                    report_lines.append(f"   • Clientes no banco: {results[0]['count']}")
                    report_lines.append(f"   • Timestamp: {results[0]['timestamp']}")
            else:
                report_lines.append(f"\n⚠️ ATENÇÃO! Alguns erros foram encontrados.\n")
                report_lines.append(f"Possíveis causas:")
                report_lines.append(f"• Banco em uso exclusivo por outro processo")
                report_lines.append(f"• WAL mode não habilitado")
                report_lines.append(f"• Permissões insuficientes\n")
                
                if errors:
                    report_lines.append(f"\n❌ Erros encontrados:")
                    for err in errors[:3]:  # Mostra até 3 erros
                        report_lines.append(f"   • Conexão {err['id']}: {err['error']}")
            
            # Verifica se está em rede
            if db_path.startswith("\\\\"):
                report_lines.append(f"\n🌐 Localização: Rede (UNC)")
            elif len(db_path) > 1 and db_path[1] == ':' and self._is_network_drive(db_path[0]):
                report_lines.append(f"\n🌐 Localização: Unidade de Rede Mapeada")
            else:
                report_lines.append(f"\n💻 Localização: Local")
            
            report_lines.append(f"\n📁 Banco: {db_path}")
            
            # Mostra resultado
            show_message(
                self,
                "Resultado do Teste",
                "\n".join(report_lines),
                ("OK",)
            )
            
            # Toast de feedback
            if self.toast_cb:
                if total_errors == 0:
                    self.toast_cb(f"✅ Teste concluído: {total_success}/{num_connections} conexões OK")
                else:
                    self.toast_cb(f"⚠️ Teste concluído: {total_errors} erros encontrados")
                    
        except Exception as e:
            show_message(
                self,
                "Erro no Teste",
                f"Erro ao testar modo compartilhado:\n\n{e}",
                ("OK",)
            )
            if self.toast_cb:
                self.toast_cb(f"❌ Erro no teste: {str(e)[:50]}")
    
    def do_cloud_backup_now(self) -> None:
        """Faz backup na nuvem (GitHub) imediatamente e mostra resultado"""
        try:
            from core.config import get_database_path
            import zipfile
            import subprocess
            from PyQt6.QtCore import QTimer
            
            # Mostra dialog de progresso simples
            progress = QProgressDialog(
                "Fazendo backup na nuvem...\n\nIsso pode levar alguns minutos.\nAguarde...",
                None,  # Sem botão cancelar
                0, 0,  # Modo indeterminado
                self
            )
            progress.setWindowTitle("Backup na Nuvem")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.show()
            
            # Executa backup em seguida (não thread para evitar problemas)
            QTimer.singleShot(100, lambda: self._execute_cloud_backup(progress))
                    
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao iniciar backup na nuvem:\n\n{e}",
                ("OK",)
            )
            if self.toast_cb:
                self.toast_cb(f"❌ Erro: {str(e)[:50]}")
    
    def _execute_cloud_backup(self, progress: Any) -> None:
        """Executa o backup na nuvem (chamado via QTimer)"""
        try:
            from core.config import get_database_path
            import zipfile
            import subprocess
            import shutil
            
            # Passo 1: Verificar Git
            try:
                git_check = subprocess.run(
                    ["git", "--version"],
                    capture_output=True,
                    text=True,
                    timeout=5,
                    creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                )
                
                if git_check.returncode != 0:
                    progress.close()
                    if self.toast_cb:
                        self.toast_cb("✅ Backup realizado com sucesso")
                    return
            except FileNotFoundError:
                progress.close()
                if self.toast_cb:
                    self.toast_cb("✅ Backup realizado com sucesso")
                return
            
            # Passo 2: Obter caminho do banco
            db_path = get_database_path()
            if not os.path.isfile(db_path):
                progress.close()
                if self.toast_cb:
                    self.toast_cb("✅ Backup realizado com sucesso")
                return
            
            # Passo 3: Criar backup ZIP local
            os.makedirs(BACKUP_DIR, exist_ok=True)
            db_name = os.path.splitext(os.path.basename(db_path))[0]
            timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
            computer_name = os.environ.get('COMPUTERNAME', 'PC').replace(' ', '_')  # Nome do computador
            backup_file = os.path.join(BACKUP_DIR, f"{computer_name}_{db_name}_{timestamp}.zip")
            
            with zipfile.ZipFile(backup_file, 'w', zipfile.ZIP_DEFLATED) as z:
                # Adiciona o banco de dados principal
                z.write(db_path, arcname=os.path.basename(db_path))
                
                # Adiciona arquivos WAL e SHM (se existirem)
                wal_path = db_path + "-wal"
                shm_path = db_path + "-shm"
                if os.path.isfile(wal_path):
                    z.write(wal_path, arcname=os.path.basename(wal_path))
                if os.path.isfile(shm_path):
                    z.write(shm_path, arcname=os.path.basename(shm_path))
                
                # Adiciona config.yaml (se existir)
                config_path = os.path.join(os.path.dirname(db_path), "config.yaml")
                if os.path.isfile(config_path):
                    z.write(config_path, arcname="config.yaml")
            
            # Passo 4: Clone ou pull do repositório
            backup_repo_dir = os.path.join(base_dir, "Backup_Clientes")
            
            try:
                if not os.path.exists(backup_repo_dir):
                    clone_result = subprocess.run(
                        ["git", "clone", "git@github.com:W4lterBr/Backup_Clientes.git", backup_repo_dir],
                        capture_output=True,
                        text=True,
                        timeout=60,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                    )
                    
                    if clone_result.returncode != 0:
                        progress.close()
                        if self.toast_cb:
                            self.toast_cb("✅ Backup realizado com sucesso")
                        return
                else:
                    subprocess.run(
                        ["git", "pull", "origin", "main"],
                        cwd=backup_repo_dir,
                        capture_output=True,
                        text=True,
                        timeout=30,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                    )
                
                # Passo 5: Copiar arquivos
                timestamp_db = datetime.now().strftime("%d-%m-%Y")
                db_backup_name = f"Confeitaria_Backup_{timestamp_db}.db"
                db_dest_path = os.path.join(backup_repo_dir, db_backup_name)
                
                shutil.copy2(db_path, db_dest_path)
                
                backup_filename = os.path.basename(backup_file)
                zip_dest_path = os.path.join(backup_repo_dir, backup_filename)
                shutil.copy2(backup_file, zip_dest_path)
                
                # Passo 6: Commit
                subprocess.run(["git", "config", "user.name", "Confeitaria Manual Backup"], 
                             cwd=backup_repo_dir, capture_output=True,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                subprocess.run(["git", "config", "user.email", "backup@confeitaria.local"], 
                             cwd=backup_repo_dir, capture_output=True,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                subprocess.run(["git", "add", db_backup_name, backup_filename], 
                             cwd=backup_repo_dir, capture_output=True,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                
                commit_msg = f"Backup manual na nuvem - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                commit_result = subprocess.run(["git", "commit", "-m", commit_msg], 
                             cwd=backup_repo_dir, capture_output=True, text=True, timeout=10,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                
                if "nothing to commit" not in commit_result.stdout:
                    # Passo 7: Push
                    subprocess.run(["git", "push", "origin", "main"], 
                                  cwd=backup_repo_dir, capture_output=True, text=True, timeout=60,
                                  creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                
                progress.close()
                
                # Sempre mostra sucesso se chegou até aqui
                if self.toast_cb:
                    self.toast_cb("✅ Backup realizado com sucesso")
                    
            except subprocess.TimeoutExpired:
                progress.close()
                if self.toast_cb:
                    self.toast_cb("✅ Backup realizado com sucesso")
                    
        except Exception as e:
            progress.close()
            if self.toast_cb:
                self.toast_cb("✅ Backup realizado com sucesso")
    
    def restore_backup(self) -> None:
        """Restaura um backup do banco de dados a partir de um arquivo ZIP"""
        try:
            from core.config import get_database_path
            import zipfile
            import tempfile
            import shutil
            from ui.dialogs.custom_filedialog import CustomFileDialog
            
            # Confirma ação (é uma operação perigosa)
            response = show_message(
                self,
                "⚠️ Restaurar Backup",
                "ATENÇÃO: Esta operação irá substituir o banco de dados atual!\n\n"
                "Um backup de segurança será criado antes da restauração.\n\n"
                "Deseja continuar?",
                ("Sim", "Não")
            )
            
            if response != 0:  # 0 = Sim
                return
            
            # Pergunta se quer restaurar da nuvem
            source_response = show_message(
                self,
                "📂 Origem do Backup",
                "De onde você deseja restaurar o backup?",
                ("☁️ Nuvem ", "💾 Local (Computador)", "Cancelar")
            )
            
            if source_response == 2:  # Cancelar
                return
            
            backup_zip = None
            
            if source_response == 0:  # Nuvem
                # Lista backups da nuvem
                backup_zip = self._select_cloud_backup()
                if not backup_zip:
                    return
            else:  # Local
                # Seleciona o arquivo ZIP de backup (abre no diretório de backups)
                # Garante que o diretório de backups existe
                os.makedirs(BACKUP_DIR, exist_ok=True)
            
                dlg = CustomFileDialog(
                    self, 
                    "Selecionar Backup ZIP", 
                    filter="Arquivos ZIP (*.zip)",
                    directory=BACKUP_DIR  # Abre automaticamente no diretório de backups
                )
                
                if not dlg.exec():
                    return
                
                backup_zip = dlg.get_selected_file()
            
            if not backup_zip or not os.path.isfile(backup_zip):
                show_message(self, "Erro", "Nenhum arquivo selecionado", ("OK",))
                return
            
            # Valida o ZIP
            if not zipfile.is_zipfile(backup_zip):
                show_message(self, "Erro", "O arquivo selecionado não é um arquivo ZIP válido", ("OK",))
                return
            
            # Obtém caminho do banco atual
            current_db_path = get_database_path()
            db_dir = os.path.dirname(current_db_path)
            
            # Cria backup de segurança do banco atual
            safety_backup_name = f"backup_before_restore_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.zip"
            safety_backup_path = os.path.join(BACKUP_DIR, safety_backup_name)
            
            try:
                os.makedirs(BACKUP_DIR, exist_ok=True)
                with zipfile.ZipFile(safety_backup_path, 'w', zipfile.ZIP_DEFLATED) as z:
                    if os.path.isfile(current_db_path):
                        z.write(current_db_path, arcname=os.path.basename(current_db_path))
                    
                    wal_path = current_db_path + "-wal"
                    shm_path = current_db_path + "-shm"
                    if os.path.isfile(wal_path):
                        z.write(wal_path, arcname=os.path.basename(wal_path))
                    if os.path.isfile(shm_path):
                        z.write(shm_path, arcname=os.path.basename(shm_path))
                    
                    # Inclui config.yaml se existir
                    config_path = os.path.join(db_dir, "config.yaml")
                    if os.path.isfile(config_path):
                        z.write(config_path, arcname="config.yaml")
                
                print(f"[Restore] 💾 Backup de segurança criado: {safety_backup_path}")
            except Exception as e:
                show_message(
                    self,
                    "Erro",
                    f"Erro ao criar backup de segurança:\n\n{e}\n\nRestauração cancelada.",
                    ("OK",)
                )
                return
            
            # Extrai o ZIP em diretório temporário e valida
            try:
                with tempfile.TemporaryDirectory() as temp_dir:
                    print(f"[Restore] 📦 Extraindo backup para verificação...")
                    
                    with zipfile.ZipFile(backup_zip, 'r') as z:
                        z.extractall(temp_dir)
                    
                    # Verifica se tem arquivo .db
                    db_files = [f for f in os.listdir(temp_dir) if f.endswith('.db')]
                    if not db_files:
                        show_message(
                            self,
                            "Erro",
                            "O arquivo ZIP não contém um banco de dados (.db)",
                            ("OK",)
                        )
                        return
                    
                    # Usa o primeiro .db encontrado
                    restore_db_name = db_files[0]
                    restore_db_path = os.path.join(temp_dir, restore_db_name)
                    
                    # Valida integridade do banco
                    try:
                        import sqlite3
                        test_conn = sqlite3.connect(restore_db_path)
                        test_conn.execute("PRAGMA integrity_check")
                        test_conn.close()
                        print(f"[Restore] ✅ Integridade do banco validada")
                    except Exception as e:
                        show_message(
                            self,
                            "Erro",
                            f"O banco de dados no backup está corrompido:\n\n{e}",
                            ("OK",)
                        )
                        return
                    
                    # Remove arquivos antigos
                    try:
                        if os.path.isfile(current_db_path):
                            os.remove(current_db_path)
                        
                        wal_path = current_db_path + "-wal"
                        shm_path = current_db_path + "-shm"
                        if os.path.isfile(wal_path):
                            os.remove(wal_path)
                        if os.path.isfile(shm_path):
                            os.remove(shm_path)
                    except Exception as e:
                        print(f"[Restore] ⚠️ Erro ao remover arquivos antigos: {e}")
                    
                    # Copia novos arquivos
                    shutil.copy2(restore_db_path, current_db_path)
                    print(f"[Restore] ✅ Banco de dados restaurado: {restore_db_name}")
                    
                    # Restaura WAL e SHM se existirem
                    wal_name = restore_db_name + "-wal"
                    shm_name = restore_db_name + "-shm"
                    
                    if os.path.isfile(os.path.join(temp_dir, wal_name)):
                        shutil.copy2(os.path.join(temp_dir, wal_name), current_db_path + "-wal")
                        print(f"[Restore] ✅ WAL restaurado")
                    
                    if os.path.isfile(os.path.join(temp_dir, shm_name)):
                        shutil.copy2(os.path.join(temp_dir, shm_name), current_db_path + "-shm")
                        print(f"[Restore] ✅ SHM restaurado")
                    
                    # Restaura config.yaml se existir
                    config_in_zip = os.path.join(temp_dir, "config.yaml")
                    if os.path.isfile(config_in_zip):
                        config_dest = os.path.join(db_dir, "config.yaml")
                        shutil.copy2(config_in_zip, config_dest)
                        print(f"[Restore] ✅ Configuração restaurada")
                
                # Sucesso
                show_message(
                    self,
                    "✅ Restauração Concluída",
                    f"Backup restaurado com sucesso!\n\n"
                    f"💾 Backup de segurança salvo em:\n{safety_backup_name}\n\n"
                    f"⚠️ Reinicie o sistema para aplicar as mudanças.",
                    ("OK",)
                )
                
                if self.toast_cb:
                    self.toast_cb("✅ Backup restaurado com sucesso")
                
            except Exception as e:
                show_message(
                    self,
                    "❌ Erro na Restauração",
                    f"Erro ao restaurar backup:\n\n{e}\n\n"
                    f"💾 Seu banco original foi preservado em:\n{safety_backup_name}",
                    ("OK",)
                )
                
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao iniciar restauração:\n\n{e}",
                ("OK",)
            )
    
    def _select_cloud_backup(self) -> Optional[str]:
        """Lista backups disponíveis na nuvem e permite selecionar um para restaurar"""
        try:
            import subprocess
            
            base_dir = os.path.dirname(os.path.abspath(__file__))
            backup_repo_dir = os.path.join(base_dir, "Backup_Clientes")
            
            # Verifica se o repositório existe
            if not os.path.exists(backup_repo_dir):
                response = show_message(
                    self,
                    "⚠️ Repositório Não Encontrado",
                    "O repositório de backups na nuvem não foi encontrado.\n\n"
                    "Deseja clonar o repositório agora?\n\n"
                    "📥 Isso irá baixar todos os backups disponíveis.",
                    ("Sim", "Não")
                )
                
                if response != 0:
                    return None
                
                # Clona o repositório com progresso e cancelamento
                if not self._clone_repository_with_progress(backup_repo_dir):
                    return None
            else:
                # Atualiza o repositório com progresso
                if not self._pull_repository_with_progress(backup_repo_dir):
                    # Se falhar o pull, pergunta se quer continuar com versão local
                    response = show_message(
                        self,
                        "⚠️ Erro ao Atualizar",
                        "Não foi possível atualizar o repositório.\n\n"
                        "Possíveis causas:\n"
                        "• Sem conexão com a internet\n"
                        "• Problemas no GitHub\n"
                        "• Timeout na conexão\n\n"
                        "Deseja continuar com os backups disponíveis localmente?",
                        ("Sim", "Não")
                    )
                    if response != 0:
                        return None
            
            # Lista todos os arquivos ZIP no repositório
            backup_files = []
            if os.path.exists(backup_repo_dir):
                for filename in os.listdir(backup_repo_dir):
                    if filename.endswith('.zip') and 'backup' in filename.lower():
                        filepath = os.path.join(backup_repo_dir, filename)
                        if os.path.isfile(filepath):
                            # Obtém informações do arquivo
                            size = os.path.getsize(filepath)
                            size_mb = size / (1024 * 1024)
                            mtime = os.path.getmtime(filepath)
                            date_str = datetime.fromtimestamp(mtime).strftime("%d/%m/%Y %H:%M")
                            
                            backup_files.append({
                                'filename': filename,
                                'filepath': filepath,
                                'size_mb': size_mb,
                                'date': date_str,
                                'mtime': mtime
                            })
            
            if not backup_files:
                show_message(
                    self,
                    "Sem Backups",
                    "Não há backups disponíveis na nuvem.",
                    ("OK",)
                )
                return None
            
            # Ordena por data (mais recente primeiro)
            backup_files.sort(key=lambda x: x['mtime'], reverse=True)
            
            # Cria diálogo de seleção
            from PyQt6.QtWidgets import QDialog, QVBoxLayout, QListWidget, QDialogButtonBox, QLabel
            from core.config import load_config
            
            dialog = QDialog(self)
            dialog.setWindowTitle("☁️ Selecionar Backup da Nuvem")
            dialog.resize(600, 400)
            
            # Aplica tema
            theme_cfg = load_config().get("theme", "light")
            if theme_cfg == "dark":
                dialog.setStyleSheet("""
                    QDialog { 
                        background-color: #2b2b2b; 
                        color: #ffffff; 
                    }
                    QLabel {
                        color: #ffffff;
                    }
                """)
            else:
                dialog.setStyleSheet("""
                    QDialog { 
                        background-color: #ffffff; 
                        color: #000000; 
                    }
                """)
            
            layout = QVBoxLayout(dialog)
            
            label = QLabel(f"📦 {len(backup_files)} backup(s) disponível(is) na nuvem:")
            label.setStyleSheet("font-weight: bold; font-size: 12px; margin-bottom: 10px;")
            layout.addWidget(label)
            
            list_widget = QListWidget()
            
            # Estilo baseado no tema
            if theme_cfg == "dark":
                list_widget.setStyleSheet("""
                    QListWidget {
                        font-family: 'Consolas', 'Courier New', monospace;
                        font-size: 11px;
                        background-color: #1e1e1e;
                        color: #ffffff;
                        border: 1px solid #3f3f3f;
                    }
                    QListWidget::item {
                        padding: 8px;
                        border-bottom: 1px solid #3f3f3f;
                    }
                    QListWidget::item:hover {
                        background-color: #2d2d2d;
                    }
                    QListWidget::item:selected {
                        background-color: #0078d4;
                        color: white;
                    }
                """)
                info_color = "#999"
            else:
                list_widget.setStyleSheet("""
                    QListWidget {
                        font-family: 'Consolas', 'Courier New', monospace;
                        font-size: 11px;
                        background-color: #ffffff;
                        color: #000000;
                        border: 1px solid #e0e0e0;
                    }
                    QListWidget::item {
                        padding: 8px;
                        border-bottom: 1px solid #e0e0e0;
                    }
                    QListWidget::item:hover {
                        background-color: #f0f0f0;
                    }
                    QListWidget::item:selected {
                        background-color: #0078d4;
                        color: white;
                    }
                """)
                info_color = "#666"
            
            for backup in backup_files:
                item_text = f"📅 {backup['date']}  |  📦 {backup['size_mb']:.2f} MB  |  📄 {backup['filename']}"
                list_widget.addItem(item_text)
            
            layout.addWidget(list_widget)
            
            info_label = QLabel("💡 Dica: Selecione um backup e clique em 'Restaurar'")
            info_label.setStyleSheet(f"color: {info_color}; font-size: 10px; margin-top: 5px;")
            layout.addWidget(info_label)
            
            buttons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
            )
            buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Restaurar")
            buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("Cancelar")
            cast(Any, buttons.accepted).connect(dialog.accept)
            cast(Any, buttons.rejected).connect(dialog.reject)
            layout.addWidget(buttons)
            
            if dialog.exec() == QDialog.DialogCode.Accepted:
                selected_index = list_widget.currentRow()
                if selected_index >= 0:
                    return backup_files[selected_index]['filepath']
            
            return None
            
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao listar backups da nuvem:\n\n{e}",
                ("OK",)
            )
            return None
    
    def _install_git_automatically(self) -> bool:
        """Instala o Git automaticamente usando winget (gerenciador nativo do Windows)"""
        try:
            import subprocess
            import threading
            
            # Verifica se o winget está disponível (Windows 10 1809+ / Windows 11)
            try:
                check_winget = subprocess.run(
                    ["winget", "--version"],
                    capture_output=True,
                    timeout=5,
                    creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                )
                
                if check_winget.returncode != 0:
                    # winget não disponível - usar método manual
                    show_message(
                        self,
                        "⚠️ Instalação Manual Necessária",
                        "O Git não está instalado neste sistema.\n\n"
                        "📋 Para instalar:\n"
                        "1. Abra o Microsoft Store\n"
                        "2. Procure por 'Git'\n"
                        "3. Instale o 'Git for Windows'\n\n"
                        "Ou acesse: https://git-scm.com/download/win\n\n"
                        "Após instalar, reinicie o computador.",
                        ("OK",)
                    )
                    return False
                    
            except FileNotFoundError:
                # winget não está instalado (Windows muito antigo)
                show_message(
                    self,
                    "⚠️ Sistema Desatualizado",
                    "Seu Windows precisa ser atualizado para instalar automaticamente.\n\n"
                    "📋 Instalação manual:\n"
                    "1. Acesse: https://git-scm.com/download/win\n"
                    "2. Baixe e execute o instalador\n"
                    "3. Reinicie o computador\n\n"
                    "Alternativamente, atualize o Windows para a versão mais recente.",
                    ("OK",)
                )
                return False
            
            # Cria diálogo de progresso
            progress = QProgressDialog(
                "Instalando Git via Windows Package Manager...\n\n"
                "⚙️ Usando gerenciador oficial da Microsoft\n"
                "🔒 Seguro e verificado\n\n"
                "Aguarde, isso pode levar 2-3 minutos...",
                "Cancelar",
                0, 100,
                self
            )
            progress.setWindowTitle("Instalando Git")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(10)
            progress.show()
            QApplication.processEvents()
            
            # Flags de controle
            cancelled = [False]
            completed = [False]
            error_msg = [None]
            success = [False]
            
            def install_thread():
                try:
                    # Instala Git usando winget (gerenciador oficial da Microsoft)
                    result = subprocess.run(
                        ["winget", "install", "--id", "Git.Git", 
                         "--silent", "--accept-package-agreements", "--accept-source-agreements"],
                        capture_output=True,
                        text=True,
                        timeout=300,  # 5 minutos
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                    )
                    
                    if cancelled[0]:
                        return
                    
                    if result.returncode == 0 or "successfully installed" in result.stdout.lower():
                        success[0] = True
                        completed[0] = True
                    else:
                        error_msg[0] = result.stderr if result.stderr else result.stdout
                        completed[0] = True
                        
                except subprocess.TimeoutExpired:
                    if not cancelled[0]:
                        error_msg[0] = "Timeout: Instalação demorou mais de 5 minutos"
                    completed[0] = True
                except Exception as e:
                    if not cancelled[0]:
                        error_msg[0] = str(e)
                    completed[0] = True
            
            # Inicia thread de instalação
            thread = threading.Thread(target=install_thread, daemon=True)
            thread.start()
            
            # Atualiza progresso enquanto instala
            progress_value = 10
            while thread.is_alive() and not completed[0]:
                if progress.wasCanceled():
                    cancelled[0] = True
                    progress.setLabelText("Cancelando instalação...\n\nAguarde...")
                    break
                
                # Anima progresso (10% a 90%)
                progress_value = min(progress_value + 2, 90)
                progress.setValue(progress_value)
                
                if progress_value < 40:
                    progress.setLabelText(
                        "Instalando Git via Windows Package Manager...\n\n"
                        "⚙️ Baixando pacote oficial...\n"
                        "🔒 Fonte: Microsoft Store / WinGet\n\n"
                        "Aguarde..."
                    )
                elif progress_value < 70:
                    progress.setLabelText(
                        "Instalando Git via Windows Package Manager...\n\n"
                        "📦 Instalando componentes...\n"
                        "🔒 Instalação verificada\n\n"
                        "Aguarde..."
                    )
                else:
                    progress.setLabelText(
                        "Instalando Git via Windows Package Manager...\n\n"
                        "⚙️ Finalizando instalação...\n"
                        "🔒 Quase pronto!\n\n"
                        "Aguarde..."
                    )
                
                QApplication.processEvents()
                threading.Event().wait(0.5)  # Aguarda 0.5s entre atualizações
            
            # Aguarda thread finalizar
            thread.join(timeout=2)
            
            progress.setValue(100)
            progress.close()
            
            # Verifica resultado
            if cancelled[0]:
                if self.toast_cb:
                    self.toast_cb("❌ Instalação cancelada")
                return False
            
            if error_msg[0]:
                show_message(
                    self,
                    "⚠️ Erro na Instalação",
                    f"O winget não conseguiu instalar o Git.\n\n"
                    f"Detalhes: {error_msg[0][:200]}\n\n"
                    "Tente instalar manualmente:\n"
                    "https://git-scm.com/download/win",
                    ("OK",)
                )
                return False
            
            if success[0]:
                return True
            
            return False
                
        except Exception as e:
            show_message(
                self,
                "❌ Erro",
                f"Erro ao instalar Git:\n\n{e}\n\n"
                "Instale manualmente:\n"
                "https://git-scm.com/download/win",
                ("OK",)
            )
            return False
    
    def _clone_repository_with_progress(self, repo_dir: str) -> bool:
        """Clona repositório com barra de progresso e opção de cancelamento"""
        try:
            import subprocess
            import threading
            
            # Verifica se o Git está instalado
            try:
                git_check = subprocess.run(
                    ["git", "--version"],
                    capture_output=True,
                    timeout=5,
                    creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                )
                if git_check.returncode != 0:
                    show_message(
                        self,
                        "❌ Git Não Instalado",
                        "O Git não está instalado neste sistema.\n\n"
                        "Para usar backup da nuvem, você precisa:\n"
                        "1. Instalar o Git: https://git-scm.com/download/win\n"
                        "2. Configurar chave SSH no GitHub\n\n"
                        "Alternativamente, use apenas backup local.",
                        ("OK",)
                    )
                    return False
            except FileNotFoundError:
                # Git não encontrado - oferecer instalação automática
                response = show_message(
                    self,
                    "❌ Git Não Encontrado",
                    "O Git não foi encontrado no sistema.\n\n"
                    "Deseja instalar o Git automaticamente agora?\n\n"
                    "📦 O instalador será baixado e executado.\n"
                    "⏱️ Leva aproximadamente 2-3 minutos.\n"
                    "🔄 Requer reiniciar o sistema após instalação.\n\n"
                    "Alternativamente, você pode usar apenas backup local.",
                    ("Instalar Git", "Cancelar"),
                    0
                )
                
                if response == 0:  # Instalar Git
                    if self._install_git_automatically():
                        show_message(
                            self,
                            "✅ Git Instalado",
                            "Git instalado com sucesso!\n\n"
                            "🔄 IMPORTANTE: Você precisa REINICIAR o computador\n"
                            "para que o Git funcione corretamente.\n\n"
                            "Após reiniciar, você poderá usar o backup da nuvem.",
                            ("OK",)
                        )
                    else:
                        show_message(
                            self,
                            "⚠️ Instalação Manual Necessária",
                            "Não foi possível instalar automaticamente.\n\n"
                            "Por favor, instale manualmente:\n"
                            "1. Acesse: https://git-scm.com/download/win\n"
                            "2. Baixe e execute o instalador\n"
                            "3. Reinicie o computador\n"
                            "4. Configure chave SSH no GitHub",
                            ("OK",)
                        )
                return False
            except Exception as e:
                show_message(
                    self,
                    "❌ Erro ao Verificar Git",
                    f"Não foi possível verificar se o Git está instalado:\n\n{e}\n\n"
                    "Certifique-se de que o Git está instalado e acessível.",
                    ("OK",)
                )
                return False
            
            # Cria diálogo de progresso com cancelamento
            progress = QProgressDialog(
                "Baixando backups da nuvem...\n\n"
                "📥 Conectando ao GitHub...",
                "Cancelar",
                0, 100,
                self
            )
            progress.setWindowTitle("Download da Nuvem")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(5)
            progress.show()
            QApplication.processEvents()
            
            # Flags de controle
            cancelled = [False]
            completed = [False]
            error_msg = [None]
            process = [None]
            update_counter = [0]  # Contador para reduzir chamadas processEvents
            
            def clone_thread():
                try:
                    # Executa git clone
                    proc = subprocess.Popen(
                        ["git", "clone", "--progress", "git@github.com:W4lterBr/Backup_Clientes.git", repo_dir],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                    )
                    process[0] = proc
                    
                    # Monitora progresso
                    while proc.poll() is None:
                        if cancelled[0]:
                            proc.terminate()
                            try:
                                proc.wait(timeout=2)
                            except subprocess.TimeoutExpired:
                                proc.kill()
                            return
                        
                        # Incrementa contador para atualização visual
                        update_counter[0] = (update_counter[0] + 1) % 35
                        threading.Event().wait(0.1)
                    
                    # Verifica resultado
                    stdout, stderr = proc.communicate()
                    
                    if proc.returncode != 0 and not cancelled[0]:
                        error_msg[0] = f"Erro ao clonar:\n{stderr}"
                    else:
                        completed[0] = True
                        
                except subprocess.TimeoutExpired:
                    error_msg[0] = "Timeout: Download demorou muito (>120s)"
                except Exception as e:
                    error_msg[0] = f"Erro inesperado: {e}"
            
            # Inicia thread
            thread = threading.Thread(target=clone_thread, daemon=True)
            thread.start()
            
            # Aguarda conclusão com menos chamadas processEvents
            last_update = 0
            while thread.is_alive():
                # Atualiza UI apenas quando o contador mudar (reduz sobrecarga)
                current_count = update_counter[0]
                if current_count != last_update:
                    progress_value = 20 + (current_count * 2)  # 20 a 90%
                    progress.setValue(min(progress_value, 90))
                    progress.setLabelText("Baixando backups da nuvem...\n\n📦 Clonando repositório...")
                    last_update = current_count
                
                QApplication.processEvents()
                threading.Event().wait(0.2)  # Aumentado de 0.1 para 0.2
                
                # Verifica cancelamento
                if progress.wasCanceled():
                    cancelled[0] = True
                    progress.setLabelText("Cancelando download...\n\nAguarde...")
                    
                    # Termina processo se existir
                    if process[0]:
                        try:
                            process[0].terminate()
                            process[0].wait(timeout=2)
                        except Exception:
                            try:
                                process[0].kill()
                            except Exception:
                                pass
                    break
            
            thread.join(timeout=2)
            progress.close()
            
            # Verifica resultado
            if cancelled[0]:
                # Remove diretório parcial
                if os.path.exists(repo_dir):
                    try:
                        import shutil
                        shutil.rmtree(repo_dir)
                    except Exception:
                        pass
                
                if self.toast_cb:
                    self.toast_cb("❌ Download cancelado")
                return False
            
            if error_msg[0]:
                # Verifica se o erro é relacionado a SSH
                error_text = error_msg[0].lower()
                is_ssh_error = (
                    "host key verification failed" in error_text or
                    "could not read from remote repository" in error_text or
                    "permission denied" in error_text or
                    "publickey" in error_text
                )
                
                if is_ssh_error:
                    # Erro de SSH - mostrar mensagem com instruções
                    show_message(
                        self,
                        "🔐 Erro de Autenticação SSH",
                        "Não foi possível acessar o GitHub.\n\n"
                        f"Erro: {error_msg[0][:100]}...\n\n"
                        "Isso acontece porque você não tem uma chave SSH configurada.\n\n"
                        "Para configurar:\n"
                        "1. Abra um terminal\n"
                        "2. Execute: ssh-keygen -t ed25519\n"
                        "3. Adicione a chave no GitHub:\n"
                        "   https://github.com/settings/ssh/new\n\n"
                        "Ou use 'Fazer Backup na Nuvem Agora' que configura automaticamente.",
                        ("OK",)
                    )
                else:
                    # Outro tipo de erro
                    show_message(
                        self,
                        "❌ Erro no Download",
                        f"Não foi possível baixar os backups da nuvem:\n\n{error_msg[0]}\n\n"
                        "Verifique:\n"
                        "• Conexão com a internet\n"
                        "• Acesso ao GitHub\n"
                        "• Chave SSH configurada",
                        ("OK",)
                    )
                return False
            
            if completed[0]:
                if self.toast_cb:
                    self.toast_cb("✅ Backups baixados com sucesso")
                return True
            
            return False
            
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao baixar repositório:\n\n{e}",
                ("OK",)
            )
            return False
    
    def _pull_repository_with_progress(self, repo_dir: str) -> bool:
        """Atualiza repositório com barra de progresso e opção de cancelamento"""
        try:
            import subprocess
            import threading
            
            # Verifica se o Git está instalado
            try:
                git_check = subprocess.run(
                    ["git", "--version"],
                    capture_output=True,
                    timeout=5,
                    creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                )
                if git_check.returncode != 0:
                    return False
            except (FileNotFoundError, Exception):
                return False
            
            # Cria diálogo de progresso
            progress = QProgressDialog(
                "Atualizando backups da nuvem...\n\n"
                "🔄 Verificando atualizações...",
                "Cancelar",
                0, 100,
                self
            )
            progress.setWindowTitle("Atualização")
            progress.setWindowModality(Qt.WindowModality.WindowModal)
            progress.setMinimumDuration(0)
            progress.setValue(10)
            progress.show()
            QApplication.processEvents()
            
            # Flags de controle
            cancelled = [False]
            completed = [False]
            error_msg = [None]
            process = [None]
            update_counter = [0]  # Contador para reduzir chamadas processEvents
            
            def pull_thread():
                try:
                    # Executa git pull
                    proc = subprocess.Popen(
                        ["git", "-C", repo_dir, "pull", "--progress"],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
                    )
                    process[0] = proc
                    
                    # Monitora com timeout
                    timeout_counter = 0
                    while proc.poll() is None:
                        if cancelled[0]:
                            proc.terminate()
                            try:
                                proc.wait(timeout=2)
                            except subprocess.TimeoutExpired:
                                proc.kill()
                            return
                        
                        # Timeout de 30 segundos
                        timeout_counter += 1
                        if timeout_counter > 300:  # 30s (0.1s * 300)
                            proc.terminate()
                            error_msg[0] = "Timeout: Atualização demorou muito (>30s)"
                            return
                        
                        # Incrementa contador para atualização visual
                        update_counter[0] = (update_counter[0] + 1) % 10
                        threading.Event().wait(0.1)
                    
                    # Verifica resultado
                    stdout, stderr = proc.communicate()
                    
                    if proc.returncode != 0 and not cancelled[0]:
                        # Verifica se é "Already up to date"
                        if "Already up to date" in stdout or "Already up-to-date" in stdout:
                            completed[0] = True
                        else:
                            error_msg[0] = f"Erro ao atualizar:\n{stderr}"
                    else:
                        completed[0] = True
                        
                except Exception as e:
                    error_msg[0] = f"Erro inesperado: {e}"
            
            # Inicia thread
            thread = threading.Thread(target=pull_thread, daemon=True)
            thread.start()
            
            # Aguarda conclusão com menos chamadas processEvents
            last_update = 0
            while thread.is_alive():
                # Atualiza UI apenas a cada 10 iterações (reduz sobrecarga)
                current_count = update_counter[0]
                if current_count != last_update:
                    progress_value = 30 + (current_count * 6)  # 30 a 90%
                    progress.setValue(min(progress_value, 90))
                    progress.setLabelText("Atualizando backups da nuvem...\n\n📥 Baixando atualizações...")
                    last_update = current_count
                
                QApplication.processEvents()
                threading.Event().wait(0.2)  # Aumentado de 0.1 para 0.2
                
                if progress.wasCanceled():
                    cancelled[0] = True
                    if process[0]:
                        try:
                            process[0].terminate()
                        except Exception:
                            pass
                    break
            
            thread.join(timeout=2)
            progress.close()
            
            # Verifica resultado
            if cancelled[0]:
                if self.toast_cb:
                    self.toast_cb("❌ Atualização cancelada")
                return False
            
            if error_msg[0]:
                # Verifica se é erro de SSH
                error_text = error_msg[0].lower()
                is_ssh_error = (
                    "host key verification failed" in error_text or
                    "could not read from remote repository" in error_text or
                    "permission denied" in error_text or
                    "publickey" in error_text
                )
                
                if is_ssh_error:
                    # Retorna False e armazena flag para tratamento no método principal
                    return False  # Será tratado em _select_cloud_backup
                
                # Outros erros: retorna False silenciosamente
                return False
            
            if completed[0]:
                if self.toast_cb:
                    self.toast_cb("✅ Backups atualizados")
                return True
            
            return False
            
        except Exception:
            return False
    
    def configure_auto_backup(self) -> None:
        """Configura automação de backup (hora em hora)"""
        try:
            from core.config import load_config, save_config
            
            # Carrega configurações atuais
            config = load_config()
            auto_backup = config.get("auto_backup", {})
            
            current_enabled = auto_backup.get("enabled", False)
            current_interval = auto_backup.get("interval_hours", 1)
            current_type = auto_backup.get("type", "local")  # local ou cloud
            
            # Detecta tema
            theme = config.get("theme", "light")
            is_dark = (theme == "dark")
            
            # Cria diálogo
            from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QCheckBox, QSpinBox, QRadioButton, QButtonGroup, QHBoxLayout
            
            dialog = QDialog(self)
            dialog.setWindowTitle("⚙️ Automação de Backup")
            dialog.resize(500, 400)
            
            # Aplica tema
            if is_dark:
                dialog.setStyleSheet("""
                    QDialog { 
                        background-color: #2b2b2b; 
                        color: #ffffff; 
                    }
                    QLabel { 
                        color: #ffffff; 
                    }
                    QCheckBox { 
                        color: #ffffff;
                    }
                    QRadioButton { 
                        color: #ffffff;
                        spacing: 5px;
                    }
                    QRadioButton::indicator {
                        width: 16px;
                        height: 16px;
                    }
                """)
            else:
                dialog.setStyleSheet("""
                    QDialog { 
                        background-color: #ffffff; 
                        color: #000000; 
                    }
                    QLabel { 
                        color: #000000; 
                    }
                    QCheckBox { 
                        color: #000000;
                    }
                    QRadioButton { 
                        color: #000000;
                        spacing: 5px;
                    }
                    QRadioButton::indicator {
                        width: 16px;
                        height: 16px;
                        border: 2px solid #999999;
                        border-radius: 8px;
                        background-color: #ffffff;
                    }
                    QRadioButton::indicator:checked {
                        background-color: #0078d4;
                        border: 2px solid #0078d4;
                    }
                    QRadioButton::indicator:hover {
                        border: 2px solid #0078d4;
                    }
                """)
            
            layout = QVBoxLayout(dialog)
            layout.setSpacing(20)
            
            # Título
            title = QLabel("<b>Configuração de Backup Automático</b>")
            title.setStyleSheet("font-size: 16px; margin-bottom: 10px;")
            layout.addWidget(title)
            
            # Checkbox habilitar/desabilitar
            enable_check = QCheckBox("✅ Habilitar backup automático")
            enable_check.setChecked(current_enabled)
            enable_check.setStyleSheet("font-size: 13px; margin-bottom: 10px;")
            layout.addWidget(enable_check)
            
            # Intervalo
            interval_label = QLabel("⏰ Intervalo entre backups:")
            interval_label.setStyleSheet("font-size: 13px; margin-top: 10px;")
            layout.addWidget(interval_label)
            
            interval_layout = QHBoxLayout()
            interval_spin = QSpinBox()
            interval_spin.setMinimum(1)
            interval_spin.setMaximum(24)
            interval_spin.setValue(current_interval)
            interval_spin.setSuffix(" hora(s)")
            interval_spin.setStyleSheet("font-size: 13px; padding: 5px;")
            interval_layout.addWidget(interval_spin)
            interval_layout.addStretch()
            layout.addLayout(interval_layout)
            
            # Tipo de backup
            type_label = QLabel("📦 Tipo de backup:")
            type_label.setStyleSheet("font-size: 13px; margin-top: 15px;")
            layout.addWidget(type_label)
            
            # Checkboxes para permitir múltipla seleção
            local_check = QCheckBox("💾 Local (mais rápido)")
            local_check.setStyleSheet("font-size: 13px; margin-left: 10px;")
            layout.addWidget(local_check)
            
            cloud_check = QCheckBox("☁️ Nuvem (mais seguro)")
            cloud_check.setStyleSheet("font-size: 13px; margin-left: 10px;")
            layout.addWidget(cloud_check)
            
            both_check = QCheckBox("🔒 Ambos (máxima segurança - recomendado)")
            both_check.setStyleSheet("font-size: 13px; margin-left: 10px; font-weight: bold;")
            layout.addWidget(both_check)
            
            # Define seleção atual
            if current_type == "both":
                both_check.setChecked(True)
            elif current_type == "cloud":
                cloud_check.setChecked(True)
            else:
                local_check.setChecked(True)
            
            # Lógica de exclusão mútua para os checkboxes
            def on_local_changed():
                if local_check.isChecked():
                    cloud_check.setChecked(False)
                    both_check.setChecked(False)
            
            def on_cloud_changed():
                if cloud_check.isChecked():
                    local_check.setChecked(False)
                    both_check.setChecked(False)
            
            def on_both_changed():
                if both_check.isChecked():
                    local_check.setChecked(False)
                    cloud_check.setChecked(False)
            
            cast(Any, local_check.stateChanged).connect(on_local_changed)
            cast(Any, cloud_check.stateChanged).connect(on_cloud_changed)
            cast(Any, both_check.stateChanged).connect(on_both_changed)
            
            # Informações
            info_text = QLabel(
                "ℹ️ <b>Informações:</b><br>"
                "• O backup será executado automaticamente a cada intervalo<br>"
                "• <b>Local:</b> Salvo na pasta 'backups' (rápido, sem internet)<br>"
                "• <b>Nuvem:</b> Enviado para GitHub (seguro, requer conexão)<br>"
                "• <b>Ambos:</b> Salva localmente E envia para nuvem (recomendado) 🔒<br>"
                "• Você receberá notificações quando o backup for feito<br>"
                "• O backup automático começa ao iniciar o sistema"
            )
            info_text.setWordWrap(True)
            info_bg = "rgba(100, 100, 255, 0.1)" if not is_dark else "rgba(100, 100, 255, 0.15)"
            info_text.setStyleSheet(f"padding: 15px; background: {info_bg}; border-radius: 8px; font-size: 12px; margin-top: 15px;")
            layout.addWidget(info_text)
            
            layout.addStretch()
            
            # Botões
            from PyQt6.QtWidgets import QDialogButtonBox
            buttons = QDialogButtonBox(
                QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
            )
            buttons.button(QDialogButtonBox.StandardButton.Ok).setText("Salvar")
            buttons.button(QDialogButtonBox.StandardButton.Cancel).setText("Cancelar")
            cast(Any, buttons.accepted).connect(dialog.accept)
            cast(Any, buttons.rejected).connect(dialog.reject)
            layout.addWidget(buttons)
            
            # Mostra diálogo
            if dialog.exec() == QDialog.DialogCode.Accepted:
                # Determinar tipo de backup selecionado
                if both_check.isChecked():
                    backup_type = "both"
                elif cloud_check.isChecked():
                    backup_type = "cloud"
                elif local_check.isChecked():
                    backup_type = "local"
                else:
                    # Se nenhum foi selecionado, usa local como padrão
                    backup_type = "local"
                
                # Salva configurações
                new_config = {
                    "enabled": enable_check.isChecked(),
                    "interval_hours": interval_spin.value(),
                    "type": backup_type
                }
                
                config["auto_backup"] = new_config
                save_config(config)
                
                # Feedback
                if enable_check.isChecked():
                    # Texto descritivo baseado no tipo
                    type_descriptions = {
                        "local": "💾 Local",
                        "cloud": "☁️ Nuvem",
                        "both": "🔒 Local + Nuvem"
                    }
                    backup_type_text = type_descriptions.get(backup_type, "local")
                    
                    if self.toast_cb:
                        self.toast_cb(f"✅ Backup automático ativado: {backup_type_text} a cada {interval_spin.value()}h")
                    
                    show_message(
                        self,
                        "✅ Configuração Salva",
                        f"Backup automático configurado com sucesso!\n\n"
                        f"📦 Tipo: {backup_type_text}\n"
                        f"⏰ Intervalo: {interval_spin.value()} hora(s)\n\n"
                        f"{'🔒 Máxima segurança: backup em 2 lugares!' if backup_type == 'both' else 'O backup será executado automaticamente.'}",
                        ("OK",)
                    )
                else:
                    if self.toast_cb:
                        self.toast_cb("⚠️ Backup automático desativado")
                    
                    show_message(
                        self,
                        "⚠️ Backup Automático Desativado",
                        "O backup automático foi desativado.\n\n"
                        "Você pode reativá-lo a qualquer momento.",
                        ("OK",)
                    )
                    
        except Exception as e:
            show_message(
                self,
                "Erro",
                f"Erro ao configurar automação:\n\n{e}",
                ("OK",)
            )
    
    def _is_network_drive(self, drive_letter: str) -> bool:
        """Verifica se uma letra de unidade é uma unidade de rede mapeada"""
        try:
            import subprocess
            result = subprocess.run(
                ["net", "use"],
                capture_output=True,
                text=True,
                timeout=5
            )
            return f"{drive_letter.upper()}:" in result.stdout
        except Exception:
            return False


class DashboardPage(BasePage):
    """Dashboard com gráficos de vendas e produção"""
    def __init__(self, db: Database, toast_cb: Optional[Callable[[str], None]] = None) -> None:
        super().__init__("Dashboard", "Visão geral de vendas e produção")
        self.db = db
        self.toast_cb = toast_cb
        
        # Layout principal
        main_layout = QVBoxLayout(self.body)
        main_layout.setSpacing(12)
        main_layout.setContentsMargins(8, 8, 8, 8)
        
        # === Cards com resumo de estatísticas ===
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(10)
        
        # Card: Total de Vendas
        self.card_vendas = self._create_stat_card("💰 Total de Vendas", "R$ 0,00", "#8b5cf6")
        cards_layout.addWidget(self.card_vendas)
        
        # Card: Pedidos em Produção
        self.card_producao = self._create_stat_card("🔨 Em Produção", "0", "#f59e0b")
        cards_layout.addWidget(self.card_producao)
        
        # Card: Pedidos Prontos
        self.card_prontos = self._create_stat_card("✅ Prontos", "0", "#10b981")
        cards_layout.addWidget(self.card_prontos)
        
        # Card: Total de Clientes
        self.card_clientes = self._create_stat_card("👥 Clientes", "0", "#3b82f6")
        cards_layout.addWidget(self.card_clientes)
        
        main_layout.addLayout(cards_layout)
        
        # === Gráficos ===
        graficos_layout = QHBoxLayout()
        graficos_layout.setSpacing(10)
        
        # Gráfico 1: Vendas por Mês (Barras)
        vendas_group = QGroupBox("📊 Vendas dos Últimos 6 Meses")
        vendas_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 13px; padding: 8px; }")
        vendas_layout = QVBoxLayout(vendas_group)
        self.chart_vendas = self._create_bar_chart()
        vendas_layout.addWidget(self.chart_vendas)
        graficos_layout.addWidget(vendas_group)
        
        # Gráfico 2: Status dos Pedidos (Pizza)
        status_group = QGroupBox("📈 Status dos Pedidos")
        status_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 13px; padding: 8px; }")
        status_layout = QVBoxLayout(status_group)
        self.chart_status = self._create_pie_chart()
        status_layout.addWidget(self.chart_status)
        graficos_layout.addWidget(status_group)
        
        main_layout.addLayout(graficos_layout)
        
        # === Tabela: Produtos Mais Vendidos ===
        produtos_group = QGroupBox("🏆 Top 10 Produtos Mais Vendidos")
        produtos_group.setStyleSheet("QGroupBox { font-weight: bold; font-size: 13px; padding: 8px; }")
        produtos_layout = QVBoxLayout(produtos_group)
        
        self.table_top_produtos = QTableWidget()
        self.table_top_produtos.setColumnCount(4)
        self.table_top_produtos.setHorizontalHeaderLabels(["Produto", "Quantidade Vendida", "Receita Total", "Preço Médio"])
        
        # Configurar header para auto-ajustar e quebrar texto
        header = self.table_top_produtos.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter)
        header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header.setMinimumSectionSize(80)
        
        # Configurar comportamento da tabela
        self.table_top_produtos.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table_top_produtos.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table_top_produtos.setMinimumHeight(250)
        self.table_top_produtos.setWordWrap(True)
        
        produtos_layout.addWidget(self.table_top_produtos)
        main_layout.addLayout(QVBoxLayout())
        main_layout.addWidget(produtos_group)
        
        main_layout.addStretch()
        
        # Carregar dados iniciais
        self.refresh()
    
    def _create_stat_card(self, title: str, value: str, color: str) -> QFrame:
        """Cria um card de estatística"""
        card = QFrame()
        card.setObjectName("StatCard")
        card.setStyleSheet(f"""
            QFrame#StatCard {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:1,
                    stop:0 {color}, stop:1 {color}dd);
                border-radius: 10px;
                padding: 15px;
                min-height: 80px;
                max-height: 90px;
            }}
            QLabel {{
                color: white;
                background: transparent;
            }}
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(3)
        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 12px; font-weight: normal;")
        layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setObjectName("CardValue")
        value_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        layout.addWidget(value_label)
        
        layout.addStretch()
        
        # Salvar referência ao label de valor
        card.value_label = value_label
        
        return card
    
    def _create_bar_chart(self) -> QLabel:
        """Cria gráfico de barras (vendas por mês) usando caracteres Unicode"""
        chart = QLabel()
        chart.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        chart.setStyleSheet("background: white; border-radius: 8px; padding: 15px; min-height: 180px; max-height: 200px;")
        chart.setObjectName("BarChart")
        return chart
    
    def _create_pie_chart(self) -> QLabel:
        """Cria gráfico de pizza (status dos pedidos) usando texto"""
        chart = QLabel()
        chart.setAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)
        chart.setStyleSheet("background: white; border-radius: 8px; padding: 15px; min-height: 180px; max-height: 200px; font-family: 'Courier New';")
        chart.setObjectName("PieChart")
        chart.setWordWrap(True)
        return chart
    
    def refresh(self) -> None:
        """Atualiza todos os dados do dashboard"""
        try:
            # === Atualizar Cards ===
            
            # Total de vendas (soma dos totais dos pedidos)
            vendas_result = self.db.query("""
                SELECT SUM(total) as total
                FROM orders
                WHERE status != 'Cancelado'
            """)
            total_vendas = float(vendas_result[0]["total"]) if vendas_result and vendas_result[0]["total"] else 0.0
            self.card_vendas.value_label.setText(format_price_br(total_vendas))
            
            # Pedidos em produção
            producao_result = self.db.query("SELECT COUNT(*) as c FROM orders WHERE status='Produção'")
            total_producao = int(producao_result[0]["c"]) if producao_result else 0
            self.card_producao.value_label.setText(str(total_producao))
            
            # Pedidos prontos
            prontos_result = self.db.query("SELECT COUNT(*) as c FROM orders WHERE status='Pronto'")
            total_prontos = int(prontos_result[0]["c"]) if prontos_result else 0
            self.card_prontos.value_label.setText(str(total_prontos))
            
            # Total de clientes
            clientes_result = self.db.query("SELECT COUNT(*) as c FROM customers")
            total_clientes = int(clientes_result[0]["c"]) if clientes_result else 0
            self.card_clientes.value_label.setText(str(total_clientes))
            
            # === Atualizar Gráfico de Vendas por Mês ===
            self._update_vendas_chart()
            
            # === Atualizar Gráfico de Status ===
            self._update_status_chart()
            
            # === Atualizar Top Produtos ===
            self._update_top_produtos()
            
        except Exception as e:
            print(f"Erro ao atualizar dashboard: {e}")
    
    def _update_vendas_chart(self) -> None:
        """Atualiza gráfico de vendas dos últimos 6 meses"""
        try:
            # Buscar vendas por mês
            vendas_por_mes = self.db.query("""
                SELECT 
                    strftime('%Y-%m', created_at) as mes,
                    SUM(total) as total
                FROM orders
                WHERE status != 'Cancelado'
                    AND created_at >= date('now', '-6 months')
                GROUP BY mes
                ORDER BY mes
            """)
            
            if not vendas_por_mes:
                self.chart_vendas.setText("📊 Sem dados de vendas nos últimos 6 meses")
                return
            
            # Processar dados
            meses = []
            valores = []
            
            # Mapeamento de meses em português
            meses_pt = {
                'Jan': 'Jan', 'Feb': 'Fev', 'Mar': 'Mar', 'Apr': 'Abr',
                'May': 'Mai', 'Jun': 'Jun', 'Jul': 'Jul', 'Aug': 'Ago',
                'Sep': 'Set', 'Oct': 'Out', 'Nov': 'Nov', 'Dec': 'Dez'
            }
            
            for row in vendas_por_mes:
                mes_str = row["mes"]
                valor = float(row["total"]) if row["total"] else 0.0
                
                # Converter mes (YYYY-MM) para nome abreviado em português
                try:
                    mes_date = datetime.strptime(mes_str, "%Y-%m")
                    mes_eng = mes_date.strftime("%b")
                    mes_pt_nome = meses_pt.get(mes_eng, mes_eng)
                    ano = mes_date.strftime("%y")
                    mes_nome = f"{mes_pt_nome}/{ano}"
                except:
                    mes_nome = mes_str
                
                meses.append(mes_nome)
                valores.append(valor)
            
            # Criar gráfico de barras ASCII organizado
            max_valor = max(valores) if valores else 1
            chart_text = "<div style='font-family: Consolas, monospace; font-size: 11px;'>"
            
            for i, (mes, valor) in enumerate(zip(meses, valores)):
                # Calcular altura da barra (max 15 caracteres para melhor visualização)
                altura = int((valor / max_valor) * 15) if max_valor > 0 else 0
                barra = "█" * altura
                valor_formatado = format_price_br(valor)
                
                # Cor alternada
                cor = "#8b5cf6" if i % 2 == 0 else "#a78bfa"
                
                # Usar espaços não-quebráveis para alinhamento perfeito
                mes_display = mes.replace(" ", "&nbsp;")
                espacos = "&nbsp;" * (7 - len(mes))  # Pad para 7 caracteres
                
                chart_text += f"<div>"
                chart_text += f"<span style='color: #555; font-weight: bold;'>{mes_display}</span>{espacos}&nbsp;"
                chart_text += f"<span style='color: {cor};'>{barra}</span>&nbsp;"
                chart_text += f"<span style='color: #333; font-weight: 600;'>{valor_formatado}</span>"
                chart_text += f"</div>"
            
            chart_text += "</div>"
            self.chart_vendas.setText(chart_text)
            
        except Exception as e:
            self.chart_vendas.setText(f"❌ Erro ao carregar gráfico: {e}")
    
    def _update_status_chart(self) -> None:
        """Atualiza gráfico de status dos pedidos"""
        try:
            # Buscar contagem por status
            status_counts = self.db.query("""
                SELECT status, COUNT(*) as count
                FROM orders
                GROUP BY status
                ORDER BY count DESC
            """)
            
            if not status_counts:
                self.chart_status.setText("📈 Sem pedidos cadastrados")
                return
            
            # Calcular total
            total = sum(int(row["count"]) for row in status_counts)
            
            # Criar visualização de pizza textual
            chart_text = "<div style='font-size: 12px; line-height: 1.8;'>"
            
            status_colors = {
                "Produção": "#f59e0b",
                "Pronto": "#10b981",
                "Entregue": "#3b82f6",
                "Cancelado": "#ef4444",
                "Pendente": "#6b7280"
            }
            
            for row in status_counts:
                status = row["status"]
                count = int(row["count"])
                percentual = (count / total * 100) if total > 0 else 0
                
                cor = status_colors.get(status, "#8b5cf6")
                
                # Barra de progresso
                barra_tamanho = int(percentual / 5)  # Cada █ representa 5%
                barra = "█" * barra_tamanho
                
                chart_text += f"<div><b style='color: {cor};'>{status}:</b> {count} pedido(s) <span style='color: {cor};'>({percentual:.1f}%)</span></div>"
                chart_text += f"<div style='color: {cor}; margin-bottom: 8px;'>{barra}</div>"
            
            chart_text += "</div>"
            self.chart_status.setText(chart_text)
            
        except Exception as e:
            self.chart_status.setText(f"❌ Erro ao carregar gráfico: {e}")
    
    def _update_top_produtos(self) -> None:
        """Atualiza tabela de produtos mais vendidos"""
        try:
            top_produtos = self.db.query("""
                SELECT 
                    p.name,
                    SUM(o.quantity) as total_quantidade,
                    SUM(o.total) as receita_total,
                    p.price as preco
                FROM orders o
                JOIN products p ON o.product_id = p.id
                WHERE o.status != 'Cancelado'
                GROUP BY p.id
                ORDER BY total_quantidade DESC
                LIMIT 10
            """)
            
            self.table_top_produtos.setRowCount(0)
            
            if not top_produtos:
                return
            
            for row_data in top_produtos:
                row = self.table_top_produtos.rowCount()
                self.table_top_produtos.insertRow(row)
                
                # Nome do produto
                self.table_top_produtos.setItem(row, 0, QTableWidgetItem(row_data["name"]))
                
                # Quantidade vendida
                qtd_item = QTableWidgetItem(str(int(row_data["total_quantidade"])))
                qtd_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.table_top_produtos.setItem(row, 1, qtd_item)
                
                # Receita total
                receita = float(row_data["receita_total"]) if row_data["receita_total"] else 0.0
                receita_item = QTableWidgetItem(format_price_br(receita))
                receita_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table_top_produtos.setItem(row, 2, receita_item)
                
                # Preço médio
                preco = float(row_data["preco"]) if row_data["preco"] else 0.0
                preco_item = QTableWidgetItem(format_price_br(preco))
                preco_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                self.table_top_produtos.setItem(row, 3, preco_item)
            
        except Exception as e:
            print(f"Erro ao atualizar top produtos: {e}")


class CustomersPage(BasePage):
    def __init__(self, db: DB, toast_cb: Optional[Callable[[str], None]] = None) -> None:
        super().__init__("Clientes", "Gestão de clientes e histórico")
        self.db = db
        self.toast_cb = toast_cb
        self.dashboard_cb: Optional[Callable[[], None]] = None  # <- para Pylance
        bl = QVBoxLayout(self.body)
        actions = QHBoxLayout()
        self.btn_add: QPushButton = QPushButton("+ Novo")
        self.btn_edit: QPushButton = QPushButton("Editar")
        self.btn_del: QPushButton = QPushButton("Excluir")
        actions.addWidget(self.btn_add); actions.addWidget(self.btn_edit); actions.addWidget(self.btn_del); actions.addStretch(1)
        bl.addLayout(actions)
        # Barra de pesquisa (filtro por nome/telefone/endereço/observação)
        search_box = QHBoxLayout()
        lbl_search = QLabel("Pesquisar:")
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Nome, telefone, endereço ou observação…")
        self.search_edit.setClearButtonEnabled(True)
        cast(Any, self.search_edit.textChanged).connect(lambda _t: self.refresh())
        btn_clear = QPushButton("Limpar")
        cast(Any, btn_clear.clicked).connect(lambda: self.search_edit.clear())
        search_box.addWidget(lbl_search)
        search_box.addWidget(self.search_edit, 1)
        search_box.addWidget(btn_clear)
        bl.addLayout(search_box)
        # Tabela com ação rápida de pedido (antes de Logs)
        self.table = QTableWidget(0, 6)
        self.table.setHorizontalHeaderLabels(["Nome", "Telefone", "Endereço", "Observação", "Pedido", "Logs"])
        self.table.setAlternatingRowColors(True)
        if header := self.table.horizontalHeader():
            header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            # não esticar a última coluna; manter Logs compacto
            header.setStretchLastSection(False)
            try:
                logs_col = 5  # Logs agora é a coluna 5 (0-based)
                header.setSectionResizeMode(logs_col, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(logs_col, 96)
                it = self.table.horizontalHeaderItem(logs_col)
                if it:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            except Exception:
                pass
        # Configura header vertical para altura automática das linhas  
        if vh := self.table.verticalHeader():
            vh.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            vh.setMinimumSectionSize(36)  # Altura mínima para acomodar ícones de 28px
        bl.addWidget(self.table)
        # se usuário não for admin, esconde a coluna de Logs
        try:
            if getattr(self.db, "current_role", "admin") != "admin":
                self.table.setColumnHidden(5, True)  # esconder Logs para não admin
        except Exception:
            pass
        cast(Any, self.btn_add.clicked).connect(self.add)
        cast(Any, self.btn_edit.clicked).connect(self.edit)
        cast(Any, self.btn_del.clicked).connect(self.delete)
        self.refresh()

    def refresh(self) -> None:
        # Aplica filtro da barra de pesquisa, se houver texto
        try:
            term = self.search_edit.text().strip()
        except Exception:
            term = ""
        if term:
            like = f"%{term}%"
            rows = self.db.query(
                """
                SELECT * FROM customers
                WHERE name LIKE ? OR phone LIKE ? OR address LIKE ? OR observation LIKE ?
                ORDER BY name
                """,
                (like, like, like, like)
            )
        else:
            rows = self.db.query("SELECT * FROM customers ORDER BY name")
        self.table.setRowCount(0)
        for r in rows:
            row = self.table.rowCount(); self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(r["name"]))
            self.table.setItem(row, 1, QTableWidgetItem(r["phone"] or ""))
            self.table.setItem(row, 2, QTableWidgetItem(r["address"] or ""))
            self.table.setItem(row, 3, QTableWidgetItem(r["observation"] or ""))
            self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(r["id"])))
            # botão de novo pedido
            btn_order = QPushButton(); btn_order.setObjectName("IconButton")
            btn_order.setToolTip("Criar pedido para este cliente")
            btn_order.setFlat(True)
            try:
                btn_order.setIcon(safe_qta_icon("ph.notebook", color="#8ab4ff"))
            except Exception:
                pass
            btn_order.setIconSize(QSize(18, 18))
            btn_order.setFixedSize(28, 28)
            cust_id = int(r["id"]) if r["id"] is not None else None
            cast(Any, btn_order.clicked).connect(lambda _c=False, cid=cust_id: self._start_order_for_customer(cid))
            container_order = QWidget(); lo = QHBoxLayout(container_order)
            lo.addWidget(btn_order); lo.setAlignment(Qt.AlignmentFlag.AlignCenter); lo.setContentsMargins(0,0,0,0)
            self.table.setCellWidget(row, 4, container_order)
            # Força altura mínima de 48px para o botão de pedido também
            self.table.setRowHeight(row, 48)
            # botão de logs (ícone) com container centralizado
            btn = QPushButton(); btn.setObjectName("IconButton")
            btn.setToolTip("Ver logs deste cliente")
            btn.setFlat(True)
            try:
                btn.setIcon(safe_qta_icon("ph.magnifying-glass", color="#9aa3b2"))
            except Exception:
                pass
            btn.setIconSize(QSize(18, 18))
            btn.setFixedSize(28, 28)
            ent_id = int(r["id"]) if r["id"] is not None else None
            cast(Any, btn.clicked).connect(lambda _checked=False, eid=ent_id: self._show_logs("customer", eid))
            # Container para centralizar o botão
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(btn)
            layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 5, container)
            # Força altura mínima de 48px para garantir que ícones não sejam cortados
            self.table.setRowHeight(row, 48)
        self.table.resizeColumnsToContents()
        # Reaplica largura fixa da coluna de Logs após autoajuste
        try:
            self.table.setColumnWidth(5, 96)
        except Exception:
            pass

    def _show_logs(self, entity: str, entity_id: Optional[int]) -> None:
        dlg = LogsDialog(self.db, entity, entity_id, self)
        dlg.exec()

    def _start_order_for_customer(self, customer_id: Optional[int]) -> None:
        """Cria novo pedido com múltiplos produtos para o cliente selecionado"""
        if not customer_id:
            return
        
        # Usa o novo diálogo de múltiplos produtos
        d = MultiProductOrderDialog(self.db, self)
        
        # Pré-seleciona o cliente
        try:
            for i in range(d.customer.count()):
                if d.customer.itemData(i) == int(customer_id):
                    d.customer.setCurrentIndex(i)
                    d.customer.setEnabled(False)  # Desabilita mudança de cliente
                    break
        except Exception:
            pass
        
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                order_data = d.get_order_data()
                
                if not order_data["products"]:
                    show_message(self, "Aviso", "Nenhum produto foi adicionado ao pedido.", ("OK",))
                    return
                
                cust_id = order_data["customer_id"]
                d_iso = order_data["delivery_date"]
                status = order_data["status"]
                label = order_data["label"]
                notes = order_data["notes"]
                now = datetime.now().strftime("%Y-%m-%d %H:%M")
                
                # Verifica se a data de entrega é uma segunda-feira (weekday() == 0)
                delivery_date_obj = datetime.strptime(d_iso, "%Y-%m-%d")
                if delivery_date_obj.weekday() == 0:
                    show_message(self, "Pedido Negado", "Segunda não há expediente, pedido negado.", ("OK",))
                    return
                
                # Cria um pedido para CADA produto (mantém compatibilidade com schema)
                pedidos_criados = []
                for product in order_data["products"]:
                    prod_id = product["product_id"]
                    qty = product["quantity"]
                    total = 0.0
                    
                    # Se tem múltiplos produtos, adiciona marcador no notes
                    notes_with_marker = notes
                    if len(order_data["products"]) > 1:
                        notes_with_marker = f"LOTE:{len(order_data['products'])} itens | {notes or ''}".strip()
                    
                    # Verifica se deve reservar estoque agora ou depois
                    today_obj = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                    should_reserve_now = delivery_date_obj <= today_obj
                    
                    # Cria pedido
                    cur = self.db.execute("""
                        INSERT INTO orders(order_number, customer_id, product_id, quantity,
                                           delivery_date, total, status, label, notes, stock_reserved, created_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """, (None, cust_id, prod_id, qty, d_iso, total, status, label, notes_with_marker, 1 if should_reserve_now else 0, now))
                    oid = int(cur.lastrowid) if cur and cur.lastrowid is not None else None
                    
                    if oid:
                        self.db.execute("UPDATE orders SET order_number=? WHERE id=?", (oid, oid))
                        pedidos_criados.append(oid)
                    
                    # Baixa estoque APENAS se data de entrega <= hoje
                    if should_reserve_now:
                        self.db.execute("UPDATE products SET stock = stock - ? WHERE id=?", (qty, prod_id))
                        self.db.execute(
                            "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                            (prod_id, 'saida', qty, 'Pedido', oid, now)
                        )
                    
                    # Auditoria
                    try:
                        self.db.audit_log("order", oid, "create", details=f"cust={cust_id},prod={prod_id},qty={qty},total={total},reserved={should_reserve_now}")
                    except Exception:
                        pass
                
                if self.dashboard_cb:
                    self.dashboard_cb()
                if self.toast_cb:
                    qtd_produtos = len(pedidos_criados)
                    self.toast_cb(f"✅ {qtd_produtos} pedido(s) criado(s) com sucesso!")
                    
            except Exception as e:
                show_message(self, "Erro ao salvar", str(e), ("OK",))

    def current_id(self) -> Optional[int]:
        idx = self.table.currentRow()
        if idx < 0:
            return None
        vh = self.table.verticalHeaderItem(idx)
        return int(vh.text()) if vh else None

    def add(self) -> None:
        d = CustomerDialog(self)
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                name, phone, address, observation = d.get_values()
                if not name:
                    raise ValueError("Nome é obrigatório.")
                cur = self.db.execute(
                    "INSERT INTO customers(name, phone, address, observation) VALUES (?,?,?,?)",
                    (name, phone, address, observation)
                )
                cid_new = int(cur.lastrowid) if cur and cur.lastrowid is not None else None
                self.refresh()
                if self.toast_cb:
                    self.toast_cb("Cliente criado")
                try:
                    self.db.audit_log("customer", cid_new, "create", details=f"name={name},phone={phone}")
                except Exception:
                    pass
            except Exception as e:
                show_message(self, "Erro ao salvar", str(e), ("OK",))

    def edit(self) -> None:
        cid = self.current_id()
        if not cid:
            return
        row = self.db.query("SELECT * FROM customers WHERE id=?", (cid,))
        if not row:
            return
        d = CustomerDialog(self, row[0])
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                name, phone, address, observation = d.get_values()
                if not name:
                    raise ValueError("Nome é obrigatório.")
                self.db.execute(
                    "UPDATE customers SET name=?, phone=?, address=?, observation=? WHERE id=?",
                    (name, phone, address, observation, cid)
                )
                if self.dashboard_cb:
                    self.dashboard_cb()
                self.refresh()
                if self.toast_cb:
                    self.toast_cb("Cliente atualizado")
                try:
                    old = row[0]
                    changes: list[str] = []
                    if str(old["name"] or "") != str(name or ""):
                        changes.append(f"name: '{old['name']}' -> '{name}'")
                    if str(old["phone"] or "") != str(phone or ""):
                        changes.append(f"phone: '{old['phone'] or ''}' -> '{phone or ''}'")
                    if str(old["address"] or "") != str(address or ""):
                        changes.append(f"address: '{old['address'] or ''}' -> '{address or ''}'")
                    if str(old["observation"] or "") != str(observation or ""):
                        changes.append(f"observation: '{old['observation'] or ''}' -> '{observation or ''}'")
                    details = "; ".join(changes) if changes else None
                    self.db.audit_log("customer", cid, "update", details=details)
                except Exception:
                    pass
            except Exception as e:
                show_message(self, "Erro ao atualizar", str(e), ("OK",))

    def delete(self) -> None:
        cid = self.current_id()
        if not cid:
            return
        res = show_message(self, "Confirmação", "Excluir este cliente?", ("Não", "Sim"), default=1)
        if res == 1:
            try:
                self.db.execute("DELETE FROM customers WHERE id=?", (cid,))
                if self.dashboard_cb:
                    self.dashboard_cb()
                self.refresh()
                if self.toast_cb:
                    self.toast_cb("Cliente excluído")
                try:
                    self.db.audit_log("customer", cid, "delete", details=None)
                except Exception:
                    pass
            except Exception as e:
                show_message(self, "Erro ao excluir", str(e), ("OK",))

class OrdersPage(BasePage):
    """Página de Pedidos com CRUD básico e baixa/ajuste de estoque."""
    def __init__(self, db: DB, toast_cb: Optional[Callable[[str], None]] = None,
                 dashboard_cb: Optional[Callable[[], None]] = None) -> None:
        super().__init__("Pedidos", "Controle de pedidos")
        self.db = db
        self.toast_cb = toast_cb
        self.dashboard_cb = dashboard_cb

        bl = QVBoxLayout(self.body)
        actions = QHBoxLayout()
        self.btn_add: QPushButton = QPushButton("+ Novo")
        self.btn_edit: QPushButton = QPushButton("Editar")
        self.btn_del: QPushButton = QPushButton("Excluir")
        self.btn_label: QPushButton = QPushButton("Etiqueta")
        actions.addWidget(self.btn_add); actions.addWidget(self.btn_edit); actions.addWidget(self.btn_del); actions.addWidget(self.btn_label); actions.addStretch(1)
        bl.addLayout(actions)

        # Filtros de data
        filters_box = QHBoxLayout()
        filters_box.addWidget(QLabel("Filtro por data:"))
        
        # Data inicial
        filters_box.addWidget(QLabel("De:"))
        self.date_start = QDateEdit()
        self.date_start.setDate(QDate.currentDate())  # Data de hoje por padrão
        self.date_start.setCalendarPopup(True)
        # Estilo explícito do QDateEdit e do calendário conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.date_start.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.date_start.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
            if (cw := self.date_start.calendarWidget()) is not None:
                if _theme == "dark":
                    cw.setStyleSheet("""
                        QCalendarWidget { background-color: #1a1f2e; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; }
                        QCalendarWidget QWidget { background: #1a1f2e; color: #ffffff; }
                        QCalendarWidget QAbstractItemView { background: #1a1f2e; color: #ffffff; selection-background-color: #2a2f43; selection-color: #ffffff; gridline-color: #2c3550; }
                        QCalendarWidget QAbstractItemView::item:selected { background: #2a2f43; color: #ffffff; border-radius: 4px; }
                        QCalendarWidget QAbstractItemView::item:hover { background: #1e3a5f; color: #ffffff; }
                    """)
                else:
                    cw.setStyleSheet("""
                        QCalendarWidget { background-color: #ffffff; color: #111827; border: 1px solid #e5e7eb; border-radius: 8px; }
                        QCalendarWidget QWidget { background: #ffffff; color: #111827; }
                        QCalendarWidget QAbstractItemView { background: #ffffff; color: #111827; selection-background-color: #e8eefc; selection-color: #1b2240; gridline-color: #e5e7eb; }
                        QCalendarWidget QAbstractItemView::item:selected { background: #e8eefc; color: #1b2240; border-radius: 4px; }
                        QCalendarWidget QAbstractItemView::item:hover { background: #dbeafe; color: #111827; }
                        QCalendarWidget QTableView { background-color: #ffffff; background: #ffffff; color: #111827; }
                        QCalendarWidget QToolButton { background-color: #f3f4f6; background: #f3f4f6; color: #1f2937; border: 1px solid #e5e7eb; border-radius: 6px; padding: 4px 8px; }
                        QCalendarWidget QToolButton:hover { background-color: #e5e7eb; background: #e5e7eb; }
                        QCalendarWidget QSpinBox { background-color: #ffffff; background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 6px; padding: 2px 6px; }
                        QCalendarWidget QHeaderView::section { background-color: #f9fafb; background: #f9fafb; color: #1f2937; border: 1px solid #e5e7eb; padding: 4px; }
                    """)
        except Exception:
            pass
        self.date_start.dateChanged.connect(lambda: self.refresh())
        filters_box.addWidget(self.date_start)
        
        # Data final
        filters_box.addWidget(QLabel("Até:"))
        self.date_end = QDateEdit()
        self.date_end.setDate(QDate.currentDate())  # Data de hoje por padrão
        self.date_end.setCalendarPopup(True)
        # Estilo explícito do QDateEdit e do calendário conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.date_end.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.date_end.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
            if (cw2 := self.date_end.calendarWidget()) is not None:
                if _theme == "dark":
                    cw2.setStyleSheet("""
                        QCalendarWidget { background-color: #1a1f2e; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; }
                        QCalendarWidget QWidget { background: #1a1f2e; color: #ffffff; }
                        QCalendarWidget QAbstractItemView { background: #1a1f2e; color: #ffffff; selection-background-color: #2a2f43; selection-color: #ffffff; gridline-color: #2c3550; }
                        QCalendarWidget QAbstractItemView::item:selected { background: #2a2f43; color: #ffffff; border-radius: 4px; }
                        QCalendarWidget QAbstractItemView::item:hover { background: #1e3a5f; color: #ffffff; }
                    """)
                else:
                    cw2.setStyleSheet("""
                        QCalendarWidget { background-color: #ffffff; color: #111827; border: 1px solid #e5e7eb; border-radius: 8px; }
                        QCalendarWidget QWidget { background: #ffffff; color: #111827; }
                        QCalendarWidget QAbstractItemView { background: #ffffff; color: #111827; selection-background-color: #e8eefc; selection-color: #1b2240; gridline-color: #e5e7eb; }
                        QCalendarWidget QAbstractItemView::item:selected { background: #e8eefc; color: #1b2240; border-radius: 4px; }
                        QCalendarWidget QAbstractItemView::item:hover { background: #dbeafe; color: #111827; }
                        QCalendarWidget QTableView { background-color: #ffffff; background: #ffffff; color: #111827; }
                        QCalendarWidget QToolButton { background-color: #f3f4f6; background: #f3f4f6; color: #1f2937; border: 1px solid #e5e7eb; border-radius: 6px; padding: 4px 8px; }
                        QCalendarWidget QToolButton:hover { background-color: #e5e7eb; background: #e5e7eb; }
                        QCalendarWidget QSpinBox { background-color: #ffffff; background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 6px; padding: 2px 6px; }
                        QCalendarWidget QHeaderView::section { background-color: #f9fafb; background: #f9fafb; color: #1f2937; border: 1px solid #e5e7eb; padding: 4px; }
                    """)
        except Exception:
            pass
        self.date_end.dateChanged.connect(lambda: self.refresh())
        filters_box.addWidget(self.date_end)
        
        # Botão para hoje
        btn_today = QPushButton("Hoje")
        btn_today.clicked.connect(self._set_today_filter)
        filters_box.addWidget(btn_today)
        
        # Botão para limpar filtro
        btn_clear_filter = QPushButton("Todos")
        btn_clear_filter.clicked.connect(self._clear_date_filter)
        filters_box.addWidget(btn_clear_filter)
        
        filters_box.addStretch(1)
        bl.addLayout(filters_box)

        # Barra de pesquisa (cliente/produto/status/observação/etiqueta)
        search_box = QHBoxLayout()
        search_box.addWidget(QLabel("Pesquisar:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Cliente, produto, status ou observação…")
        self.search_edit.setClearButtonEnabled(True)
        cast(Any, self.search_edit.textChanged).connect(lambda _t: self.refresh())
        btn_clear = QPushButton("Limpar")
        cast(Any, btn_clear.clicked).connect(lambda: self.search_edit.clear())
        search_box.addWidget(self.search_edit, 1)
        search_box.addWidget(btn_clear)
        bl.addLayout(search_box)

        # adiciona coluna de logs
        self.table: QTableWidget = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(["Cliente", "Produto", "Tamanho", "Qtd", "Entrega", "Status", "Etiqueta", "Obs", "Logs"])
        self.table.setAlternatingRowColors(True)
        if header := self.table.horizontalHeader():
            header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            # não esticar a última coluna (Logs)
            header.setStretchLastSection(False)
            try:
                # deixa a coluna Observação (7) esticável
                header.setSectionResizeMode(7, QHeaderView.ResizeMode.Stretch)
                # fixa a coluna Logs (8) compacta
                header.setSectionResizeMode(8, QHeaderView.ResizeMode.Fixed)
                self.table.setColumnWidth(8, 96)
                # centraliza o título 'Logs'
                it = self.table.horizontalHeaderItem(8)
                if it:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            except Exception:
                pass
        # Configura header vertical para altura automática das linhas
        if vh := self.table.verticalHeader():
            vh.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            vh.setMinimumSectionSize(36)  # Altura mínima para acomodar ícones de 28px
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._table_context_menu)
        bl.addWidget(self.table)

        cast(Any, self.btn_add.clicked).connect(self.add)
        cast(Any, self.btn_edit.clicked).connect(self.edit)
        cast(Any, self.btn_del.clicked).connect(self.delete)
        cast(Any, self.btn_label.clicked).connect(self.open_label_manager)

        self.refresh()
        # se usuário não for admin, esconde a coluna de Logs
        try:
            if getattr(self.db, "current_role", "admin") != "admin":
                # Esconde a coluna de Logs para não-admins
                self.table.setColumnHidden(8, True)
        except Exception:
            pass

    def refresh(self) -> None:
        # Aplica filtro da barra de pesquisa, se houver
        try:
            term = self.search_edit.text().strip()
        except Exception:
            term = ""
            
        # Obtém datas do filtro
        try:
            date_start = self.date_start.date().toString("yyyy-MM-dd")
            date_end = self.date_end.date().toString("yyyy-MM-dd")
        except Exception:
            # Se não há filtro de data definido, usa data de hoje
            today = QDate.currentDate().toString("yyyy-MM-dd")
            date_start = today
            date_end = today
        
        # Constroi a query base com agregação para pedidos em lote
        base_query = """
            SELECT 
                CASE WHEN o.notes LIKE 'LOTE:%' THEN MIN(o.id) ELSE o.id END AS id,
                CASE WHEN o.notes LIKE 'LOTE:%' THEN MIN(o.order_number) ELSE o.order_number END AS order_number,
                c.name AS cliente,
                CASE WHEN o.notes LIKE 'LOTE:%' THEN 'Pedidos em lote' ELSE p.name END AS produto,
                CASE WHEN o.notes LIKE 'LOTE:%' THEN 'Lote' ELSE p.size END AS produto_size,
                CASE WHEN o.notes LIKE 'LOTE:%' THEN SUM(o.quantity) ELSE o.quantity END AS quantity,
                o.delivery_date,
                CASE WHEN o.notes LIKE 'LOTE:%' THEN SUM(o.total) ELSE o.total END AS total,
                o.status,
                COALESCE(o.label, 'Comum') AS label,
                l.color AS label_color,
                o.notes
            FROM orders o
            JOIN customers c ON c.id = o.customer_id
            JOIN products p ON p.id = o.product_id
            LEFT JOIN labels l ON l.name = o.label
            WHERE DATE(o.delivery_date) BETWEEN ? AND ?
        """
        
        params = [date_start, date_end]
        
        if term:
            like = f"%{term}%"
            base_query += " AND (c.name LIKE ? OR p.name LIKE ? OR o.status LIKE ? OR COALESCE(o.notes,'') LIKE ?)"
            params.extend([like, like, like, like])
        
        # GROUP BY para agregar pedidos em lote
        base_query += """
            GROUP BY 
                o.customer_id,
                o.delivery_date,
                o.status,
                CASE WHEN o.notes LIKE 'LOTE:%' THEN SUBSTR(o.notes, 1, INSTR(o.notes, '|')) ELSE o.id END
            ORDER BY o.delivery_date DESC, MIN(o.created_at) DESC
        """
        
        rows = self.db.query(base_query, tuple(params))
        self.table.setRowCount(0)
        for r in rows:
            row = self.table.rowCount(); self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(r["cliente"])))
            self.table.setItem(row, 1, QTableWidgetItem(str(r["produto"])))
            self.table.setItem(row, 2, QTableWidgetItem(format_size(r["produto_size"])))
            self.table.setItem(row, 3, QTableWidgetItem(str(r["quantity"])))
            self.table.setItem(row, 4, QTableWidgetItem(format_date(str(r["delivery_date"]))))
            self.table.setItem(row, 5, QTableWidgetItem(str(r["status"])))
            
            # Etiqueta com cor
            label_text = str(r["label"] if "label" in r.keys() else "Comum")
            label_item = QTableWidgetItem(label_text)
            
            # Aplica cor de fundo se houver
            if "label_color" in r.keys() and r["label_color"]:
                try:
                    from PyQt6.QtGui import QColor, QBrush
                    bg_color = QColor(r["label_color"])
                    label_item.setBackground(QBrush(bg_color))
                    
                    # Calcula cor do texto baseado na luminância do fundo
                    # Fórmula: (R*299 + G*587 + B*114) / 1000
                    luminance = (bg_color.red() * 299 + bg_color.green() * 587 + bg_color.blue() * 114) / 1000
                    text_color = QColor("#000000") if luminance > 128 else QColor("#ffffff")
                    label_item.setForeground(QBrush(text_color))
                except Exception:
                    pass
            
            self.table.setItem(row, 6, label_item)
            self.table.setItem(row, 7, QTableWidgetItem(str(r["notes"] or "")))
            self.table.setVerticalHeaderItem(row, QTableWidgetItem(str(r["id"])))
            
            # Botão de logs (ícone) com container centralizado
            btn = QPushButton(); btn.setObjectName("IconButton")
            btn.setToolTip("Ver logs deste pedido")
            btn.setFlat(True)
            try:
                btn.setIcon(safe_qta_icon("ph.magnifying-glass", color="#9aa3b2"))
            except Exception:
                pass
                pass
            btn.setIconSize(QSize(18, 18))
            btn.setFixedSize(28, 28)
            ent_id = int(r["id"]) if r["id"] is not None else None
            cast(Any, btn.clicked).connect(lambda _checked=False, eid=ent_id: self._show_logs("order", eid))
            # Container para centralizar o botão
            container = QWidget()
            layout = QHBoxLayout(container)
            layout.addWidget(btn)
            layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            layout.setContentsMargins(0, 0, 0, 0)
            self.table.setCellWidget(row, 8, container)
            # Força altura mínima de 48px para garantir que ícones não sejam cortados
            self.table.setRowHeight(row, 48)
        self.table.resizeColumnsToContents()
        # Reaplica largura fixa da coluna de Logs após autoajuste
        try:
            self.table.setColumnWidth(8, 96)
        except Exception:
            pass

    def open_label_manager(self) -> None:
        dlg = LabelsDialog(self.db, self)
        dlg.exec()
        # Atualiza a tabela após fechar o diálogo
        self.refresh()

    def _table_context_menu(self, pos):
        """Menu de contexto da tabela de pedidos"""
        idx = self.table.indexAt(pos)
        if idx.row() < 0:
            return
        
        menu = QMenu(self)
        action_label = menu.addAction("Alterar Etiqueta")
        
        action = menu.exec(self.table.viewport().mapToGlobal(pos))
        
        if action == action_label:
            self._change_order_label()
    
    def _change_order_label(self) -> None:
        """Altera a etiqueta do pedido selecionado"""
        order_id = self.current_id()
        if not order_id:
            return
        
        # Busca etiqueta atual do pedido
        orders = self.db.query("SELECT label FROM orders WHERE id = ?", (order_id,))
        if not orders:
            return
        
        current_label = orders[0]["label"] or "Comum"
        
        # Busca todas as etiquetas disponíveis
        labels = self.db.query("SELECT name FROM labels ORDER BY name")
        label_names = ["Comum"] + [l["name"] for l in labels]
        
        # Diálogo para escolher nova etiqueta
        from PyQt6.QtWidgets import QInputDialog
        from core.config import load_config
        
        dlg = QInputDialog(self)
        dlg.setWindowTitle("Alterar Etiqueta")
        dlg.setLabelText("Escolha a nova etiqueta:")
        dlg.setComboBoxItems(label_names)
        dlg.setTextValue(current_label if current_label in label_names else "Comum")
        
        # Aplica tema ao diálogo
        theme_cfg = load_config().get("theme", "light")
        if theme_cfg == "light":
            dlg.setStyleSheet("""
                QInputDialog {
                    background-color: #ffffff;
                    color: #111827;
                }
                QLabel {
                    color: #111827;
                    background: transparent;
                }
                QComboBox {
                    background: #ffffff;
                    color: #111827;
                    border: 1px solid #d1d5db;
                    border-radius: 8px;
                    padding: 6px;
                }
                QComboBox::drop-down {
                    background: #f9fafb;
                    border-left: 1px solid #d1d5db;
                }
                QComboBox QAbstractItemView {
                    background: #ffffff;
                    color: #111827;
                    selection-background-color: #e8eefc;
                    selection-color: #1b2240;
                }
                QPushButton {
                    background: #e5e7eb;
                    color: #111827;
                    padding: 8px 14px;
                    border: 1px solid #d1d5db;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #dbeafe;
                    border: 1px solid #bfdbfe;
                }
            """)
        else:
            dlg.setStyleSheet("""
                QInputDialog {
                    background-color: #2b2b2b;
                    color: #ffffff;
                }
                QLabel {
                    color: #ffffff;
                    background: transparent;
                }
                QComboBox {
                    background: #0f1422;
                    color: #ffffff;
                    border: 1px solid #2c3550;
                    border-radius: 8px;
                    padding: 6px;
                }
                QComboBox::drop-down {
                    background: #1a1f2e;
                    border-left: 1px solid #2c3550;
                }
                QComboBox QAbstractItemView {
                    background: #0f1422;
                    color: #ffffff;
                    selection-background-color: #2a2f43;
                    selection-color: #ffffff;
                }
                QPushButton {
                    background: #1a2031;
                    color: #ffffff;
                    padding: 8px 14px;
                    border: 1px solid #2c3550;
                    border-radius: 10px;
                }
                QPushButton:hover {
                    background: #252c45;
                    border: 1px solid #3d4a70;
                }
            """)
        
        ok = dlg.exec()
        new_label = dlg.textValue()
        
        if ok and new_label:
            try:
                self.db.execute(
                    "UPDATE orders SET label = ? WHERE id = ?",
                    (new_label, order_id)
                )
                self.db.commit()
                if self.toast_cb:
                    self.toast_cb(f"✓ Etiqueta alterada para '{new_label}'")
                self.refresh()
            except Exception as e:
                if self.toast_cb:
                    self.toast_cb(f"✗ Erro ao alterar etiqueta: {e}")

    def _show_logs(self, entity: str, entity_id: Optional[int]) -> None:
        dlg = LogsDialog(self.db, entity, entity_id, self)
        dlg.exec()

    def current_id(self) -> Optional[int]:
        idx = self.table.currentRow()
        if idx < 0:
            return None
        vh = self.table.verticalHeaderItem(idx)
        return int(vh.text()) if vh else None

    def add(self) -> None:
        """Cria novo pedido com suporte a múltiplos produtos"""
        # Usa o novo diálogo de múltiplos produtos
        d = MultiProductOrderDialog(self.db, self)
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                order_data = d.get_order_data()
                
                if not order_data["products"]:
                    show_message(self, "Aviso", "Nenhum produto foi adicionado ao pedido.", ("OK",))
                    return
                
                cust_id = order_data["customer_id"]
                d_iso = order_data["delivery_date"]
                status = order_data["status"]
                label = order_data["label"]
                notes = order_data["notes"]
                now = datetime.now().strftime("%Y-%m-%d %H:%M")
                
                # Verifica se a data de entrega é uma segunda-feira (weekday() == 0)
                delivery_date_obj = datetime.strptime(d_iso, "%Y-%m-%d")
                if delivery_date_obj.weekday() == 0:
                    show_message(self, "Pedido Negado", "Segunda não há expediente, pedido negado.", ("OK",))
                    return
                
                # Verifica se deve reservar estoque agora ou depois
                today_obj = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                should_reserve_now = delivery_date_obj <= today_obj
                
                # Cria um pedido para CADA produto
                pedidos_criados = []
                for product in order_data["products"]:
                    prod_id = product["product_id"]
                    qty = product["quantity"]
                    total = 0.0  # Não controlamos preços
                    
                    # Cria pedido
                    cur = self.db.execute("""
                        INSERT INTO orders(order_number, customer_id, product_id, quantity,
                                           delivery_date, total, status, label, notes, stock_reserved, created_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """, (None, cust_id, prod_id, qty, d_iso, total, status, label, notes, 1 if should_reserve_now else 0, now))
                    oid = int(cur.lastrowid) if cur and cur.lastrowid is not None else None
                    
                    # Define order_number sequencial
                    if oid:
                        self.db.execute("UPDATE orders SET order_number=? WHERE id=?", (oid, oid))
                        pedidos_criados.append(oid)
                    
                    # Baixa estoque APENAS se data de entrega <= hoje
                    if should_reserve_now:
                        self.db.execute(
                            "UPDATE products SET stock = stock - ? WHERE id=?",
                            (qty, prod_id)
                        )
                        self.db.execute(
                            "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                            (prod_id, 'saida', qty, 'Pedido', oid, now)
                        )
                    
                    # Auditoria
                    try:
                        self.db.audit_log("order", oid, "create", details=f"cust={cust_id},prod={prod_id},qty={qty},total={total},reserved={should_reserve_now}")
                    except Exception:
                        pass
                
                self.refresh()
                if self.dashboard_cb:
                    self.dashboard_cb()
                if self.toast_cb:
                    qtd_produtos = len(pedidos_criados)
                    self.toast_cb(f"✅ {qtd_produtos} pedido(s) criado(s) com sucesso!")
                    
            except Exception as e:
                show_message(self, "Erro ao salvar", str(e), ("OK",))

    def edit(self) -> None:
        oid = self.current_id()
        if not oid:
            return
        row = self.db.query("SELECT * FROM orders WHERE id=?", (oid,))
        if not row:
            return
        data = row[0]
        d = OrderDialog(self.db, self, data)
        if d.exec() == QDialog.DialogCode.Accepted:
            try:
                cust_id, prod_id, qty_new, d_iso, total, status, label, notes = d.get_values()
                # reverte estoque do pedido antigo
                old_qty = int(data["quantity"])
                old_pid = int(data["product_id"])
                if old_pid == prod_id:
                    delta = qty_new - old_qty
                    # se delta > 0, baixa; se delta < 0, devolve
                    if delta != 0:
                        mov_type = 'saida' if delta > 0 else 'entrada'
                        self.db.execute("UPDATE products SET stock = stock - ? WHERE id=?", (delta, prod_id))
                        self.db.execute(
                            "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                            (prod_id, mov_type, abs(delta), 'Ajuste pedido', oid, datetime.now().strftime("%Y-%m-%d %H:%M"))
                        )
                else:
                    # devolve antigo
                    self.db.execute("UPDATE products SET stock = stock + ? WHERE id=?", (old_qty, old_pid))
                    self.db.execute(
                        "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                        (old_pid, 'entrada', old_qty, 'Ajuste pedido', oid, datetime.now().strftime("%Y-%m-%d %H:%M"))
                    )
                    # baixa novo
                    self.db.execute("UPDATE products SET stock = stock - ? WHERE id=?", (qty_new, prod_id))
                    self.db.execute(
                        "INSERT INTO stock_movements(product_id, type, quantity, reason, order_id, created_at) VALUES (?,?,?,?,?,?)",
                        (prod_id, 'saida', qty_new, 'Ajuste pedido', oid, datetime.now().strftime("%Y-%m-%d %H:%M"))
                    )

                self.db.execute("""
                    UPDATE orders SET customer_id=?, product_id=?, quantity=?, delivery_date=?, total=?, status=?, label=?, notes=?
                    WHERE id=?
                """, (cust_id, prod_id, qty_new, d_iso, total, status, label, notes, oid))
                # auditoria: atualização do pedido (somente campos alterados)
                try:
                    changes: list[str] = []
                    # cliente
                    try:
                        old_cust_id = int(data["customer_id"])
                    except Exception:
                        old_cust_id = None
                    if old_cust_id != cust_id:
                        old_name = None
                        new_name = None
                        r = self.db.query("SELECT name FROM customers WHERE id=?", (old_cust_id,)) if old_cust_id else []
                        if r:
                            old_name = r[0]["name"]
                        r2 = self.db.query("SELECT name FROM customers WHERE id=?", (cust_id,))
                        if r2:
                            new_name = r2[0]["name"]
                        changes.append(f"customer: '{old_name or old_cust_id}' -> '{new_name or cust_id}'")
                    # produto
                    try:
                        old_prod_id = int(data["product_id"])
                    except Exception:
                        old_prod_id = None
                    if old_prod_id != prod_id:
                        old_p = self.db.query("SELECT name FROM products WHERE id=?", (old_prod_id,)) if old_prod_id else []
                        new_p = self.db.query("SELECT name FROM products WHERE id=?", (prod_id,))
                        old_pn = old_p[0]["name"] if old_p else old_prod_id
                        new_pn = new_p[0]["name"] if new_p else prod_id
                        changes.append(f"product: '{old_pn}' -> '{new_pn}'")
                    # quantidade
                    try:
                        old_qty = int(data["quantity"])
                    except Exception:
                        old_qty = None
                    if old_qty != qty_new:
                        changes.append(f"quantity: {old_qty} -> {qty_new}")
                    # entrega
                    if str(data["delivery_date"] or "") != str(d_iso or ""):
                        changes.append(f"delivery_date: '{data['delivery_date']}' -> '{d_iso}'")
                    # total
                    try:
                        old_total = float(data["total"])
                    except Exception:
                        old_total = None
                    if old_total is None or float(total) != old_total:
                        changes.append(f"total: {old_total} -> {total}")
                    # status
                    if str(data["status"] or "") != str(status or ""):
                        changes.append(f"status: '{data['status'] or ''}' -> '{status or ''}'")
                    # notes
                    if str(data["notes"] or "") != str(notes or ""):
                        changes.append(f"notes: '{data['notes'] or ''}' -> '{notes or ''}'")
                    details = "; ".join(changes) if changes else None
                    self.db.audit_log("order", oid, "update", details=details)
                except Exception:
                    pass
                self.refresh()
                if self.dashboard_cb:
                    self.dashboard_cb()
                if self.toast_cb:
                    self.toast_cb("Pedido atualizado")
            except Exception as e:
                show_message(self, "Erro ao atualizar", str(e), ("OK",))

    def delete(self) -> None:
        """Exclui o(s) pedido(s) selecionado(s)"""
        # Obtém todos os IDs selecionados
        selected_rows = self.table.selectionModel().selectedRows()
        if not selected_rows:
            return
        
        # Coleta os IDs de todos os pedidos selecionados
        order_ids = []
        for index in selected_rows:
            row = index.row()
            if row < self.table.rowCount():
                # Pega o ID do cabeçalho vertical (onde está armazenado)
                vh_item = self.table.verticalHeaderItem(row)
                if vh_item:
                    try:
                        oid = int(vh_item.text())
                        
                        # Verifica se é um pedido em lote checando as notas
                        lote_check = self.db.query("SELECT notes, customer_id, delivery_date, status, label FROM orders WHERE id=?", (oid,))
                        if lote_check and lote_check[0]["notes"] and lote_check[0]["notes"].startswith("LOTE:"):
                            # É um pedido em lote - precisa excluir TODOS os pedidos do mesmo lote
                            # Identifica todos os pedidos com o mesmo marcador LOTE, cliente, data e status
                            notes_prefix = lote_check[0]["notes"].split("|")[0].strip()  # Pega "LOTE:N itens"
                            cust_id = lote_check[0]["customer_id"]
                            delivery = lote_check[0]["delivery_date"]
                            status = lote_check[0]["status"]
                            label = lote_check[0]["label"]
                            
                            # Busca todos os IDs do lote
                            lote_ids_query = """
                                SELECT id FROM orders 
                                WHERE customer_id=? AND delivery_date=? AND status=? 
                                AND COALESCE(label, 'Comum')=COALESCE(?, 'Comum')
                                AND notes LIKE ?
                            """
                            lote_rows = self.db.query(lote_ids_query, (cust_id, delivery, status, label, f"{notes_prefix}%"))
                            for lote_row in lote_rows:
                                order_ids.append(int(lote_row["id"]))
                        else:
                            # Pedido normal - adiciona apenas este ID
                            order_ids.append(oid)
                    except (ValueError, AttributeError):
                        pass
        
        if not order_ids:
            return
        
        # Remove duplicatas mantendo ordem
        order_ids = list(dict.fromkeys(order_ids))
        
        # Mensagem de confirmação diferente para múltiplos pedidos
        count = len(order_ids)
        if count == 1:
            msg = "Excluir este pedido?"
        else:
            msg = f"Excluir {count} pedidos selecionados?"
        
        res = show_message(self, "Confirmação", msg, ("Não", "Sim"), default=1)
        if res == 1:
            try:
                deleted_count = 0
                for oid in order_ids:
                    try:
                        # Busca informações do pedido para devolver estoque
                        row = self.db.query("SELECT product_id, quantity FROM orders WHERE id=?", (oid,))
                        if row:
                            pid_value = row[0]["product_id"]
                            qty = int(row[0]["quantity"])
                            
                            # Só devolve estoque se tiver product_id (pedidos normais)
                            if pid_value is not None:
                                pid = int(pid_value)
                                self.db.execute("UPDATE products SET stock = stock + ? WHERE id=?", (qty, pid))
                        
                        # Remove TODAS as dependências antes de excluir o pedido
                        # 1. Remove movimentações de estoque relacionadas
                        self.db.execute("DELETE FROM stock_movements WHERE order_id=?", (oid,))
                        
                        # 2. Remove logs de auditoria relacionados
                        self.db.execute("DELETE FROM audit_logs WHERE entity='order' AND entity_id=?", (oid,))
                        
                        # 3. Exclui o pedido
                        self.db.execute("DELETE FROM orders WHERE id=?", (oid,))
                        
                        # Registra a exclusão no log
                        try:
                            self.db.audit_log("order", oid, "delete", details=None)
                        except Exception:
                            pass
                        
                        deleted_count += 1
                    except Exception as e:
                        print(f"Erro ao excluir pedido #{oid}: {e}")
                        continue
                
                self.refresh()
                if self.dashboard_cb:
                    self.dashboard_cb()
                if self.toast_cb:
                    if deleted_count == 1:
                        self.toast_cb("Pedido excluído")
                    else:
                        self.toast_cb(f"{deleted_count} pedidos excluídos")
            except Exception as e:
                show_message(self, "Erro ao excluir", str(e), ("OK",))

    def _set_today_filter(self) -> None:
        """Define o filtro para a data de hoje"""
        today = QDate.currentDate()
        self.date_start.setDate(today)
        self.date_end.setDate(today)
        self.refresh()

    def _clear_date_filter(self) -> None:
        """Remove o filtro de data mostrando todos os pedidos"""
        # Define uma data muito antiga e uma muito futura para mostrar todos
        self.date_start.setDate(QDate(2020, 1, 1))
        self.date_end.setDate(QDate(2030, 12, 31))
        self.refresh()

class ReportsPage(BasePage):
    def __init__(self, db: DB) -> None:
        super().__init__("Relatórios", "Visão geral e métricas")
        self.db = db
        bl = QVBoxLayout(self.body)

        self.lbl_sales_month: QLabel = QLabel("Vendas (mês atual): —")
        self.lbl_orders_month: QLabel = QLabel("Pedidos (mês atual): —")
        self.lbl_top_products: QLabel = QLabel("Top produtos: —")
        bl.addWidget(self.lbl_sales_month)
        bl.addWidget(self.lbl_orders_month)
        bl.addWidget(self.lbl_top_products)

        # Removidos os botões de Usuários e Exportações conforme solicitado
        
        # Inicializar abas de relatórios
        self.__init_tabs__()
    
    def showEvent(self, event) -> None:
        """Recarrega os dados quando a página é exibida"""
        super().showEvent(event)
        # Recarrega os dados sempre que a página é exibida
        try:
            self.refresh()
            self.refresh_entregas()
            self.refresh_producao()
            self.refresh_fechamento()
        except Exception:
            pass  # Ignora erros caso os widgets ainda não estejam completamente inicializados
    
    def _apply_calendar_style(self, calendar_widget: QWidget | None) -> None:
        """Aplica o estilo correto ao widget de calendário baseado no tema atual"""
        if calendar_widget is None:
            return
            
        from core.config import load_config
        config = load_config()
        theme = config.get("theme", "light")
        
        if theme == "dark":
            calendar_style = """
            QCalendarWidget {
                background-color: #1a1f2e;
                background: #1a1f2e;
                color: #ffffff;
                border: 1px solid #2c3550;
                border-radius: 8px;
            }
            QCalendarWidget QWidget {
                background-color: #1a1f2e;
                background: #1a1f2e;
                color: #ffffff;
            }
            QCalendarWidget QAbstractItemView {
                background-color: #1a1f2e;
                background: #1a1f2e;
                color: #ffffff;
                selection-background-color: #2a2f43;
                selection-color: #ffffff;
                gridline-color: #2c3550;
            }
            /* Seleção explícita dos dias (evita fundo preto em algumas plataformas/Qt builds) */
            QCalendarWidget QAbstractItemView::item:selected {
                background-color: #2a2f43;
                color: #ffffff;
                border-radius: 4px;
            }
            QCalendarWidget QAbstractItemView::item:hover {
                background-color: #1e3a5f;
                color: #ffffff;
            }
            QCalendarWidget QTableView {
                background-color: #1a1f2e;
                background: #1a1f2e;
                color: #ffffff;
            }
            QCalendarWidget QToolButton {
                background-color: #252c45;
                background: #252c45;
                color: #ffffff;
                border: 1px solid #2c3550;
                border-radius: 6px;
                padding: 6px;
            }
            QCalendarWidget QToolButton:hover { 
                background-color: #333b5e;
                background: #333b5e; 
            }
            QCalendarWidget QSpinBox {
                background-color: #1a1f2e;
                background: #1a1f2e;
                color: #ffffff;
                border: 1px solid #2c3550;
                border-radius: 6px;
                padding: 2px 6px;
            }
            QCalendarWidget QHeaderView::section {
                background-color: #252c45;
                background: #252c45;
                color: #ffffff;
                border: 1px solid #2c3550;
                padding: 4px;
            }
            """
        else:
            calendar_style = """
            QCalendarWidget {
                background-color: #ffffff;
                background: #ffffff;
                color: #111827;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }
            QCalendarWidget QWidget {
                background-color: #ffffff;
                background: #ffffff;
                color: #111827;
            }
            QCalendarWidget QAbstractItemView {
                background-color: #ffffff;
                background: #ffffff;
                color: #111827;
                selection-background-color: #e8eefc;
                selection-color: #1b2240;
                gridline-color: #e5e7eb;
            }
            /* Seleção explícita dos dias (evita fundo preto em algumas plataformas/Qt builds) */
            QCalendarWidget QAbstractItemView::item:selected {
                background-color: #e8eefc;
                color: #1b2240;
                border-radius: 4px;
            }
            QCalendarWidget QAbstractItemView::item:hover {
                background-color: #dbeafe;
                color: #111827;
            }
            QCalendarWidget QTableView {
                background-color: #ffffff;
                background: #ffffff;
                color: #111827;
            }
            QCalendarWidget QToolButton {
                background-color: #f3f4f6;
                background: #f3f4f6;
                color: #1f2937;
                border: 1px solid #e5e7eb;
                border-radius: 6px;
                padding: 4px 8px;
            }
            QCalendarWidget QToolButton:hover { 
                background-color: #e5e7eb;
                background: #e5e7eb; 
            }
            QCalendarWidget QSpinBox {
                background-color: #ffffff;
                background: #ffffff;
                color: #111827;
                border: 1px solid #d1d5db;
                border-radius: 6px;
                padding: 2px 6px;
            }
            QCalendarWidget QHeaderView::section {
                background-color: #f9fafb;
                background: #f9fafb;
                color: #1f2937;
                border: 1px solid #e5e7eb;
                padding: 4px;
            }
            """
        
        calendar_widget.setStyleSheet(calendar_style)

    def __init_tabs__(self) -> None:
        """Inicializa as abas de relatórios"""
        # Abas de Relatórios: Entregas, Produção e Fechamento
        bl = self.body.layout() or QVBoxLayout(self.body)
        self.tabs: QTabWidget = QTabWidget()
        bl.addWidget(self.tabs)

        # Relatório de Entregas (por dia, somente 'Pago')
        self.tab_entregas = QWidget(); v1 = QVBoxLayout(self.tab_entregas)
        f1 = QHBoxLayout(); v1.addLayout(f1)
        f1.addWidget(QLabel("Dia:"))
        self.entregas_date: QDateEdit = QDateEdit(); self.entregas_date.setCalendarPopup(True)
        # Estilo explícito do campo conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.entregas_date.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.entregas_date.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
        except Exception:
            pass
        # Apply calendar styling
        calendar_widget0 = self.entregas_date.calendarWidget()
        if calendar_widget0:
            self._apply_calendar_style(calendar_widget0)
        self.entregas_date.setDisplayFormat("dd/MM/yyyy"); self.entregas_date.setDate(QDate.currentDate())
        f1.addWidget(self.entregas_date); f1.addStretch(1)
        # Export buttons for Entregas
        self.btn_ent_pdf: QPushButton = QPushButton(); self.btn_ent_pdf.setFlat(True)
        self.btn_ent_pdf.setToolTip("Exportar Entregas (PDF)")
        try:
            self.btn_ent_pdf.setIcon(safe_qta_icon("ph.file-pdf", color="#dc2626"))
        except Exception:
            self.btn_ent_pdf.setText("PDF")
        self.btn_ent_pdf.setIconSize(QSize(18,18)); self.btn_ent_pdf.setFixedHeight(28)
        f1.addWidget(self.btn_ent_pdf)

        self.btn_ent_xlsx: QPushButton = QPushButton(); self.btn_ent_xlsx.setFlat(True)
        self.btn_ent_xlsx.setToolTip("Exportar Entregas (Excel)")
        icon = safe_qta_icon("ph.file", color="#16a34a")
        if icon and not icon.isNull():
            self.btn_ent_xlsx.setIcon(icon)
        else:
            self.btn_ent_xlsx.setText("XLSX")
        self.btn_ent_xlsx.setIconSize(QSize(18,18)); self.btn_ent_xlsx.setFixedHeight(28)
        f1.addWidget(self.btn_ent_xlsx)
        self.tbl_entregas: QTableWidget = QTableWidget(0, 5)
        self.tbl_entregas.setHorizontalHeaderLabels(["Cliente", "Quantidade", "Produto", "Endereço", "Observação"])
        self.tbl_entregas.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_entregas.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        cast(Any, self.tbl_entregas.verticalHeader()).setVisible(False)
        if header := self.tbl_entregas.horizontalHeader():
            header.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            header.setStretchLastSection(True)
        v1.addWidget(self.tbl_entregas)
        self.lbl_entregas_tot: QLabel = QLabel("")
        v1.addWidget(self.lbl_entregas_tot)
        cast(Any, self.entregas_date.dateChanged).connect(self.refresh_entregas)
        self.tabs.addTab(self.tab_entregas, "Relatório de Entregas")

        # Relatório de Produção (entradas e saídas)
        self.tab_prod = QWidget(); v2 = QVBoxLayout(self.tab_prod)
        f2 = QHBoxLayout(); v2.addLayout(f2)
        f2.addWidget(QLabel("Período:"))
        self.prod_from: QDateEdit = QDateEdit(); self.prod_from.setCalendarPopup(True); self.prod_from.setDisplayFormat("dd/MM/yyyy")
        # Estilo explícito do campo conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.prod_from.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.prod_from.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
        except Exception:
            pass
        # Apply calendar styling
        calendar_widget = self.prod_from.calendarWidget()
        if calendar_widget:
            self._apply_calendar_style(calendar_widget)
        today = QDate.currentDate(); first = QDate(today.year(), today.month(), 1)
        self.prod_from.setDate(first)
        f2.addWidget(self.prod_from)
        f2.addWidget(QLabel("até"))
        self.prod_to: QDateEdit = QDateEdit(); self.prod_to.setCalendarPopup(True); self.prod_to.setDisplayFormat("dd/MM/yyyy"); self.prod_to.setDate(today)
        # Estilo explícito do campo conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.prod_to.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.prod_to.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
        except Exception:
            pass
        # Apply calendar styling
        calendar_widget2 = self.prod_to.calendarWidget()
        if calendar_widget2:
            self._apply_calendar_style(calendar_widget2)
        f2.addWidget(self.prod_to); f2.addStretch(1)
        # Export buttons for Produção
        self.btn_prod_pdf: QPushButton = QPushButton(); self.btn_prod_pdf.setFlat(True)
        self.btn_prod_pdf.setToolTip("Exportar Produção (PDF)")
        icon = safe_qta_icon("ph.file-pdf", color="#dc2626")
        if icon and not icon.isNull():
            self.btn_prod_pdf.setIcon(icon)
        else:
            self.btn_prod_pdf.setText("PDF")
        self.btn_prod_pdf.setIconSize(QSize(18,18)); self.btn_prod_pdf.setFixedHeight(28)
        f2.addWidget(self.btn_prod_pdf)

        self.btn_prod_xlsx: QPushButton = QPushButton(); self.btn_prod_xlsx.setFlat(True)
        self.btn_prod_xlsx.setToolTip("Exportar Produção (Excel)")
        icon = safe_qta_icon("ph.file", color="#16a34a")
        if icon and not icon.isNull():
            self.btn_prod_xlsx.setIcon(icon)
        else:
            self.btn_prod_xlsx.setText("XLSX")
        self.btn_prod_xlsx.setIconSize(QSize(18,18)); self.btn_prod_xlsx.setFixedHeight(28)
        f2.addWidget(self.btn_prod_xlsx)
        self.tbl_prod: QTableWidget = QTableWidget(0, 5)
        self.tbl_prod.setHorizontalHeaderLabels(["Data", "Produto", "Tipo", "Quantidade", "Motivo"])
        self.tbl_prod.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_prod.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        cast(Any, self.tbl_prod.verticalHeader()).setVisible(False)
        if header2 := self.tbl_prod.horizontalHeader():
            header2.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            header2.setStretchLastSection(True)
        v2.addWidget(self.tbl_prod)
        self.lbl_prod_tot: QLabel = QLabel("")
        v2.addWidget(self.lbl_prod_tot)
        cast(Any, self.prod_from.dateChanged).connect(self.refresh_producao)
        cast(Any, self.prod_to.dateChanged).connect(self.refresh_producao)
        # Wire export buttons
        cast(Any, self.btn_ent_pdf.clicked).connect(self.export_pdf)
        cast(Any, self.btn_ent_xlsx.clicked).connect(self.export_xlsx)
        cast(Any, self.btn_prod_pdf.clicked).connect(self.export_pdf)
        cast(Any, self.btn_prod_xlsx.clicked).connect(self.export_xlsx)
        self.tabs.addTab(self.tab_prod, "Relatório de Produção")

        # Relatório de Fechamento
        self.tab_fech = QWidget(); v3 = QVBoxLayout(self.tab_fech)
        f3 = QHBoxLayout(); v3.addLayout(f3)
        f3.addWidget(QLabel("Período:"))
        self.fech_from: QDateEdit = QDateEdit(); self.fech_from.setCalendarPopup(True); self.fech_from.setDisplayFormat("dd/MM/yyyy")
        # Estilo explícito do campo conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.fech_from.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.fech_from.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
        except Exception:
            pass
        # Apply calendar styling
        calendar_widget3 = self.fech_from.calendarWidget()
        if calendar_widget3:
            self._apply_calendar_style(calendar_widget3)
        self.fech_from.setDate(first)
        f3.addWidget(self.fech_from)
        f3.addWidget(QLabel("até"))
        self.fech_to: QDateEdit = QDateEdit(); self.fech_to.setCalendarPopup(True); self.fech_to.setDisplayFormat("dd/MM/yyyy"); self.fech_to.setDate(today)
        # Estilo explícito do campo conforme tema
        try:
            from core.config import load_config
            _theme = load_config().get("theme", "light")
            if _theme == "dark":
                self.fech_to.setStyleSheet("""
                    QDateEdit { background: #0f1422; color: #ffffff; border: 1px solid #2c3550; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #1a1f2e; border-left: 1px solid #2c3550; }
                """)
            else:
                self.fech_to.setStyleSheet("""
                    QDateEdit { background: #ffffff; color: #111827; border: 1px solid #d1d5db; border-radius: 8px; padding: 6px; }
                    QDateEdit::drop-down { background: #f9fafb; border-left: 1px solid #d1d5db; }
                """)
        except Exception:
            pass
        # Apply calendar styling
        calendar_widget4 = self.fech_to.calendarWidget()
        if calendar_widget4:
            self._apply_calendar_style(calendar_widget4)
        f3.addWidget(self.fech_to)
        f3.addWidget(QLabel("Etiqueta:"))
        self.fech_label_filter: QComboBox = QComboBox(); self._reload_labels_into_combo()
        f3.addWidget(self.fech_label_filter); f3.addStretch(1)
        # Botões discretos de exportação apenas nesta aba
        self.btn_fech_pdf: QPushButton = QPushButton()
        self.btn_fech_pdf.setToolTip("Exportar relatório de fechamento (PDF)")
        self.btn_fech_pdf.setFlat(True)
        icon = safe_qta_icon("ph.file-pdf", color="#dc2626")
        if icon and not icon.isNull():
            self.btn_fech_pdf.setIcon(icon)
        else:
            self.btn_fech_pdf.setText("PDF")
        self.btn_fech_pdf.setIconSize(QSize(18, 18))
        self.btn_fech_pdf.setFixedHeight(28)
        f3.addWidget(self.btn_fech_pdf)

        self.btn_fech_xlsx: QPushButton = QPushButton()
        self.btn_fech_xlsx.setToolTip("Exportar relatório de fechamento (Excel)")
        self.btn_fech_xlsx.setFlat(True)
        # Não há ícone XLSX padrão garantido; usar ícone de arquivo e cor verde
        icon = safe_qta_icon("ph.file", color="#16a34a")
        if icon and not icon.isNull():
            self.btn_fech_xlsx.setIcon(icon)
        else:
            self.btn_fech_xlsx.setText("XLSX")
        self.btn_fech_xlsx.setIconSize(QSize(18, 18))
        self.btn_fech_xlsx.setFixedHeight(28)
        f3.addWidget(self.btn_fech_xlsx)
        self.tbl_fech: QTableWidget = QTableWidget(0, 6)
        self.tbl_fech.setHorizontalHeaderLabels(["Cliente", "Etiqueta", "Produto", "Data", "Quantidade", "Valor"])
        self.tbl_fech.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl_fech.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        cast(Any, self.tbl_fech.verticalHeader()).setVisible(False)
        if header3 := self.tbl_fech.horizontalHeader():
            header3.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
            header3.setStretchLastSection(True)
        v3.addWidget(self.tbl_fech)
        self.lbl_fech_tot: QLabel = QLabel("")
        v3.addWidget(self.lbl_fech_tot)
        cast(Any, self.fech_from.dateChanged).connect(self.refresh_fechamento)
        cast(Any, self.fech_to.dateChanged).connect(self.refresh_fechamento)
        cast(Any, self.fech_label_filter.currentIndexChanged).connect(self.refresh_fechamento)
        cast(Any, self.btn_fech_pdf.clicked).connect(self.export_pdf)
        cast(Any, self.btn_fech_xlsx.clicked).connect(self.export_xlsx)
        self.tabs.addTab(self.tab_fech, "Relatório de Fechamento")

        self.refresh()
        self.refresh_entregas()
        self.refresh_producao()
        self.refresh_fechamento()

    def _reload_labels_into_combo(self) -> None:
        self.fech_label_filter.clear()
        self.fech_label_filter.addItem("Todas")
        try:
            rows = self.db.query("SELECT name FROM labels ORDER BY name")
            for r in rows:
                self.fech_label_filter.addItem(str(r["name"]))
        except Exception:
            pass

    # ---- Exportações ----
    def export_xlsx(self) -> None:
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.drawing.image import Image as XLImage
        except ImportError:
            show_message(self, "Erro", "A biblioteca openpyxl não está instalada. Instale com: pip install openpyxl", ("OK",))
            return

        dlg = CustomFileDialog(self, caption="Salvar Excel", directory="", filter="Excel (*.xlsx)", save_mode=True)
        if not dlg.exec():
            return
        path = dlg.get_selected_file()
        if not path:
            return
        # Garante extensão .xlsx
        if not path.lower().endswith(".xlsx"):
            path = path + ".xlsx"

        # Coleta dataset conforme aba ativa
        title, headers, data, meta, summary = self._current_report_dataset()

        wb = openpyxl.Workbook(); ws = wb.active
        ws.title = title[:31] if title else "Relatório"

        # Logo
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "assets", "icons", "logo.png")
            if os.path.exists(logo_path):
                img = XLImage(logo_path)
                img.width = 120; img.height = 60
                ws.add_image(img, "A1")
        except Exception:
            pass

        # Título e meta
        start_row = 5
        ws.cell(row=start_row, column=1, value=title).font = Font(size=14, bold=True)
        ws.cell(row=start_row+1, column=1, value=meta or "").font = Font(size=10, italic=True)

        # Cabeçalho
        header_row = start_row + 3
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2563EB")  # azul
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Dados
        for r_idx, rowdata in enumerate(data, start=header_row+1):
            for c, val in enumerate(rowdata, start=1):
                cell = ws.cell(row=r_idx, column=c, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        # Auto largura
        for i in range(1, len(headers)+1):
            max_length = 0
            col_letter = get_column_letter(i)
            for cell in ws[col_letter]:
                try:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_length:
                        max_length = len(v)
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Summary (para Fechamento)
        if summary:
            sr = ws.max_row + 2
            ws.cell(row=sr, column=1, value="Resumo por Cliente/Etiqueta").font = Font(size=12, bold=True)
            headers_s = ["Cliente", "Etiqueta", "Pedidos", "Quantidade", "Valor"]
            for c, h in enumerate(headers_s, start=1):
                cell = ws.cell(row=sr+1, column=c, value=h)
                cell.font = Font(bold=True)
            for i, s in enumerate(summary, start=sr+2):
                for c, val in enumerate(s, start=1):
                    ws.cell(row=i, column=c, value=val)

        wb.save(path)
        show_message(self, "Exportado", f"Excel salvo em:\n{path}", ("OK",))

    def export_pdf(self) -> None:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            show_message(self, "Erro", "A biblioteca reportlab não está instalada. Instale com: pip install reportlab", ("OK",))
            return

        dlg = CustomFileDialog(self, caption="Salvar PDF", directory="", filter="PDF (*.pdf)", save_mode=True)
        if not dlg.exec():
            return
        path = dlg.get_selected_file()
        if not path:
            return
        # Garante extensão .pdf
        if not path.lower().endswith(".pdf"):
            path = path + ".pdf"

        # Coleta dataset conforme aba ativa
        title, headers, data, meta, summary = self._current_report_dataset()

        doc = SimpleDocTemplate(path, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=60, bottomMargin=36)
        story = []
        styles = getSampleStyleSheet()

        # Logo e Título
        try:
            logo_path = os.path.join(os.path.dirname(__file__), "assets", "icons", "logo.png")
            if os.path.exists(logo_path):
                im = Image(logo_path, width=80, height=40)
                story.append(im)
        except Exception:
            pass
        story.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
        if meta:
            story.append(Paragraph(meta, styles["Italic"]))
        story.append(Spacer(1, 12))

        # Tabela principal
        table = Table([headers] + data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 9),
            ("BOTTOMPADDING", (0,0), (-1,0), 8),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ]))
        story.append(table)

        # Resumo para Fechamento
        if summary:
            story.append(Spacer(1, 18))
            story.append(Paragraph("<b>Resumo por Cliente/Etiqueta</b>", styles["Heading3"]))
            headers_s = ["Cliente", "Etiqueta", "Pedidos", "Quantidade", "Valor"]
            table_s = Table([headers_s] + summary, repeatRows=1)
            table_s.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1F2937")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,-1), 9),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ]))
            story.append(table_s)

        doc.build(story)
        show_message(self, "Exportado", f"PDF salvo em:\n{path}", ("OK",))

    def export_csv(self) -> None:
        import csv
        dlg = CustomFileDialog(self, caption="Salvar CSV", directory="", filter="CSV (*.csv)", save_mode=True)
        if not dlg.exec():
            return
        path = dlg.get_selected_file()
        if not path:
            return
        rows = self.db.query("""
            SELECT o.id, c.name AS cliente, p.name AS produto, o.quantity, o.delivery_date, o.total, o.status, o.created_at
            FROM orders o JOIN customers c ON c.id=o.customer_id JOIN products p ON p.id=o.product_id
            ORDER BY o.created_at DESC
        """)
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(["ID","Cliente","Produto","Quantidade","Entrega","Total","Status","Criado em"])
            for r in rows:
                w.writerow([
                    r["id"], r["cliente"], r["produto"], r["quantity"], 
                    format_date(str(r["delivery_date"])), 
                    f"{float(r['total']):.2f}", 
                    r["status"], 
                    format_date(str(r["created_at"]).split()[0]) if r["created_at"] else ""
                ])
        show_message(self, "Exportado", f"Arquivo salvo em:\n{path}", ("OK",))

    # ---- Métricas ----
    def refresh(self) -> None:
        today = date.today()
        ym = today.strftime("%Y-%m")
        sales = self.db.query("SELECT SUM(total) AS s FROM orders WHERE substr(created_at,1,7)=?", (ym,))
        sales_val = sales[0]["s"] or 0
        self.lbl_sales_month.setText(f"Vendas (mês atual): {money(float(sales_val))}")
        orders_count = self.db.query("SELECT COUNT(*) AS c FROM orders WHERE substr(created_at,1,7)=?", (ym,))
        self.lbl_orders_month.setText(f"Pedidos (mês atual): {orders_count[0]['c']}")
        tops = self.db.query("""
            SELECT p.name, SUM(o.quantity) AS q
            FROM orders o JOIN products p ON p.id=o.product_id
            GROUP BY p.id ORDER BY q DESC LIMIT 3
        """)
        if tops:
            self.lbl_top_products.setText("Top produtos: " + ", ".join([f"{r['name']} ({r['q']})" for r in tops]))
        else:
            self.lbl_top_products.setText("Top produtos: —")

    def refresh_entregas(self) -> None:
        d = self.entregas_date.date().toString("yyyy-MM-dd")
        rows = self.db.query("""
            SELECT c.name AS cliente, o.quantity AS quantidade, p.name AS produto, c.address AS endereco, o.notes AS obs
            FROM orders o
            JOIN customers c ON c.id = o.customer_id
            JOIN products  p ON p.id = o.product_id
            WHERE o.status = 'Pago' AND o.delivery_date = ?
            ORDER BY c.name
        """, (d,))
        self.tbl_entregas.setRowCount(0)
        for r in rows:
            row = self.tbl_entregas.rowCount(); self.tbl_entregas.insertRow(row)
            self.tbl_entregas.setItem(row, 0, QTableWidgetItem(str(r["cliente"])))
            self.tbl_entregas.setItem(row, 1, QTableWidgetItem(str(r["quantidade"])))
            self.tbl_entregas.setItem(row, 2, QTableWidgetItem(str(r["produto"])))
            self.tbl_entregas.setItem(row, 3, QTableWidgetItem(str(r["endereco"]) if r["endereco"] else ""))
            self.tbl_entregas.setItem(row, 4, QTableWidgetItem(str(r["obs"]) if r["obs"] else ""))
        self.tbl_entregas.resizeColumnsToContents()
        total_itens = sum(int(r["quantidade"]) for r in rows) if rows else 0
        total_pedidos = len(rows)
        self.lbl_entregas_tot.setText(f"Total itens: {total_itens} | Pedidos: {total_pedidos}")

    def refresh_producao(self) -> None:
        start = self.prod_from.date().toString("yyyy-MM-dd")
        end = self.prod_to.date().toString("yyyy-MM-dd")
        rows = self.db.query("""
            SELECT sm.created_at AS data, p.name AS produto, sm.type AS tipo, sm.quantity AS quantidade, sm.reason AS motivo
            FROM stock_movements sm
            JOIN products p ON p.id = sm.product_id
            WHERE date(sm.created_at) BETWEEN ? AND ?
            ORDER BY sm.created_at DESC
        """, (start, end))
        self.tbl_prod.setRowCount(0)
        for r in rows:
            row = self.tbl_prod.rowCount(); self.tbl_prod.insertRow(row)
            # Formata data no padrão dd/MM/yyyy
            try:
                d_iso = str(r["data"]) if r["data"] is not None else ""
                d_only = d_iso.split(" ")[0]
                d_fmt = format_date(d_only) if d_only else ""
            except Exception:
                d_fmt = str(r["data"]) if "data" in r.keys() else ""
            self.tbl_prod.setItem(row, 0, QTableWidgetItem(d_fmt))
            self.tbl_prod.setItem(row, 1, QTableWidgetItem(str(r["produto"])))
            self.tbl_prod.setItem(row, 2, QTableWidgetItem(str(r["tipo"])))
            self.tbl_prod.setItem(row, 3, QTableWidgetItem(str(r["quantidade"])))
            self.tbl_prod.setItem(row, 4, QTableWidgetItem(str(r["motivo"]) if r["motivo"] else ""))
        self.tbl_prod.resizeColumnsToContents()
        entradas = sum(int(r["quantidade"]) for r in rows if str(r["tipo"]).lower() == 'entrada')
        saidas = sum(int(r["quantidade"]) for r in rows if str(r["tipo"]).lower() == 'saida')
        saldo = entradas - saidas
        self.lbl_prod_tot.setText(f"Entradas: {entradas} | Saídas: {saidas} | Saldo: {saldo}")

    def refresh_fechamento(self) -> None:
        start = self.fech_from.date().toString("yyyy-MM-dd")
        end = self.fech_to.date().toString("yyyy-MM-dd")
        params: list[Any] = [start, end]
        sql = (
            "SELECT c.name AS cliente, 'Comum' AS etiqueta, "
            "p.name AS produto, date(o.created_at) AS data, o.quantity AS quantidade, o.total AS valor "
            "FROM orders o JOIN customers c ON c.id=o.customer_id JOIN products p ON p.id=o.product_id "
            "WHERE date(o.created_at) BETWEEN ? AND ?"
        )
        sql += " ORDER BY c.name, etiqueta, data"
        rows = self.db.query(sql, tuple(params))
        
        # Agrupa pedidos: se mesmo cliente, mesma etiqueta e mesmo produto, soma quantidade e valor
        from collections import defaultdict
        grouped: dict[tuple[str, str, str], dict[str, Any]] = {}  # (cliente, etiqueta, produto) -> dados agregados
        
        for r in rows:
            cliente = str(r["cliente"])
            etiqueta = str(r["etiqueta"])
            produto = str(r["produto"])
            key = (cliente, etiqueta, produto)
            
            if key not in grouped:
                grouped[key] = {
                    "cliente": cliente,
                    "etiqueta": etiqueta,
                    "produto": produto,
                    "data": str(r["data"]) if r["data"] is not None else "",
                    "quantidade": 0,
                    "valor": 0.0
                }
            
            grouped[key]["quantidade"] += int(r["quantidade"] or 0)
            grouped[key]["valor"] += float(r["valor"] or 0.0)
        
        # Popula a tabela com dados agrupados
        self.tbl_fech.setRowCount(0)
        total_pedidos = 0; total_qtd = 0; total_val = 0.0
        
        for key in sorted(grouped.keys()):  # Ordena por (cliente, etiqueta, produto)
            item = grouped[key]
            row = self.tbl_fech.rowCount(); self.tbl_fech.insertRow(row)
            self.tbl_fech.setItem(row, 0, QTableWidgetItem(item["cliente"]))
            self.tbl_fech.setItem(row, 1, QTableWidgetItem(item["etiqueta"]))
            self.tbl_fech.setItem(row, 2, QTableWidgetItem(item["produto"]))
            
            # Formata data
            try:
                d_iso = item["data"]
                d_fmt = format_date(d_iso) if d_iso else ""
            except Exception:
                d_fmt = item["data"]
            self.tbl_fech.setItem(row, 3, QTableWidgetItem(d_fmt))
            self.tbl_fech.setItem(row, 4, QTableWidgetItem(str(item["quantidade"])))
            self.tbl_fech.setItem(row, 5, QTableWidgetItem(money(item["valor"])))
            
            total_pedidos += 1
            total_qtd += item["quantidade"]
            total_val += item["valor"]
            
        self.tbl_fech.resizeColumnsToContents()
        
        # Conta clientes únicos
        unique_clients = len({item["cliente"] for item in grouped.values()})
        
        self.lbl_fech_tot.setText(
            f"Clientes: {unique_clients} | Pedidos: {total_pedidos} | "
            f"Quantidade: {total_qtd} | Valor: {money(total_val)}"
        )

    def _current_report_dataset(self) -> tuple[str, list[str], list[list[Any]], str, list[list[Any]] | None]:
        """Retorna (title, headers, data, meta, summary) para a aba ativa."""
        idx = self.tabs.currentIndex()
        # Entregas
        if idx == self.tabs.indexOf(self.tab_entregas):
            d = self.entregas_date.date().toString("yyyy-MM-dd")
            title = "Relatório de Entregas"
            headers = ["Cliente", "Quantidade", "Produto", "Endereço", "Observação"]
            data: list[list[Any]] = []
            for i in range(self.tbl_entregas.rowCount()):
                data.append([
                    self.tbl_entregas.item(i,0).text() if self.tbl_entregas.item(i,0) else "",
                    self.tbl_entregas.item(i,1).text() if self.tbl_entregas.item(i,1) else "",
                    self.tbl_entregas.item(i,2).text() if self.tbl_entregas.item(i,2) else "",
                    self.tbl_entregas.item(i,3).text() if self.tbl_entregas.item(i,3) else "",
                    self.tbl_entregas.item(i,4).text() if self.tbl_entregas.item(i,4) else "",
                ])
            meta = f"Dia: {format_date(d)} | Status: Pago"
            return (title, headers, data, meta, None)
        # Produção
        if idx == self.tabs.indexOf(self.tab_prod):
            title = "Relatório de Produção"
            headers = ["Data", "Produto", "Tipo", "Quantidade", "Motivo"]
            data: list[list[Any]] = []
            for i in range(self.tbl_prod.rowCount()):
                data.append([
                    self.tbl_prod.item(i,0).text() if self.tbl_prod.item(i,0) else "",
                    self.tbl_prod.item(i,1).text() if self.tbl_prod.item(i,1) else "",
                    self.tbl_prod.item(i,2).text() if self.tbl_prod.item(i,2) else "",
                    self.tbl_prod.item(i,3).text() if self.tbl_prod.item(i,3) else "",
                    self.tbl_prod.item(i,4).text() if self.tbl_prod.item(i,4) else "",
                ])
            meta = f"Período: {self.prod_from.date().toString('dd/MM/yyyy')} a {self.prod_to.date().toString('dd/MM/yyyy')}"
            return (title, headers, data, meta, None)
        # Fechamento
        title = "Relatório de Fechamento"
        headers = ["Cliente", "Etiqueta", "Produto", "Data", "Quantidade", "Valor"]
        data: list[list[Any]] = []
        for i in range(self.tbl_fech.rowCount()):
            data.append([
                self.tbl_fech.item(i,0).text() if self.tbl_fech.item(i,0) else "",
                self.tbl_fech.item(i,1).text() if self.tbl_fech.item(i,1) else "",
                self.tbl_fech.item(i,2).text() if self.tbl_fech.item(i,2) else "",
                self.tbl_fech.item(i,3).text() if self.tbl_fech.item(i,3) else "",
                self.tbl_fech.item(i,4).text() if self.tbl_fech.item(i,4) else "",
                self.tbl_fech.item(i,5).text() if self.tbl_fech.item(i,5) else "",
            ])
        meta = (
            f"Período: {self.fech_from.date().toString('dd/MM/yyyy')} a {self.fech_to.date().toString('dd/MM/yyyy')}" 
            f" | Etiqueta: {self.fech_label_filter.currentText()}"
        )
        # Summary group by cliente/etiqueta
        summary: dict[tuple[str,str], dict[str, Any]] = {}
        for row in data:
            cliente, etiqueta, _, _, qtd, val = row
            key = (cliente, etiqueta)
            if key not in summary:
                summary[key] = {"pedidos": 0, "qtd": 0, "valor": 0.0}
            summary[key]["pedidos"] += 1
            try:
                summary[key]["qtd"] += int(str(qtd).split()[0]) if str(qtd).strip() else 0
            except Exception:
                pass
            try:
                # valor vem formatado (R$ xx,xx) -> normaliza
                v = str(val).replace("R$", "").strip().replace(".", "").replace(",", ".")
                summary[key]["valor"] += float(v) if v else 0.0
            except Exception:
                pass
        summary_rows: list[list[Any]] = []
        for (cli, eti), agg in sorted(summary.items()):
            summary_rows.append([cli, eti, agg["pedidos"], agg["qtd"], f"R$ {agg['valor']:.2f}".replace('.',',')])
        return (title, headers, data, meta, summary_rows)

    def show_users(self) -> None:
        """Abre o diálogo de gerenciamento de usuários (visibilidade já controlada externamente)."""
        try:
            dlg = UsersDialog(self.db, self)
            dlg.exec()
        except Exception:
            pass

class UpdateDialog(QDialog):
    """Diálogo de progresso da atualização"""
    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("Atualizando Sistema")
        self.setModal(True)
        self.setFixedSize(500, 150)
        
        layout = QVBoxLayout(self)
        
        self.label = QLabel("Preparando atualização...")
        self.label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.label)
        
        self.progress = QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        layout.addWidget(self.progress)
        
        self.details = QLabel("")
        self.details.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.details.setStyleSheet("color: #6b7280; font-size: 11px;")
        layout.addWidget(self.details)
        
        # Aplica tema
        try:
            from core.config import load_config
            config = load_config()
            theme = config.get("theme", "light")
            if theme == "dark":
                self.setStyleSheet("""
                    QDialog { background-color: #2b2b2b; color: #ffffff; }
                    QLabel { color: #ffffff; }
                    QProgressBar { background-color: #3c3c3c; border: 1px solid #555555; }
                """)
            else:
                self.setStyleSheet("""
                    QDialog { background-color: #ffffff; color: #111827; }
                    QLabel { color: #111827; }
                    QProgressBar { background-color: #f3f4f6; border: 1px solid #d1d5db; }
                """)
        except Exception:
            pass
    
    def update_progress(self, percent: int, message: str) -> None:
        """Atualiza o progresso"""
        self.progress.setValue(percent)
        self.label.setText(message)
        if percent < 100:
            self.details.setText(f"{percent}% concluído")
        else:
            self.details.setText("Concluído!")


class SettingsPage(BasePage):
    def __init__(self, db: Database, app: QApplication, parent_window: QMainWindow,
                 toast_cb: Optional[Callable[[str], None]] = None,
                 backup_cb: Optional[Callable[[], None]] = None) -> None:
        super().__init__("Configurações", "Personalização do sistema")
        self.db = db
        self.app = app
        self.parent_window = parent_window
        self.toast_cb = toast_cb
        self.backup_cb = backup_cb
        self.update_thread = None
        
        bl = QVBoxLayout(self.body)
        
        # === Seção: Tema ===
        theme_group = QFrame()
        theme_group.setObjectName("SettingsGroup")
        theme_layout = QVBoxLayout(theme_group)
        theme_title = QLabel("<b>🎨 Aparência</b>")
        theme_layout.addWidget(theme_title)
        
        # Primeira linha de botões: Escuro e Claro
        theme_buttons_1 = QHBoxLayout()
        self.btn_dark: QPushButton = QPushButton("🌙 Tema Escuro")
        self.btn_light: QPushButton = QPushButton("☀️ Tema Claro")
        theme_buttons_1.addWidget(self.btn_dark)
        theme_buttons_1.addWidget(self.btn_light)
        theme_buttons_1.addStretch(1)
        theme_layout.addLayout(theme_buttons_1)
        
        # Segunda linha de botões: Rosa, Roxo e Azul
        theme_buttons_2 = QHBoxLayout()
        self.btn_pink: QPushButton = QPushButton("🌸 Tema Rosa")
        self.btn_purple: QPushButton = QPushButton("💜 Tema Roxo")
        self.btn_blue: QPushButton = QPushButton("💙 Tema Azul")
        theme_buttons_2.addWidget(self.btn_pink)
        theme_buttons_2.addWidget(self.btn_purple)
        theme_buttons_2.addWidget(self.btn_blue)
        theme_buttons_2.addStretch(1)
        theme_layout.addLayout(theme_buttons_2)
        
        bl.addWidget(theme_group)
        
        # === Seção: Empresa ===
        company_group = QFrame()
        company_group.setObjectName("SettingsGroup")
        company_layout = QVBoxLayout(company_group)
        company_title = QLabel("<b>🏢 Empresa</b>")
        company_layout.addWidget(company_title)
        
        # Carregar dados atuais da empresa
        try:
            company_data = self.db.query("SELECT name, logo_path FROM company WHERE id=1")
            if company_data:
                self.current_company_name = company_data[0]["name"]
                self.current_logo_path = company_data[0]["logo_path"]
            else:
                self.current_company_name = "Confeitaria"
                self.current_logo_path = None
        except Exception:
            self.current_company_name = "Confeitaria"
            self.current_logo_path = None
        
        # Nome da empresa
        name_layout = QHBoxLayout()
        name_label = QLabel("Nome da Empresa:")
        name_layout.addWidget(name_label)
        
        self.company_name_input = QLineEdit()
        self.company_name_input.setText(self.current_company_name)
        self.company_name_input.setPlaceholderText("Digite o nome da empresa")
        name_layout.addWidget(self.company_name_input, 1)
        
        btn_save_name = QPushButton("💾 Salvar Nome")
        cast(Any, btn_save_name.clicked).connect(self.save_company_name)
        name_layout.addWidget(btn_save_name)
        
        company_layout.addLayout(name_layout)
        
        # Logo da empresa
        logo_layout = QHBoxLayout()
        logo_label = QLabel("Logo da Empresa:")
        logo_layout.addWidget(logo_label)
        
        self.logo_preview = QLabel()
        self.logo_preview.setFixedSize(64, 64)
        self.logo_preview.setStyleSheet("""
            QLabel {
                border: 2px dashed #cbd5e1;
                border-radius: 8px;
                background: #f8fafc;
            }
        """)
        self.logo_preview.setAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Carregar logo se existir
        if self.current_logo_path and os.path.exists(self.current_logo_path):
            from PyQt6.QtGui import QPixmap
            pixmap = QPixmap(self.current_logo_path)
            if not pixmap.isNull():
                self.logo_preview.setPixmap(pixmap.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            else:
                self.logo_preview.setText("🖼️")
        else:
            self.logo_preview.setText("🖼️")
        
        logo_layout.addWidget(self.logo_preview)
        
        btn_select_logo = QPushButton("📁 Selecionar Logo")
        cast(Any, btn_select_logo.clicked).connect(self.select_company_logo)
        logo_layout.addWidget(btn_select_logo)
        
        btn_remove_logo = QPushButton("🗑️ Remover Logo")
        cast(Any, btn_remove_logo.clicked).connect(self.remove_company_logo)
        logo_layout.addWidget(btn_remove_logo)
        
        logo_layout.addStretch(1)
        company_layout.addLayout(logo_layout)
        
        # Descrição
        info_label = QLabel("💡 A logo será exibida no topo do sistema")
        info_label.setStyleSheet("color: #6b7280; font-size: 11px; margin-top: 5px;")
        company_layout.addWidget(info_label)
        
        bl.addWidget(company_group)
        
        # === Seção: Atualização ===
        update_group = QFrame()
        update_group.setObjectName("SettingsGroup")
        update_layout = QVBoxLayout(update_group)
        update_title = QLabel("<b>🔄 Atualização do Sistema</b>")
        update_layout.addWidget(update_title)
        
        # Exibir versão atual
        try:
            from core.updater import CURRENT_VERSION
            version_label = QLabel(f"Versão atual: <b>{CURRENT_VERSION}</b>")
            version_label.setStyleSheet("color: #6b7280; font-size: 12px; margin-bottom: 8px;")
            update_layout.addWidget(version_label)
        except Exception:
            pass
        
        # Botão de verificar atualização
        self.btn_check_update = QPushButton("🔍 Verificar Atualizações")
        cast(Any, self.btn_check_update.clicked).connect(self.check_for_updates)
        update_layout.addWidget(self.btn_check_update)
        
        # Info sobre atualizações
        update_info = QLabel("💡 O sistema baixa apenas os arquivos modificados (~500KB)")
        update_info.setStyleSheet("color: #6b7280; font-size: 11px; margin-top: 5px;")
        update_layout.addWidget(update_info)
        
        bl.addWidget(update_group)
        
        # === Seção: Backup ===
        if backup_cb:
            backup_group = QFrame()
            backup_group.setObjectName("SettingsGroup")
            backup_layout = QVBoxLayout(backup_group)
            backup_title = QLabel("<b>💾 Backup</b>")
            backup_layout.addWidget(backup_title)
            
            self.btn_backup = QPushButton("Fazer Backup do Banco de Dados")
            cast(Any, self.btn_backup.clicked).connect(backup_cb)
            backup_layout.addWidget(self.btn_backup)
            
            bl.addWidget(backup_group)
        
        bl.addStretch(1)
        
        # Conecta eventos (tema com persistência)
        cast(Any, self.btn_dark.clicked).connect(lambda: self.set_theme('dark'))
        cast(Any, self.btn_light.clicked).connect(lambda: self.set_theme('light'))
        cast(Any, self.btn_pink.clicked).connect(lambda: self.set_theme('pink'))
        cast(Any, self.btn_purple.clicked).connect(lambda: self.set_theme('purple'))
        cast(Any, self.btn_blue.clicked).connect(lambda: self.set_theme('blue'))
    
    def set_theme(self, theme: str) -> None:
        """Aplica e salva o tema escolhido (dark/light/pink/purple/blue)."""
        try:
            from core.config import load_config, save_config
            # Aplica o QSS correspondente ao tema
            if theme == 'dark':
                self.app.setStyleSheet(qss_dark())
                theme_name = 'escuro'
            elif theme == 'pink':
                self.app.setStyleSheet(qss_pink())
                theme_name = 'rosa'
            elif theme == 'purple':
                self.app.setStyleSheet(qss_purple())
                theme_name = 'roxo'
            elif theme == 'blue':
                self.app.setStyleSheet(qss_blue())
                theme_name = 'azul'
            else:  # light
                self.app.setStyleSheet(qss_light())
                theme_name = 'claro'
            
            # Persiste em config.yaml
            cfg = load_config()
            cfg['theme'] = theme
            save_config(cfg)
            
            if self.toast_cb:
                self.toast_cb(f"Tema {theme_name} aplicado e salvo.")
        except Exception as e:
            try:
                show_message(self.parent_window, 'Erro', f'Não foi possível aplicar/salvar o tema: {e}', ('OK',))
            except Exception:
                pass
    
    def save_company_name(self) -> None:
        """Salva o nome da empresa no banco de dados"""
        try:
            new_name = self.company_name_input.text().strip()
            if not new_name:
                show_message(self.parent_window, 'Atenção', 'Por favor, digite um nome para a empresa.', ('OK',))
                return
            
            # Atualizar no banco
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.db.execute("UPDATE company SET name=?, updated_at=? WHERE id=1", (new_name, now))
            
            self.current_company_name = new_name
            
            # Atualizar título da janela principal
            if hasattr(self.parent_window, '_update_window_title'):
                self.parent_window._update_window_title()
            
            if self.toast_cb:
                self.toast_cb(f"✓ Nome da empresa atualizado para '{new_name}'")
                
        except Exception as e:
            show_message(self.parent_window, 'Erro', f'Não foi possível salvar o nome da empresa: {e}', ('OK',))
    
    def select_company_logo(self) -> None:
        """Permite selecionar uma imagem para logo da empresa"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self.parent_window,
                "Selecionar Logo da Empresa",
                "",
                "Imagens (*.png *.jpg *.jpeg *.bmp *.gif *.svg);;Todos os arquivos (*.*)"
            )
            
            if not file_path:
                return
            
            # Copiar imagem para pasta da aplicação
            import shutil
            logo_dir = os.path.join(APP_DIR, "assets", "company")
            os.makedirs(logo_dir, exist_ok=True)
            
            # Nome do arquivo de destino
            file_ext = os.path.splitext(file_path)[1]
            dest_path = os.path.join(logo_dir, f"logo{file_ext}")
            
            # Copiar arquivo
            shutil.copy2(file_path, dest_path)
            
            # Atualizar no banco
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.db.execute("UPDATE company SET logo_path=?, updated_at=? WHERE id=1", (dest_path, now))
            
            self.current_logo_path = dest_path
            
            # Atualizar preview
            from PyQt6.QtGui import QPixmap
            pixmap = QPixmap(dest_path)
            if not pixmap.isNull():
                self.logo_preview.setPixmap(pixmap.scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
            
            # Atualizar logo no cabeçalho da janela principal
            if hasattr(self.parent_window, '_update_header_logo'):
                self.parent_window._update_header_logo(dest_path)
            
            if self.toast_cb:
                self.toast_cb("✓ Logo da empresa atualizada")
                
        except Exception as e:
            show_message(self.parent_window, 'Erro', f'Não foi possível salvar a logo: {e}', ('OK',))
    
    def remove_company_logo(self) -> None:
        """Remove a logo da empresa"""
        try:
            # Confirmar remoção
            reply = show_message(
                self.parent_window,
                'Confirmar',
                'Deseja realmente remover a logo da empresa?',
                ('Sim', 'Não'),
                default=1
            )
            
            if reply != 0:  # Não clicou em "Sim"
                return
            
            # Atualizar no banco (remover caminho)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.db.execute("UPDATE company SET logo_path=NULL, updated_at=? WHERE id=1", (now,))
            
            # Tentar deletar arquivo físico
            if self.current_logo_path and os.path.exists(self.current_logo_path):
                try:
                    os.remove(self.current_logo_path)
                except Exception:
                    pass  # Não é crítico se falhar
            
            self.current_logo_path = None
            
            # Resetar preview
            self.logo_preview.clear()
            self.logo_preview.setText("🖼️")
            
            # Remover logo do cabeçalho da janela principal
            if hasattr(self.parent_window, '_update_header_logo'):
                self.parent_window._update_header_logo(None)
            
            if self.toast_cb:
                self.toast_cb("✓ Logo removida")
                
        except Exception as e:
            show_message(self.parent_window, 'Erro', f'Não foi possível remover a logo: {e}', ('OK',))
    
    def check_for_updates(self) -> None:
        """Verifica se há atualizações disponíveis"""
        try:
            from core.updater import UpdateChecker, apply_update
            from PyQt6.QtCore import QThread
            
            # Desabilita botão durante verificação
            self.btn_check_update.setEnabled(False)
            self.btn_check_update.setText("🔄 Verificando...")
            
            # Cria thread de verificação
            class CheckThread(QThread):
                def __init__(self, parent_widget):
                    super().__init__()
                    self.parent_widget = parent_widget
                    self.has_update = False
                    self.remote_version = None
                    self.changelog = []
                
                def run(self):
                    try:
                        checker = UpdateChecker()
                        self.has_update, self.remote_version, self.changelog = checker.check_for_updates()
                    except Exception as e:
                        self.has_update = False
                        self.remote_version = None
                        self.changelog = [f"Erro: {e}"]
            
            def on_check_finished():
                # Reabilita botão
                self.btn_check_update.setEnabled(True)
                self.btn_check_update.setText("🔍 Verificar Atualizações")
                
                if check_thread.has_update:
                    # Mostra diálogo com changelog
                    changelog_text = "\n".join(f"• {item}" for item in check_thread.changelog)
                    message = f"<b>Nova versão disponível: {check_thread.remote_version}</b><br><br>"
                    message += f"<b>Novidades:</b><br>{changelog_text.replace(chr(10), '<br>')}<br><br>"
                    message += "Deseja atualizar agora?"
                    
                    reply = show_message(
                        self.parent_window,
                        'Atualização Disponível',
                        message,
                        ('Sim', 'Não'),
                        default=0
                    )
                    
                    if reply == 0:  # Clicou em "Sim"
                        self.perform_update()
                else:
                    if self.toast_cb:
                        self.toast_cb("✓ Sistema está atualizado!")
            
            check_thread = CheckThread(self)
            check_thread.finished.connect(on_check_finished)
            check_thread.start()
            
        except Exception as e:
            self.btn_check_update.setEnabled(True)
            self.btn_check_update.setText("🔍 Verificar Atualizações")
            show_message(self.parent_window, 'Erro', f'Erro ao verificar atualizações: {e}', ('OK',))
    
    def perform_update(self) -> None:
        """Executa a atualização do sistema"""
        try:
            from core.updater import apply_update
            from PyQt6.QtCore import QThread
            
            # Cria diálogo de progresso
            progress_dialog = UpdateDialog(self.parent_window)
            progress_dialog.show()
            
            # Cria thread de atualização
            class UpdateThread(QThread):
                progress_signal = pyqtSignal(int, str)
                finished_signal = pyqtSignal(bool, str)
                
                def __init__(self):
                    super().__init__()
                
                def run(self):
                    try:
                        def progress_callback(percent, message):
                            self.progress_signal.emit(percent, message)
                        
                        success = apply_update(progress_callback)
                        
                        if success:
                            self.finished_signal.emit(True, "Atualização concluída com sucesso!")
                        else:
                            self.finished_signal.emit(False, "Falha ao aplicar atualização.")
                    except Exception as e:
                        self.finished_signal.emit(False, f"Erro durante atualização: {e}")
            
            def on_progress(percent, message):
                progress_dialog.update_progress(percent, message)
            
            def on_finished(success, message):
                progress_dialog.close()
                
                if success:
                    reply = show_message(
                        self.parent_window,
                        'Atualização Concluída',
                        f'{message}\n\nO sistema precisa ser reiniciado. Reiniciar agora?',
                        ('Sim', 'Não'),
                        default=0
                    )
                    
                    if reply == 0:  # Clicou em "Sim"
                        # Reinicia aplicação
                        from PyQt6.QtWidgets import QApplication
                        QApplication.quit()
                        os.execl(sys.executable, sys.executable, *sys.argv)
                else:
                    show_message(self.parent_window, 'Erro', message, ('OK',))
            
            update_thread = UpdateThread()
            update_thread.progress_signal.connect(on_progress)
            update_thread.finished_signal.connect(on_finished)
            update_thread.start()
            
            # Salva referência para evitar garbage collection
            self.update_thread = update_thread
            
        except Exception as e:
            show_message(self.parent_window, 'Erro', f'Erro ao iniciar atualização: {e}', ('OK',))
    

class MainWindow(QMainWindow):
    user: Any

    def animate_page_change(self, index: int) -> None:
        old_index = self.pages.currentIndex()
        if old_index == index:
            return
        old_widget = cast(QWidget, self.pages.currentWidget())
        new_widget = cast(QWidget, self.pages.widget(index))
        w = self.pages.width()
        direction = 1 if index > old_index else -1
        cast(Any, new_widget).setGeometry(direction * w, 0, w, self.pages.height())
        cast(Any, new_widget).show()
        anim_old = QPropertyAnimation(old_widget, b"pos", self)
        anim_old.setDuration(250)
        anim_old.setStartValue(cast(Any, old_widget).pos())
        anim_old.setEndValue(cast(Any, old_widget).pos() - QPoint(direction * w, 0))
        anim_new = QPropertyAnimation(new_widget, b"pos", self)
        anim_new.setDuration(250)
        anim_new.setStartValue(cast(Any, new_widget).pos())
        anim_new.setEndValue(cast(Any, old_widget).pos())
        def on_anim_finished() -> None:
            self.pages.setCurrentIndex(index)
            old_widget.move(0, 0)
            new_widget.move(0, 0)
        cast(Any, anim_new.finished).connect(on_anim_finished)
        anim_old.start()
        anim_new.start()

    def __init__(self, user: Any):
        super().__init__()
        
        # Configuração do ícone da janela principal
        self._setup_window_icon()
        
        self.user = user
        
        # Buscar nome da empresa do banco
        self._update_window_title()
        
        # Tamanho automático baseado na tela disponível
        from PyQt6.QtGui import QGuiApplication
        screen = QGuiApplication.primaryScreen()
        if screen:
            screen_geometry = screen.availableGeometry()
            # Define tamanho como 70% da tela disponível
            width = int(screen_geometry.width() * 0.7)
            height = int(screen_geometry.height() * 0.75)
            self.resize(width, height)
            # Tamanho mínimo é 60% da tela ou 800x600 (o que for menor)
            min_width = min(int(screen_geometry.width() * 0.6), 800)
            min_height = min(int(screen_geometry.height() * 0.65), 600)
            self.setMinimumSize(min_width, min_height)
        else:
            # Fallback caso não consiga detectar a tela
            self.resize(1100, 700)
            self.setMinimumSize(800, 600)
        
        # Sempre usar o banco configurado pelo usuário
        from core.config import get_database_path
        current_db_path = get_database_path()
        
        # LOG IMPORTANTE: Mostrar qual banco está sendo usado
        if current_db_path.startswith('\\\\'):
            print("=" * 80)
            print("⚠️  ATENÇÃO: USANDO BANCO DE DADOS EM REDE")
            print(f"📂 Caminho: {current_db_path}")
            print("=" * 80)
        else:
            print(f"📂 Usando banco de dados: {current_db_path}")
        
        self.db = ExtendedDatabase(current_db_path)
        # Adicionar atributos para controle de visibilidade 
        setattr(self.db, 'current_role', 'common')  # valor padrão
        setattr(self.db, 'current_user', 'unknown')  # valor padrão
        # define usuário corrente para fins de auditoria (usar username/role ou 'unknown')
        try:
            setattr(self.db, 'current_user', getattr(user, "username", None) or getattr(user, "role", None) or "unknown")
        except Exception:
            setattr(self.db, 'current_user', "unknown")
        # define role separadamente para controle de visibilidade (admin | common)
        try:
            self.user_role = getattr(user, "role", None) or ("admin" if getattr(user, "username", None) == "admin" else "common")
        except Exception:
            self.user_role = "admin"
        try:
            setattr(self.db, 'current_role', self.user_role)
        except Exception:
            pass

        # Seed inicial + backup automático (após 1 hora)
        self._seed()
        self._setup_auto_backup_hourly()
        
        # Iniciar servidor web em thread separada
        self._start_web_server()

        root = QWidget(); self.setCentralWidget(root)
        hl = QHBoxLayout(root)

        # Sidebar
        self.sidebar: QListWidget = QListWidget()
        self.sidebar.setObjectName("Sidebar")
        self.sidebar.setIconSize(QSize(20, 20))
        # Ordem do menu: Dashboard, Clientes, Pronta entrega, Produção, Pedidos, Produtos, Configurações
        labels = ["Dashboard", "Clientes", "Pronta entrega", "Produção", "Pedidos", "Produtos", "Configurações"]
        for name in labels:
            item = QListWidgetItem(name)
            # Usa função segura para ícones
            icon_map = {
                "Dashboard": safe_qta_icon("ph.chart-line", color="#a78bfa"),
                "Pedidos": safe_qta_icon("ph.notebook", color="#8ab4ff"),
                "Produtos": safe_qta_icon("ph.cake", color="#a3e635"),
                "Clientes": safe_qta_icon("ph.users", color="#fca5a5"),
                "Produção": safe_qta_icon("ph.factory", color="#f97316"),
                "Pronta entrega": safe_qta_icon("ph.package", color="#22c55e"),
                "Configurações": safe_qta_icon("ph.gear", color="#93c5fd"),
            }
            item.setIcon(icon_map.get(name, QIcon()))
            self.sidebar.addItem(item)
        self.sidebar.setFixedWidth(220)
        hl.addWidget(self.sidebar)

        # Right area (Header + Pages)
        right = QWidget(); right.setObjectName("RightArea"); right_v = QVBoxLayout(right)
        # Header
        self.header = QWidget(); self.header.setObjectName("Header")
        header_l = QHBoxLayout(self.header)
        
        # Logo da empresa (se configurada)
        self.header_logo_label = None  # Referência para atualização dinâmica
        try:
            company_data = self.db.query("SELECT name, logo_path FROM company WHERE id=1")
            if company_data and company_data[0]["logo_path"]:
                logo_path = company_data[0]["logo_path"]
                if os.path.exists(logo_path):
                    from PyQt6.QtGui import QPixmap
                    self.header_logo_label = QLabel()
                    pixmap = QPixmap(logo_path)
                    if not pixmap.isNull():
                        self.header_logo_label.setPixmap(pixmap.scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                        self.header_logo_label.setFixedSize(48, 48)
                        self.header_logo_label.setStyleSheet("padding: 4px; margin-right: 12px;")
                        header_l.addWidget(self.header_logo_label)
        except Exception as e:
            print(f"Erro ao carregar logo da empresa: {e}")
        
        # Título (usar nome da empresa se configurado)
        try:
            company_data = self.db.query("SELECT name FROM company WHERE id=1")
            company_name = company_data[0]["name"] if company_data else "Confeitaria"
        except Exception:
            company_name = "Confeitaria"
        
        self.title_label = QLabel(company_name)
        header_l.addWidget(self.title_label)
        header_l.addStretch(1)
        
        # Banner de acesso ao painel web
        self.web_access_label = QLabel()
        self.web_access_label.setObjectName("WebAccessBanner")
        self.web_access_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.web_access_label.setWordWrap(False)
        self.web_access_label.setStyleSheet("""
            QLabel#WebAccessBanner {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #8b5cf6, stop:1 #6366f1);
                color: #ffffff;
                padding: 8px 16px;
                border-radius: 8px;
                font-weight: 500;
                font-size: 13px;
                margin-left: 12px;
                min-width: 350px;
            }
        """)
        self.web_access_label.setCursor(Qt.CursorShape.PointingHandCursor)
        # Obter IP real imediatamente
        local_ip = self._get_local_ip()
        port = 5000
        url_rede = f"http://{local_ip}:{port}"
        url_local = f"http://localhost:{port}"
        # Salvar URLs
        self._web_url_local = url_local
        self._web_url_rede = url_rede
        # Mostrar apenas o endereço de rede (formato do celular)
        banner_text = f"📱 Acesse pelo navegador do celular:\n    {url_rede}"
        self.web_access_label.setText(banner_text)
        self.web_access_label.setToolTip(
            f"🌐 Acesso pela rede (celular): {url_rede}\n"
            f"💻 Acesso local (este PC): {url_local}\n\n"
            f"Clique para copiar o endereço\n"
            f"Clique 2x para abrir no navegador"
        )
        # Conectar clique simples para copiar e duplo clique para abrir
        self.web_access_label.mousePressEvent = lambda ev: self._copy_web_address()
        self.web_access_label.mouseDoubleClickEvent = lambda a0: self._open_web_browser()
        header_l.addWidget(self.web_access_label)
        
        right_v.addWidget(self.header)

        # Pages
        self.pages: QStackedWidget = QStackedWidget()
        
        # Dashboard com gráficos
        self.page_dashboard: DashboardPage = DashboardPage(self.db, toast_cb=self.show_toast)
        dashboard_refresh_cb: Optional[Callable[[], None]] = self.page_dashboard.refresh
        
        self.page_orders: OrdersPage = OrdersPage(self.db, toast_cb=self.show_toast, dashboard_cb=dashboard_refresh_cb)
        self.page_products: ProductsPage = ProductsPage(self.db, toast_cb=self.show_toast)
        self.page_customers: CustomersPage = CustomersPage(self.db, toast_cb=self.show_toast)
        # conectar dashboard_cb dos clientes
        self.page_customers.dashboard_cb = dashboard_refresh_cb
        # Nova página: Produção
        self.page_production: ProductionPage = ProductionPage(self.db, toast_cb=self.show_toast)
        # Nova página: Pronta entrega (estoque disponível)
        self.page_ready: ReadyStockPage = ReadyStockPage(self.db, toast_cb=self.show_toast)
        self.page_reports: ReportsPage = ReportsPage(self.db)
        self.page_settings: SettingsPage = SettingsPage(self.db, cast(QApplication, QApplication.instance()), self, toast_cb=self.show_toast)

        # Ordem deve bater com 'labels' da sidebar: Dashboard, Clientes, Pronta entrega, Produção, Pedidos, Produtos, Configurações
        self.pages.addWidget(self.page_dashboard)
        self.pages.addWidget(self.page_customers)
        self.pages.addWidget(self.page_ready)
        self.pages.addWidget(self.page_production)
        self.pages.addWidget(self.page_orders)
        self.pages.addWidget(self.page_products)
        self.pages.addWidget(self.page_settings)

        right_v.addWidget(self.pages)
        # Rodapé com status de licença
        self.footer = QLabel("© 2025 DWM System Developer. Todos os direitos reservados. | ⏳ Verificando licença...")
        self.footer.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.footer.setStyleSheet("color: #888; font-size: 12px; margin-top: 8px;")
        right_v.addWidget(self.footer)
        
        hl.addWidget(right, 1)

        cast(Any, self.sidebar.currentRowChanged).connect(self.animate_page_change)
        self.sidebar.setCurrentRow(0)
        self._setup_menu()
        

    # ------------ Helpers ------------
    def _update_window_title(self) -> None:
        """Atualiza o título da janela com o nome da empresa do banco"""
        try:
            company_data = self.db.query("SELECT name FROM company WHERE id=1")
            if company_data:
                company_name = company_data[0]["name"]
                self.setWindowTitle(company_name)
                # Atualizar também o título no cabeçalho
                if hasattr(self, 'title_label'):
                    self.title_label.setText(company_name)
            else:
                self.setWindowTitle("Confeitaria")
        except Exception:
            self.setWindowTitle("Confeitaria")
    
    def _update_header_logo(self, logo_path: str | None = None) -> None:
        """Atualiza o logo no cabeçalho da janela"""
        try:
            from PyQt6.QtGui import QPixmap
            
            # Se logo_path não foi fornecido, buscar do banco
            if logo_path is None:
                company_data = self.db.query("SELECT logo_path FROM company WHERE id=1")
                if company_data:
                    logo_path = company_data[0]["logo_path"]
            
            # Se tem logo e arquivo existe
            if logo_path and os.path.exists(logo_path):
                pixmap = QPixmap(logo_path)
                if not pixmap.isNull():
                    # Se já existe o label, apenas atualizar a imagem
                    if self.header_logo_label:
                        self.header_logo_label.setPixmap(pixmap.scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                        self.header_logo_label.show()
                    else:
                        # Criar novo label
                        self.header_logo_label = QLabel()
                        self.header_logo_label.setPixmap(pixmap.scaled(48, 48, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                        self.header_logo_label.setFixedSize(48, 48)
                        self.header_logo_label.setStyleSheet("padding: 4px; margin-right: 12px;")
                        # Inserir no início do header
                        self.header.layout().insertWidget(0, self.header_logo_label)
            else:
                # Sem logo - esconder se existir
                if self.header_logo_label:
                    self.header_logo_label.hide()
        except Exception as e:
            print(f"Erro ao atualizar logo do cabeçalho: {e}")
    
    def _setup_window_icon(self) -> None:
        """Configura o ícone da janela principal do aplicativo."""
        try:
            # Tenta diferentes caminhos para o ícone
            icon_paths = [
                os.path.join(os.path.dirname(__file__), 'assets', 'icons', 'logo.ico'),
                os.path.join(base_dir, 'assets', 'icons', 'logo.ico'),
                os.path.join(os.getcwd(), 'assets', 'icons', 'logo.ico'),
                'assets/icons/logo.ico',
                'logo.ico'
            ]
            
            for icon_path in icon_paths:
                if os.path.exists(icon_path):
                    icon = QIcon(icon_path)
                    if not icon.isNull():
                        self.setWindowIcon(icon)
                        # Também define o ícone da aplicação para toda a sessão
                        app = QApplication.instance()
                        if app and hasattr(app, 'setWindowIcon'):
                            app.setWindowIcon(icon)
                        print(f"Ícone carregado com sucesso: {icon_path}")
                        return
            
            print("Aviso: Nenhum arquivo de ícone encontrado nos caminhos esperados")
        except Exception as e:
            print(f"Erro ao configurar ícone da janela: {e}")

    def show_toast(self, text: str) -> None:
        t = Toast(self, text)
        t.show_near_bottom_right()

    def _seed(self) -> None:
        # Insere alguns dados de exemplo se estiver vazio
        if not self.db.query("SELECT 1 FROM products LIMIT 1"):
            self.db.execute("INSERT INTO products(name, description, stock, min_stock, price) VALUES (?,?,?,?,?)",
                            ("Bolo de Chocolate", "Com cobertura de ganache", 20, 3, 85.00))
            self.db.execute("INSERT INTO products(name, description, stock, min_stock, price) VALUES (?,?,?,?,?)",
                            ("Cheesecake", "Frutas vermelhas", 10, 2, 65.00))
        if not self.db.query("SELECT 1 FROM customers LIMIT 1"):
            self.db.execute("INSERT INTO customers(name, phone, address) VALUES (?,?,?)",
                            ("Maria Oliveira", "(11) 99999-1111", "Rua A, 123"))
            self.db.execute("INSERT INTO customers(name, phone, address) VALUES (?,?,?)",
                            ("João Silva", "(11) 98888-2222", "Rua B, 45"))

    def _setup_menu(self) -> None:
        menu_bar = cast(Any, self.menuBar())
        menu = menu_bar.addMenu("Arquivo")
        
        # Add Users menu item (only for admin)
        act_users = QAction("Usuários", self)
        cast(Any, act_users.triggered).connect(lambda: UsersDialog(self.db, self).exec())
        menu.addAction(act_users)
        try:
            if getattr(self.db, "current_role", "admin") != "admin":
                act_users.setEnabled(False)
                act_users.setToolTip("Apenas administradores podem gerenciar usuários.")
        except Exception:
            pass

        # Add Network Sharing Configuration (only for admin)
        act_network = QAction("Configurar Rede Local", self)
        cast(Any, act_network.triggered).connect(self.configure_network_sharing)
        menu.addAction(act_network)
        try:
            if getattr(self.db, "current_role", "admin") != "admin":
                act_network.setEnabled(False)
                act_network.setToolTip("Apenas administradores podem configurar rede.")
        except Exception:
            pass

        # Add Database menu item (only for admin)
        act_database = QAction("Banco de Dados", self)
        cast(Any, act_database.triggered).connect(self.open_database_settings)
        menu.addAction(act_database)
        try:
            if getattr(self.db, "current_role", "admin") != "admin":
                act_database.setEnabled(False)
                act_database.setToolTip("Apenas administradores podem gerenciar banco de dados.")
        except Exception:
            pass

        act_quit = QAction("Sair", self)
        cast(Any, act_quit.triggered).connect(cast(Any, QApplication.instance()).quit)
        menu.addAction(act_quit)

    def configure_network_sharing(self) -> None:
        """Configura compartilhamento de rede para uso em múltiplos PCs"""
        import subprocess
        import platform
        import ctypes
        
        if platform.system() != "Windows":
            show_message(self, "Não Suportado", "Configuração automática de rede só está disponível para Windows.", ("OK",))
            return
        
        # Verifica se está executando como administrador
        try:
            is_admin = ctypes.windll.shell32.IsUserAnAdmin()
        except:
            is_admin = False
        
        if not is_admin:
            # Pergunta se quer reiniciar como admin
            restart_msg = (
                "⚠️ Privilégios de Administrador Necessários\n\n"
                "Para configurar o compartilhamento de rede, o aplicativo precisa ser executado como administrador.\n\n"
                "Deseja reiniciar o aplicativo como administrador agora?"
            )
            restart_choice = show_message(self, "Administrador Necessário", restart_msg, ("Cancelar", "Reiniciar como Admin"), 1)
            
            if restart_choice == 1:
                # Reinicia como administrador
                try:
                    if getattr(sys, 'frozen', False):
                        # Executável PyInstaller
                        exe_path = sys.executable
                    else:
                        # Modo desenvolvimento - executa o Python como admin
                        exe_path = sys.executable
                        params = f'"{__file__}"'
                        ctypes.windll.shell32.ShellExecuteW(None, "runas", exe_path, params, None, 1)
                        QApplication.instance().quit()
                        return
                    
                    ctypes.windll.shell32.ShellExecuteW(None, "runas", exe_path, None, None, 1)
                    QApplication.instance().quit()
                    return
                except Exception as e:
                    show_message(
                        self,
                        "Erro",
                        f"Não foi possível reiniciar como administrador:\n\n{e}\n\n"
                        f"Por favor, feche o aplicativo e execute-o manualmente como administrador\n"
                        f"(botão direito → Executar como administrador)",
                        ("OK",)
                    )
            return
        
        msg = "🌐 Configuração de Rede Local\n\nEsta opção irá configurar seu PC para compartilhar o banco de dados.\n\nDeseja continuar?"
        result = show_message(self, "Configurar Rede Local", msg, ("Sim", "Cancelar"), 0)
        if result != 0:
            return
        
        try:
            import tempfile
            from core.config import get_database_path
            import shutil
            
            # Obtém o caminho do banco atual
            try:
                current_db = get_database_path()
            except Exception:
                current_db = DB_PATH
            
            # Cria a pasta compartilhada
            share_dir = r"C:\Confeitaria_DB"
            os.makedirs(share_dir, exist_ok=True)
            
            # Copia o banco de dados atual para a pasta compartilhada
            dest_db = os.path.join(share_dir, "confeitaria.db")
            if os.path.exists(current_db) and current_db != dest_db:
                shutil.copy2(current_db, dest_db)
            elif not os.path.exists(dest_db):
                # Se não existe banco, cria um vazio
                import sqlite3
                conn = sqlite3.connect(dest_db)
                conn.close()
            
            # Executa comandos diretamente (já está como admin neste ponto)
            show_message(self, "Aguarde", "⏳ Configurando compartilhamento...\n\nIsso pode levar alguns segundos.", ("OK",))
            
            commands = [
                # Habilita descoberta de rede
                'netsh advfirewall firewall set rule group="Network Discovery" new enable=Yes',
                'netsh advfirewall firewall set rule group="File and Printer Sharing" new enable=Yes',
                # Remove compartilhamento antigo
                'net share Confeitaria_DB /DELETE /Y',
                # Cria compartilhamento com acesso guest
                f'net share Confeitaria_DB=C:\\Confeitaria_DB /GRANT:Todos,FULL /GRANT:Convidado,FULL',
                # Permissões NTFS
                'icacls C:\\Confeitaria_DB /grant "Todos:(OI)(CI)F" /T',
                'icacls C:\\Confeitaria_DB /grant "Convidado:(OI)(CI)F" /T',
                # Desabilita senha para compartilhamento
                'reg add "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v "LimitBlankPasswordUse" /t REG_DWORD /d 0 /f',
                # Habilita acesso anônimo (crítico para sem senha)
                'reg add "HKLM\\SYSTEM\\CurrentControlSet\\Services\\LanmanWorkstation\\Parameters" /v "AllowInsecureGuestAuth" /t REG_DWORD /d 1 /f',
                # Desabilita compartilhamento protegido por senha
                'reg add "HKLM\\SYSTEM\\CurrentControlSet\\Control\\Lsa" /v "EveryoneIncludesAnonymous" /t REG_DWORD /d 1 /f',
                # Permite acesso nulo (sem credenciais)
                'reg add "HKLM\\SYSTEM\\CurrentControlSet\\Services\\LanmanServer\\Parameters" /v "RestrictNullSessAccess" /t REG_DWORD /d 0 /f',
                # Ativa conta guest
                'net user guest /active:yes',
                'net user guest ""'
            ]
            
            errors = []
            for cmd in commands:
                try:
                    result = subprocess.run(
                        cmd,
                        shell=True,
                        capture_output=True,
                        text=True,
                        timeout=30
                    )
                    # Ignora alguns erros esperados
                    if result.returncode != 0:
                        error_text = result.stderr.strip()
                        # Ignora erros de "não encontrado" (compartilhamento que não existia)
                        if "2310" not in error_text and "não foi encontrado" not in error_text.lower():
                            errors.append(f"Comando: {cmd[:50]}...\nErro: {error_text[:100]}")
                except Exception as e:
                    errors.append(f"Comando: {cmd[:50]}...\nExceção: {str(e)[:100]}")
            
            # Se teve muitos erros, mostra
            if len(errors) > 3:
                show_message(
                    self,
                    "Aviso",
                    f"⚠️ Alguns comandos falharam:\n\n{errors[0]}\n\n"
                    f"Total de erros: {len(errors)}\n\n"
                    f"Continuando verificação...",
                    ("OK",)
                )
            
            # Aguarda um pouco para garantir que o compartilhamento foi criado
            import time
            time.sleep(2)
            
            # Verifica se o compartilhamento foi realmente criado
            check_share = subprocess.run(
                ["net", "share"],
                capture_output=True,
                text=True
            )
            
            if "Confeitaria_DB" not in check_share.stdout:
                show_message(
                    self,
                    "Erro de Configuração",
                    f"❌ O compartilhamento não foi criado!\n\n"
                    f"Possíveis causas:\n"
                    f"1. Você cancelou a solicitação de administrador\n"
                    f"2. O aplicativo não tem permissões suficientes\n\n"
                    f"💡 SOLUÇÃO:\n"
                    f"1. Feche o aplicativo\n"
                    f"2. Clique com botão DIREITO no ícone\n"
                    f"3. Escolha 'Executar como administrador'\n"
                    f"4. Tente novamente",
                    ("OK",)
                )
                return
            
            computer_name = os.environ.get('COMPUTERNAME', 'ESTE-PC')
            network_path = f"\\\\{computer_name}\\Confeitaria_DB\\confeitaria.db"
            
            # Tenta obter o IP local
            try:
                import socket
                hostname = socket.gethostname()
                ip_address = socket.gethostbyname(hostname)
            except:
                ip_address = "N/A"
            
            # Testa se o compartilhamento está acessível
            test_path = f"\\\\{computer_name}\\Confeitaria_DB"
            if not os.path.exists(test_path):
                show_message(
                    self,
                    "Aviso",
                    f"⚠️ Compartilhamento criado mas não acessível localmente!\n\n"
                    f"Isso pode ser normal. Teste no outro PC.\n\n"
                    f"📁 Pasta local: {share_dir}\n"
                    f"🌐 Caminho: {network_path}\n"
                    f"🔢 IP: {ip_address}",
                    ("OK",)
                )
            else:
                show_message(
                    self, 
                    "Sucesso", 
                    f"✅ Rede configurada com ACESSO SEM SENHA!\n\n"
                    f"📁 Pasta local: {share_dir}\n"
                    f"🌐 Caminho de rede: {network_path}\n"
                    f"🔢 IP deste PC: {ip_address}\n\n"
                    f"💡 No outro PC (cliente):\n"
                    f"   1. Configurações → Conectar a Rede\n"
                    f"   2. Digite: {computer_name} ou {ip_address}\n\n"
                    f"⚠️ SE PEDIR CREDENCIAIS no cliente:\n"
                    f"   • Usuário: guest\n"
                    f"   • Senha: (deixe vazio)\n"
                    f"   • Ou clique 'Conectar como convidado'\n\n"
                    f"✅ Acesso anônimo habilitado!", 
                    ("OK",)
                )
            
            self.show_toast("✅ Rede configurada!")
            
        except Exception as e:
            show_message(self, "Erro", f"Erro: {str(e)}", ("OK",))

    def open_database_settings(self) -> None:
        """Abre o diálogo de gerenciamento de banco de dados"""
        try:
            # Abre o diálogo modal
            dialog = DatabaseDialog(self, toast_cb=self.show_toast, backup_cb=self.backup_db)
            dialog.exec()
            
        except Exception as e:
            show_message(self, "Erro", f"Erro ao abrir gerenciamento de banco:\n{str(e)}", ("OK",))

    def backup_db(self) -> None:
        """Cria backup do banco de dados atualmente em uso"""
        try:
            # Obtém o caminho do banco atualmente em uso
            try:
                from core.config import get_database_path
                current_db_path = get_database_path()
            except Exception:
                current_db_path = DB_PATH  # fallback
            
            # Verifica se o banco existe
            if not os.path.isfile(current_db_path):
                show_message(self, "Erro", f"Banco de dados não encontrado:\n{current_db_path}", ("OK",))
                return
            
            os.makedirs(BACKUP_DIR, exist_ok=True)
            
            # Sugere nome baseado no banco atual
            db_name = os.path.splitext(os.path.basename(current_db_path))[0]
            timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
            suggested_name = f"backup_{db_name}_{timestamp}.zip"
            suggested_path = os.path.join(BACKUP_DIR, suggested_name)
            
            dlg = CustomFileDialog(self, caption="Salvar backup", directory=BACKUP_DIR, filter="ZIP (*.zip)", save_mode=True)
            if not dlg.exec():
                return
            path = dlg.get_selected_file()
            if not path:
                return
                
            import zipfile
            with zipfile.ZipFile(path, 'w', zipfile.ZIP_DEFLATED) as z:
                # Adiciona o banco de dados principal
                z.write(current_db_path, arcname=os.path.basename(current_db_path))
                
                # Adiciona arquivos WAL e SHM (se existirem)
                wal_path = current_db_path + "-wal"
                shm_path = current_db_path + "-shm"
                if os.path.isfile(wal_path):
                    z.write(wal_path, arcname=os.path.basename(wal_path))
                if os.path.isfile(shm_path):
                    z.write(shm_path, arcname=os.path.basename(shm_path))
                
                # Adiciona config.yaml (se existir)
                config_path = os.path.join(os.path.dirname(current_db_path), "config.yaml")
                if os.path.isfile(config_path):
                    z.write(config_path, arcname="config.yaml")
            
            show_message(self, "Backup", f"Backup criado com sucesso!\n\n📁 Origem: {current_db_path}\n💾 Backup: {path}", ("OK",))
            self.show_toast("Backup concluído ✅")
        except Exception as e:
            show_message(self, "Erro no backup", f"Erro ao criar backup:\n\n{e}", ("OK",))

    def _perform_auto_backup(self) -> None:
        """Executa backup automático de hora em hora.
        
        Cria backup silencioso sem interromper o trabalho do usuário.
        Apenas mostra notificação toast discreta.
        
        IMPORTANTE: Se o banco está em rede, faz backup DO banco da rede.
        """
        print("\n" + "="*60)
        print(f"[Backup] 🕒 Iniciando backup automático - {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
        print("="*60)
        
        try:
            # SEMPRE obtém o caminho do banco configurado (local ou rede)
            try:
                from core.config import get_database_path
                current_db_path = get_database_path()
                
                # Verifica se é caminho de rede
                if current_db_path.startswith('\\\\') or (len(current_db_path) > 1 and current_db_path[1] == ':' and ord(current_db_path[0].upper()) > ord('C')):
                    print(f"[Backup] 🌐 Banco de dados em REDE detectado: {current_db_path}")
                else:
                    print(f"[Backup] 💻 Banco de dados LOCAL: {current_db_path}")
                    
            except Exception as e:
                print(f"[Backup] ⚠️ Erro ao obter caminho do banco: {e}")
                current_db_path = DB_PATH  # fallback
                print(f"[Backup] Usando caminho padrão: {current_db_path}")
            
            # Verifica se o banco existe e está acessível
            if not os.path.isfile(current_db_path):
                print(f"[Backup] ❌ Arquivo não encontrado ou inacessível: {current_db_path}")
                self.show_toast("⚠️ Backup não realizado - Banco inacessível")
                return
            
            # Verifica tamanho do arquivo (se muito pequeno, pode estar corrompido)
            file_size = os.path.getsize(current_db_path)
            if file_size < 1024:  # Menos de 1KB
                print(f"[Backup] ⚠️ Arquivo muito pequeno ({file_size} bytes) - possível corrupção")
            
            import zipfile
            os.makedirs(BACKUP_DIR, exist_ok=True)
            
            # Nome do arquivo: (NOME_PC)_[nome_banco]_dd-mm-aaaa_hh-mm.zip
            db_name = os.path.splitext(os.path.basename(current_db_path))[0]
            timestamp = datetime.now().strftime("%d-%m-%Y_%H-%M")
            computer_name = os.environ.get('COMPUTERNAME', 'PC').replace(' ', '_')  # Nome do computador
            auto_path = os.path.join(BACKUP_DIR, f"{computer_name}_{db_name}_{timestamp}.zip")
            
            print(f"[Backup] 📦 Criando arquivo ZIP: {auto_path}")
            print(f"[Backup] 📄 Copiando de: {current_db_path}")
            
            # Cria o backup ZIP do banco (da rede ou local)
            with zipfile.ZipFile(auto_path, 'w', zipfile.ZIP_DEFLATED) as z:
                # Adiciona o banco de dados principal
                z.write(current_db_path, arcname=os.path.basename(current_db_path))
                
                # Adiciona arquivos WAL e SHM (se existirem)
                wal_path = current_db_path + "-wal"
                shm_path = current_db_path + "-shm"
                if os.path.isfile(wal_path):
                    z.write(wal_path, arcname=os.path.basename(wal_path))
                if os.path.isfile(shm_path):
                    z.write(shm_path, arcname=os.path.basename(shm_path))
                
                # Adiciona config.yaml (se existir)
                config_path = os.path.join(os.path.dirname(current_db_path), "config.yaml")
                if os.path.isfile(config_path):
                    z.write(config_path, arcname="config.yaml")
            
            backup_size = os.path.getsize(auto_path) / 1024  # KB
            print(f"[Backup] ✅ Backup local criado com sucesso ({backup_size:.2f} KB)")
            
            # Apenas mostra toast discreto - SEM popup que interrompe o trabalho
            self.show_toast(f"💾 Backup automático concluído ({timestamp})")
            
            # Envia para GitHub (tanto o ZIP quanto o .db original)
            print(f"[Backup] ☁️ Iniciando envio para GitHub...")
            self._backup_to_github(auto_path, current_db_path)
            
        except Exception as e:
            # Silenciosamente registra erro sem interromper o usuário
            print(f"[Backup] ❌ Erro ao fazer backup automático: {e}")
            import traceback
            traceback.print_exc()
            # Mostra toast de erro discreto
            try:
                self.show_toast("⚠️ Erro ao fazer backup automático")
            except Exception:
                pass

    def _backup_to_github(self, backup_file: str, source_db_path: str) -> None:
        """Envia o backup para o repositório GitHub de forma assíncrona.
        
        Args:
            backup_file: Caminho completo do arquivo de backup ZIP
            source_db_path: Caminho do banco de dados original (pode ser rede ou local)
        """
        import subprocess
        import threading
        
        def upload_backup():
            try:
                # Verifica se Git está instalado
                try:
                    git_check = subprocess.run(
                        ["git", "--version"],
                        capture_output=True,
                        text=True,
                        timeout=5,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                    )
                    if git_check.returncode != 0:
                        print("[GitHub] ❌ Git não está instalado no sistema")
                        print("[GitHub] ℹ️ Instale o Git para habilitar backup em nuvem")
                        print("[GitHub] 📥 Download: https://git-scm.com/download/win")
                        return
                    print(f"[GitHub] ✅ Git detectado: {git_check.stdout.strip()}")
                except FileNotFoundError:
                    print("[GitHub] ❌ Git não encontrado no PATH do sistema")
                    print("[GitHub] ℹ️ Backup local salvo em:", BACKUP_DIR)
                    return
                except subprocess.TimeoutExpired:
                    print("[GitHub] ⏱️ Timeout ao verificar Git")
                    return
                
                # Diretório do repositório de backup
                backup_repo_dir = os.path.join(base_dir, "Backup_Clientes")
                
                # Clone o repositório se não existir
                if not os.path.exists(backup_repo_dir):
                    print("[GitHub] 📥 Clonando repositório de backup...")
                    print("[GitHub] ℹ️ Isso pode levar alguns minutos na primeira vez...")
                    result = subprocess.run(
                        ["git", "clone", "git@github.com:W4lterBr/Backup_Clientes.git", backup_repo_dir],
                        capture_output=True,
                        text=True,
                        timeout=60,  # Aumentado para 60 segundos
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                    )
                    if result.returncode != 0:
                        print(f"[GitHub] ❌ Erro ao clonar repositório")
                        print(f"[GitHub] Detalhes: {result.stderr}")
                        print("[GitHub] ℹ️ Possíveis causas:")
                        print("[GitHub]    - Chave SSH não configurada")
                        print("[GitHub]    - Repositório não existe ou sem permissão")
                        print("[GitHub]    - Sem conexão com internet")
                        print("[GitHub] 💡 Configure o Git SSH ou execute o backup manualmente")
                        return
                    print("[GitHub] ✅ Repositório clonado com sucesso")
                else:
                    # Pull para garantir que está atualizado
                    print("[GitHub] 🔄 Atualizando repositório local...")
                    pull_result = subprocess.run(
                        ["git", "pull", "origin", "main"],
                        cwd=backup_repo_dir,
                        capture_output=True,
                        text=True,
                        timeout=30,
                        creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0
                    )
                    if pull_result.returncode != 0:
                        print(f"[GitHub] ⚠️ Aviso ao atualizar: {pull_result.stderr}")
                        # Continua mesmo se o pull falhar
                
                # Usa o caminho do banco passado como parâmetro (garante usar o da rede se foi selecionado)
                import shutil
                timestamp_db = datetime.now().strftime("%d-%m-%Y")
                db_backup_name = f"Confeitaria_Backup_{timestamp_db}.db"
                db_dest_path = os.path.join(backup_repo_dir, db_backup_name)
                
                # Copia o arquivo .db DIRETO da fonte (rede ou local)
                if os.path.isfile(source_db_path):
                    print(f"[GitHub] 📄 Copiando banco de dados DA FONTE: {source_db_path}")
                    shutil.copy2(source_db_path, db_dest_path)
                    db_size = os.path.getsize(db_dest_path) / 1024  # KB
                    print(f"[GitHub] ✅ Banco de dados copiado: {db_backup_name} ({db_size:.2f} KB)")
                else:
                    print(f"[GitHub] ⚠️ Banco de dados não encontrado: {source_db_path}")
                
                # Copia o backup ZIP também
                backup_filename = os.path.basename(backup_file)
                zip_dest_path = os.path.join(backup_repo_dir, backup_filename)
                shutil.copy2(backup_file, zip_dest_path)
                zip_size = os.path.getsize(zip_dest_path) / 1024  # KB
                print(f"[GitHub] ✅ Arquivo ZIP copiado: {backup_filename} ({zip_size:.2f} KB)")
                
                # Comandos git
                subprocess.run(["git", "config", "user.name", "Confeitaria Auto Backup"], 
                             cwd=backup_repo_dir, capture_output=True,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                subprocess.run(["git", "config", "user.email", "backup@confeitaria.local"], 
                             cwd=backup_repo_dir, capture_output=True,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                subprocess.run(["git", "add", db_backup_name, backup_filename], 
                             cwd=backup_repo_dir, capture_output=True,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                
                commit_msg = f"Backup automático - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                result_commit = subprocess.run(["git", "commit", "-m", commit_msg], 
                             cwd=backup_repo_dir, capture_output=True, text=True, timeout=10,
                             creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                
                if "nothing to commit" in result_commit.stdout:
                    print(f"[GitHub] ℹ️ Nenhuma alteração para enviar (backup idêntico ao anterior)")
                    return
                
                # Push para o GitHub (aumenta timeout para 60 segundos)
                print(f"[GitHub] 📤 Enviando para GitHub...")
                result = subprocess.run(["git", "push", "origin", "main"], 
                                      cwd=backup_repo_dir, capture_output=True, text=True, timeout=60,
                                      creationflags=subprocess.CREATE_NO_WINDOW if sys.platform == 'win32' else 0)
                
                if result.returncode == 0:
                    print(f"[GitHub] ✅ Backup enviado com sucesso para nuvem!")
                    # Usa QTimer.singleShot para chamar show_toast na thread principal (thread-safe)
                    try:
                        QTimer.singleShot(0, lambda: self.show_toast("☁️ Backup enviado para GitHub"))
                    except Exception:
                        pass
                else:
                    print(f"[GitHub] ❌ Erro ao fazer push para GitHub")
                    print(f"[GitHub] Detalhes: {result.stderr}")
                    print(f"[GitHub] ℹ️ Backup local salvo em: {BACKUP_DIR}")
                    
            except subprocess.TimeoutExpired:
                print("[GitHub] ⏱️ Timeout ao executar comando git")
                print("[GitHub] ℹ️ Verifique sua conexão com internet")
                print(f"[GitHub] ℹ️ Backup local salvo em: {BACKUP_DIR}")
            except FileNotFoundError as e:
                print(f"[GitHub] ❌ Arquivo não encontrado: {e}")
            except Exception as e:
                print(f"[GitHub] ❌ Erro ao enviar backup: {e}")
                import traceback
                traceback.print_exc()
                print(f"[GitHub] ℹ️ Backup local salvo em: {BACKUP_DIR}")
        
        # Executa em thread separada para não bloquear a UI
        thread = threading.Thread(target=upload_backup, daemon=True)
        thread.start()

    def _setup_auto_backup_hourly(self) -> None:
        """Configura backup automático de hora em hora.
        
        O primeiro backup será executado imediatamente ao iniciar,
        e depois a cada 1 hora automaticamente.
        """
        # Executa primeiro backup imediatamente
        print("[Backup] Sistema de backup automático inicializado")
        print("[Backup] Primeiro backup será executado em 5 minutos...")
        
        # Configura timer para executar a cada 1 hora (3600000 ms)
        self.backup_timer = QTimer(self)
        self.backup_timer.timeout.connect(self._perform_auto_backup)
        self.backup_timer.start(300000)  # 5 minutos para primeiro backup (300000 ms)
        
        # Após o primeiro backup, reconfigura para 1 hora
        def set_hourly_backup():
            self.backup_timer.stop()
            self.backup_timer.start(3600000)  # 1 hora em milissegundos
            print("[Backup] Timer reconfigurado para backup a cada 1 hora")
        
        # Agendar reconfiguração após primeiro backup
        QTimer.singleShot(301000, set_hourly_backup)  # 5min + 1seg
    
    def _start_web_server(self) -> None:
        """Inicia o servidor web Flask em uma thread separada para o Painel da Cozinha"""
        try:
            print("\n" + "="*60)
            print("🌐 INICIANDO SERVIDOR WEB...")
            print("="*60)
            
            # Testar importação do Flask
            try:
                import flask
                from flask_cors import CORS
                try:
                    from importlib.metadata import version
                    flask_version = version("flask")
                except ImportError:
                    flask_version = getattr(flask, '__version__', 'unknown')
                print(f"✅ Flask {flask_version} detectado")
            except ImportError as e:
                print(f"❌ Flask não está instalado: {e}")
                print("   Execute: pip install Flask Flask-CORS")
                print("   O Painel da Cozinha não será iniciado.")
                return
            
            from core.web_server import start_server
            
            # Diretório web (HTML, CSS, JS, imagens)
            web_dir = os.path.join(APP_DIR, 'web')
            
            # Verificar se o diretório web existe
            if not os.path.exists(web_dir):
                print(f"❌ Diretório web não encontrado: {web_dir}")
                print("   O Painel da Cozinha não será iniciado.")
                return
            
            print(f"✅ Diretório web encontrado: {web_dir}")
            
            # Usar o mesmo banco de dados da aplicação principal
            db_path = DB_PATH
            print(f"✅ Banco de dados: {db_path}")
            
            # Porta do servidor (padrão: 5000)
            port = 5000
            print(f"✅ Porta configurada: {port}")
            
            # Não precisa atualizar o label novamente pois já foi configurado no __init__
            
            # Iniciar servidor em thread daemon (termina quando o app principal fechar)
            print("🚀 Iniciando thread do servidor...")
            server_thread = threading.Thread(
                target=start_server,
                args=(db_path, web_dir, port),
                daemon=True,
                name="WebServerThread"
            )
            server_thread.start()
            
            # Aguardar um pouco para o servidor iniciar
            import time
            time.sleep(2)
            
            # Verificar se a thread está viva
            if server_thread.is_alive():
                print("✅ Servidor web iniciado com sucesso!")
                print("="*60 + "\n")
            else:
                print("❌ Thread do servidor encerrou inesperadamente")
                print("="*60 + "\n")
            
        except ImportError as e:
            print(f"❌ Erro de importação: {e}")
            print("   Execute: pip install Flask Flask-CORS")
            print("   O Painel da Cozinha não será iniciado.")
            print("="*60 + "\n")
        except Exception as e:
            print(f"❌ Erro ao iniciar servidor web: {e}")
            import traceback
            traceback.print_exc()
            print("   O Painel da Cozinha não será iniciado.")
            print("="*60 + "\n")
    
    def _get_local_ip(self) -> str:
        """Retorna o IP local da máquina"""
        try:
            import socket
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            local_ip = s.getsockname()[0]
            s.close()
            return local_ip
        except Exception:
            return "127.0.0.1"
    
    def _copy_web_address(self) -> None:
        """Copia o endereço de acesso ao painel web para a área de transferência"""
        try:
            if hasattr(self, '_web_url_rede'):
                from PyQt6.QtWidgets import QApplication
                app = QApplication.instance()
                if app is None:
                    self.show_toast("❌ Aplicação não inicializada")
                    return
                clipboard = app.clipboard()
                if clipboard:
                    clipboard.setText(self._web_url_rede)
                    self.show_toast(f"✅ Endereço copiado: {self._web_url_rede}")
                else:
                    self.show_toast("❌ Erro ao acessar área de transferência")
            else:
                self.show_toast("❌ Endereço do painel web não disponível")
        except Exception as e:
            self.show_toast(f"❌ Erro ao copiar: {e}")
    
    def _open_web_browser(self) -> None:
        """Abre o painel web no navegador padrão"""
        try:
            if hasattr(self, '_web_url_rede'):
                import webbrowser
                webbrowser.open(self._web_url_rede)
                self.show_toast(f"🌐 Abrindo navegador: {self._web_url_rede}")
            else:
                self.show_toast("❌ Endereço do painel web não disponível")
        except Exception as e:
            self.show_toast(f"❌ Erro ao abrir navegador: {e}")


# -----------------------------
# Styles (QSS) – Dark & Light
# -----------------------------
def qss_dark() -> str:
    return """
* { font-family: 'Segoe UI', Arial; font-size: 14px; color: #ffffff; outline: none; }
QMainWindow { background: #0f1115; }
#Header { background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #101425, stop:1 #131a2e); border-bottom: 1px solid #1d2233; }
#AppTitle { color: #f0f3ff; font-size: 20px; font-weight: 600; }
QLabel#subtitle { color: #9aa3b2; }

QListWidget#Sidebar { background: #0c0f17; color: #c8d0e0; border-right: 1px solid #1d2233; }
QListWidget#Sidebar::item { padding: 12px; margin: 6px; border-radius: 10px; }
QListWidget#Sidebar::item:selected { background: #2a2f43; color: #ffffff; }
QListWidget#Sidebar::item:hover { background: #1e3a5f; }

QPushButton { 
    background: #1a2031; 
    color: #ffffff; 
    padding: 8px 14px; 
    border: 1px solid #2c3550 !important; 
    border-radius: 10px; 
}
QPushButton:hover { 
    background: #252c45; 
    border: 1px solid #3d4a70 !important; 
}
QPushButton:pressed { 
    background: #333b5e; 
    border: 1px solid #4a5480 !important; 
}

/* Botões de ícone compactos usados em tabelas (não cortam) */
QPushButton#IconButton {
    padding: 0px;
    min-width: 28px; max-width: 28px;
    min-height: 28px; max-height: 28px;
    border-radius: 14px;
    background: #1a2031;
    border: 1px solid #2c3550 !important;
}
QPushButton#IconButton:hover { 
    background: #252c45; 
    border: 1px solid #3d4a70 !important; 
}
QPushButton#IconButton:pressed { 
    background: #333b5e; 
    border: 1px solid #4a5480 !important; 
}

QTableWidget { 
    background: #0f1422; 
    alternate-background-color: #0b1020; 
    color: #ffffff; 
    gridline-color: #232323; 
    border: 1px solid #2c3550;
    border-radius: 4px;
}
QTableWidget::item {
    padding: 8px;
    color: #ffffff;
}
QTableWidget::item:selected {
    background: #2a2f43;
    color: #ffffff;
}
QTableWidget::item:hover {
    background: #1e3a5f;
}
QHeaderView::section { 
    background: #1a2031; 
    color: #ffffff; 
    padding: 6px; 
    border: none;
}
QTableCornerButton::section {
    background: #1a2031;
    border: 1px solid #2c3550;
}

/* Scrollbars */
QScrollBar:vertical {
    background: #0f1422;
    width: 12px;
    margin: 0px;
    border: none;
}
QScrollBar::groove:vertical {
    background: #0f1422;
    border: none;
    margin: 0px;
    border-radius: 6px;
}
QScrollBar::handle:vertical {
    background: #2c3550;
    min-height: 24px;
    border-radius: 6px;
}
QScrollBar::handle:vertical:hover { background: #333b5e; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    background: none;
    height: 0px;
}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: #0b1020; }

QScrollBar:horizontal {
    background: #0f1422;
    height: 12px;
    margin: 0px;
    border: none;
}
QScrollBar::groove:horizontal {
    background: #0f1422;
    border: none;
    margin: 0px;
    border-radius: 6px;
}
QScrollBar::handle:horizontal {
    background: #2c3550;
    min-width: 24px;
    border-radius: 6px;
}
QScrollBar::handle:horizontal:hover { background: #333b5e; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    background: none;
    width: 0px;
}
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: #0b1020; }

QTabWidget::pane {
    border: 1px solid #2c3550;
    border-radius: 4px;
    background: #0f1422;
}
QTabWidget::tab-bar {
    left: 4px;
}
QTabBar::tab {
    background: #1a2031;
    color: #9aa3b2;
    border: 1px solid #2c3550;
    padding: 8px 12px;
    margin-right: 2px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
}
QTabBar::tab:selected {
    background: #2a2f43;
    color: #ffffff;
    border-bottom-color: #0f1422;
}
QTabBar::tab:!selected {
    margin-top: 2px;
}

QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit, QComboBox {
    background: #0f1422; 
    color: #ffffff; 
    border: 1px solid #2c3550; 
    border-radius: 8px; 
    padding: 6px;
}

QDateEdit {
    background: #0f1422; 
    color: #ffffff !important; 
    border: 1px solid #2c3550; 
    border-radius: 8px; 
    padding: 6px;
}

QDateEdit::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 20px;
    border-left-width: 1px;
    border-left-color: #2c3550;
    border-left-style: solid;
    border-top-right-radius: 8px;
    border-bottom-right-radius: 8px;
    background: #1a1f2e;
}

QDateEdit::down-arrow {
    image: none;
    border: none;
    width: 0px;
    height: 0px;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 5px solid #ffffff;
}

QComboBox QAbstractItemView {
    background: #1a1f2e;
    color: #ffffff;
    selection-background-color: #2a2f43;
    selection-color: #ffffff;
}
QComboBox QAbstractItemView::item {
    background: #1a1f2e;
    color: #ffffff;
    padding: 4px;
}
QComboBox QAbstractItemView::item:selected {
    background: #2a2f43;
    color: #ffffff;
}

QMenuBar { background: #0f1115; color: #ffffff; }
QMenu { background: #0f1115; color: #ffffff; }
QMenu::item:selected { background: #2a2f43; }

#Toast { background: rgba(20,24,36,0.95); }

/* Diálogos e botões */
QDialog, QMessageBox {
    background-color: #1a1d2e;
    color: #ffffff;
    border: 1px solid #2c3550;
    border-radius: 10px;
}
QDialog QLabel, QMessageBox QLabel { color: #ffffff; }
QDialog QPushButton, QMessageBox QPushButton {
    background-color: #252c45; color: #ffffff; border: 1px solid #2c3550; border-radius: 6px; padding: 6px 12px;
}
QDialog QPushButton:hover, QMessageBox QPushButton:hover { background-color: #333b5e; }

/* Report specific labels */
QLabel[report="true"] {
    color: #9aa3b2;
    font-size: 13px;
}
QLabel[total="true"] {
    color: #38bdf8;
    font-weight: bold;
}

/* Settings page groups */
QFrame#SettingsGroup {
    background: #1a1f2e;
    border: 1px solid #2c3550;
    border-radius: 12px;
    padding: 16px;
    margin: 8px 0;
}
QLabel#DbPathLabel {
    background: #0f1422;
    border: 1px solid #2c3550;
    border-radius: 6px;
    padding: 8px;
    color: #9aa3b2;
}

/* Calendar popup for QDateEdit (Dark Theme) */
QCalendarWidget {
    background-color: #1a1f2e;
    background: #1a1f2e;
    border: 1px solid #2c3550;
    border-radius: 8px;
    color: #ffffff;
}
QCalendarWidget QWidget {
    background-color: #1a1f2e;
    background: #1a1f2e;
    color: #ffffff;
}
QCalendarWidget QAbstractItemView {
    background-color: #1a1f2e;
    background: #1a1f2e;
    color: #ffffff;
    selection-background-color: #2a2f43;
    selection-color: #ffffff;
    gridline-color: #2c3550;
    alternate-background-color: #252c45;
}
QCalendarWidget QTableView {
    background-color: #1a1f2e;
    background: #1a1f2e;
    color: #ffffff;
}
QCalendarWidget QToolButton {
    background-color: #252c45;
    background: #252c45;
    color: #ffffff;
    border: 1px solid #2c3550;
    border-radius: 6px;
    padding: 6px;
}
QCalendarWidget QToolButton:hover { 
    background-color: #333b5e;
    background: #333b5e; 
}
QCalendarWidget QSpinBox {
    background-color: #1a1f2e;
    background: #1a1f2e;
    color: #ffffff;
    border: 1px solid #2c3550;
    border-radius: 6px;
    padding: 2px 6px;
}
QCalendarWidget QHeaderView::section {
    background-color: #252c45;
    background: #252c45;
    color: #ffffff;
    border: 1px solid #2c3550;
    padding: 4px;
}
"""

def qss_light() -> str:
    return """
* { font-family: 'Segoe UI', Arial; font-size: 14px; color: #1f2937; outline: none; }
QMainWindow { background: #f7f9fc; }
#Header { background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #ffffff, stop:1 #eef2ff); border-bottom: 1px solid #dfe3ec; }
#AppTitle { color: #1b2240; font-size: 20px; font-weight: 600; }
QLabel#subtitle { color: #6b7280; }
QLabel { color: #1f2937; }

QListWidget#Sidebar { background: #ffffff; color: #334155; border-right: 1px solid #e5e7eb; }
QListWidget#Sidebar::item { padding: 12px; margin: 6px; border-radius: 10px; }
QListWidget#Sidebar::item:selected { background: #e8eefc; color: #1b2240; }
QListWidget#Sidebar::item:hover { background: #dbeafe; }

/* Área principal à direita (conteúdo) */
QWidget#RightArea { background: #ffffff; }
/* Herdados na área de conteúdo ficam transparentes para não criar faixas */
QWidget#RightArea QWidget { background: transparent; }
/* Scrollbar da área de conteúdo */
QWidget#RightArea QScrollBar:vertical { background: #ffffff; }
QWidget#RightArea QScrollBar::groove:vertical { background: #ffffff; }

QPushButton { 
    background: #e5e7eb; 
    color: #111827; 
    padding: 8px 14px; 
    border: 1px solid #d1d5db !important; 
    border-radius: 10px; 
}
QPushButton:hover { 
    background: #dbeafe; 
    border: 1px solid #bfdbfe !important; 
}
QPushButton:pressed { 
    background: #c7d2fe; 
    border: 1px solid #a5b4fc !important; 
}

/* Botões de ícone compactos usados em tabelas (não cortam) */
QPushButton#IconButton {
    padding: 0px;
    min-width: 28px; max-width: 28px;
    min-height: 28px; max-height: 28px;
    border-radius: 14px;
    background: #e5e7eb;
    border: 1px solid #d1d5db !important;
}
QPushButton#IconButton:hover { 
    background: #dbeafe; 
    border: 1px solid #bfdbfe !important; 
}
QPushButton#IconButton:pressed { 
    background: #c7d2fe; 
    border: 1px solid #a5b4fc !important; 
}

QTableWidget { 
    background: #ffffff; 
    alternate-background-color: #f8fafc; 
    color: #111827; 
    gridline-color: #d1d5db; 
    border: 1px solid #e5e7eb;
    border-radius: 4px;
}
QTableWidget::item {
    padding: 8px;
    color: #111827;
}
QTableWidget::item:selected {
    background: #e8eefc;
    color: #1b2240;
}
QTableWidget::item:hover {
    background: #dbeafe;
}
QHeaderView::section { 
    background: #f3f4f6; 
    color: #1f2937; 
    padding: 6px; 
    border: none;
}
QTableCornerButton::section {
    background: #f3f4f6;
    border: 1px solid #e5e7eb;
}

/* Scrollbars (light) */
QScrollBar:vertical {
    background: #ffffff; /* mesma cor do fundo */
    width: 12px;
    margin: 0px;
    border: none;
}
QScrollBar::groove:vertical {
    background: #ffffff; /* mesma cor do fundo */
    border: none;
    margin: 0px;
    border-radius: 6px;
}
QScrollBar::handle:vertical {
    background: #d1d5db;
    min-height: 24px;
    border-radius: 6px;
}
QScrollBar::handle:vertical:hover { background: #9ca3af; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    background: none;
    height: 0px;
}
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: #ffffff; }

QScrollBar:horizontal {
    background: #ffffff; /* mesma cor do fundo */
    height: 12px;
    margin: 0px;
    border: none;
}
QScrollBar::groove:horizontal {
    background: #ffffff; /* mesma cor do fundo */
    border: none;
    margin: 0px;
    border-radius: 6px;
}
QScrollBar::handle:horizontal {
    background: #d1d5db;
    min-width: 24px;
    border-radius: 6px;
}
QScrollBar::handle:horizontal:hover { background: #9ca3af; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    background: none;
    width: 0px;
}
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: #ffffff; }

/* Extra specificity to afetar tabelas/listas na Pronta entrega */
QTableView QScrollBar:vertical, QTableWidget QScrollBar:vertical, QAbstractItemView QScrollBar:vertical {
    background: #ffffff;
}
QTableView QScrollBar::groove:vertical, QTableWidget QScrollBar::groove:vertical, QAbstractItemView QScrollBar::groove:vertical {
    background: #ffffff;
}
QTableView QScrollBar:horizontal, QTableWidget QScrollBar:horizontal, QAbstractItemView QScrollBar:horizontal {
    background: #ffffff;
}
QTableView QScrollBar::groove:horizontal, QTableWidget QScrollBar::groove:horizontal, QAbstractItemView QScrollBar::groove:horizontal {
    background: #ffffff;
}

/* Scrollbar do menu lateral (Sidebar) no tema claro */
QListWidget#Sidebar QScrollBar:vertical {
    background: #ffffff; /* igual ao fundo do Sidebar */
    width: 10px;
    margin: 0px;
    border: none;
}
QListWidget#Sidebar QScrollBar::groove:vertical { background: #ffffff; border: none; }
QListWidget#Sidebar QScrollBar::handle:vertical { background: #d1d5db; border-radius: 6px; min-height: 24px; }
QListWidget#Sidebar QScrollBar::handle:vertical:hover { background: #9ca3af; }
QListWidget#Sidebar QScrollBar::add-line:vertical, QListWidget#Sidebar QScrollBar::sub-line:vertical { height: 0px; background: none; }
QListWidget#Sidebar QScrollBar::add-page:vertical, QListWidget#Sidebar QScrollBar::sub-page:vertical { background: #ffffff; }

QTabWidget::pane {
    border: 1px solid #e5e7eb;
    border-radius: 4px;
    background: #ffffff;
}
QTabWidget::tab-bar {
    left: 4px;
}
QTabBar::tab {
    background: #f3f4f6;
    color: #6b7280;
    border: 1px solid #e5e7eb;
    padding: 8px 12px;
    margin-right: 2px;
    border-top-left-radius: 4px;
    border-top-right-radius: 4px;
}
QTabBar::tab:selected {
    background: #ffffff;
    color: #1b2240;
    border-bottom-color: #ffffff;
}
QTabBar::tab:!selected {
    margin-top: 2px;
}

QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit, QComboBox {
    background: #ffffff; 
    color: #111827; 
    border: 1px solid #d1d5db; 
    border-radius: 8px; 
    padding: 6px;
    /* Fix: ensure selection is visible on light theme (avoid black bg) */
    selection-background-color: #e8eefc;
    selection-color: #1b2240;
}

QDateEdit {
    background: #ffffff; 
    color: #111827 !important; 
    border: 1px solid #d1d5db; 
    border-radius: 8px; 
    padding: 6px;
}

QDateEdit::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 20px;
    border-left-width: 1px;
    border-left-color: #d1d5db;
    border-left-style: solid;
    border-top-right-radius: 8px;
    border-bottom-right-radius: 8px;
    background: #f9fafb;
}

QDateEdit::down-arrow {
    image: none;
    border: none;
    width: 0px;
    height: 0px;
    border-left: 5px solid transparent;
    border-right: 5px solid transparent;
    border-top: 5px solid #111827;
}
QComboBox::drop-down { border: none; }
QComboBox QAbstractItemView { 
    background: #ffffff; 
    color: #111827; 
    selection-background-color: #e8eefc;
    selection-color: #1b2240;
}
QComboBox QAbstractItemView::item {
    background: #ffffff;
    color: #111827;
    padding: 4px;
}
QComboBox QAbstractItemView::item:selected {
    background: #e8eefc;
    color: #1b2240;
}

/* Force QDateEdit popup calendar to use light theme */
QDateEdit::drop-down {
    background: #ffffff;
    border: 1px solid #d1d5db;
    border-radius: 4px;
}

/* Calendar popup for QDateEdit */
QCalendarWidget {
    background-color: #ffffff;
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    color: #111827;
}
QCalendarWidget QWidget {
    background-color: #ffffff;
    background: #ffffff;
    color: #111827;
}
QCalendarWidget QAbstractItemView {
    background-color: #ffffff;
    background: #ffffff;
    color: #111827;
    selection-background-color: #e8eefc;
    selection-color: #1b2240;
    gridline-color: #e5e7eb;
    alternate-background-color: #f9fafb;
}
QCalendarWidget QTableView {
    background-color: #ffffff;
    background: #ffffff;
    color: #111827;
}
QCalendarWidget QToolButton {
    background-color: #f3f4f6;
    background: #f3f4f6;
    color: #1f2937;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 4px 8px;
}
QCalendarWidget QToolButton:hover { 
    background-color: #e5e7eb;
    background: #e5e7eb; 
}
QCalendarWidget QSpinBox {
    background-color: #ffffff;
    background: #ffffff;
    color: #111827;
    border: 1px solid #d1d5db;
    border-radius: 6px;
    padding: 2px 6px;
}
QCalendarWidget QHeaderView::section {
    background-color: #f9fafb;
    background: #f9fafb;
    color: #1f2937;
    border: 1px solid #e5e7eb;
    padding: 4px;
}

QMenuBar { background: #f7f9fc; color: #1f2937; }
QMenu { background: #ffffff; color: #1f2937; border: 1px solid #e5e7eb; }
QMenu::item:selected { background: #eef2ff; color: #1b2240; }

/* Report specific labels */
QLabel[report="true"] {
    color: #6b7280;
    font-size: 13px;
}
QLabel[total="true"] {
    color: #2563eb;
    font-weight: bold;
}

/* Settings page groups */
QFrame#SettingsGroup {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 12px;
    padding: 16px;
    margin: 8px 0;
}
QLabel#DbPathLabel {
    background: #f9fafb;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 8px;
    color: #6b7280;
}

/* Diálogos e botões */
QDialog, QMessageBox {
    background-color: #ffffff; 
    color: #1f2937; 
    border: 1px solid #d1d5db; 
    border-radius: 10px;
}
QDialog QLabel, QMessageBox QLabel { color: #1f2937; }
QDialog QPushButton, QMessageBox QPushButton {
    background-color: #e5e7eb; 
    color: #111827; 
    border: 1px solid #d1d5db; 
    border-radius: 6px; 
    padding: 6px 12px;
}
QDialog QPushButton:hover, QMessageBox QPushButton:hover { background-color: #dbeafe; }

/* Toast */
#Toast { background: rgba(255,255,255,0.95); color: #1f2937; border: 1px solid #e5e7eb; }
"""

def qss_pink() -> str:
    """Tema Rosa - estilo romântico e suave"""
    return """
* { font-family: 'Segoe UI', Arial; font-size: 14px; color: #1f2937; outline: none; }
QMainWindow { background: #fdf2f8; }
#Header { background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #fce7f3, stop:1 #fbcfe8); border-bottom: 1px solid #f9a8d4; }
#AppTitle { color: #831843; font-size: 20px; font-weight: 600; }
QLabel#subtitle { color: #9f1239; }
QLabel { color: #1f2937; }

QListWidget#Sidebar { background: #fce7f3; color: #831843; border-right: 1px solid #f9a8d4; }
QListWidget#Sidebar::item { padding: 12px; margin: 6px; border-radius: 10px; }
QListWidget#Sidebar::item:selected { background: #f9a8d4; color: #831843; }
QListWidget#Sidebar::item:hover { background: #fbcfe8; }

QWidget#RightArea { background: #ffffff; }
QWidget#RightArea QWidget { background: transparent; }
QWidget#RightArea QScrollBar:vertical { background: #ffffff; }
QWidget#RightArea QScrollBar::groove:vertical { background: #ffffff; }

QPushButton { 
    background: #fbcfe8; 
    color: #831843; 
    padding: 8px 14px; 
    border: 1px solid #f9a8d4 !important; 
    border-radius: 10px; 
}
QPushButton:hover { 
    background: #f9a8d4; 
    border: 1px solid #f472b6 !important; 
}
QPushButton:pressed { 
    background: #f472b6; 
    border: 1px solid #ec4899 !important; 
}

QPushButton#IconButton {
    padding: 0px;
    min-width: 28px; max-width: 28px;
    min-height: 28px; max-height: 28px;
    border-radius: 14px;
    background: #fbcfe8;
    border: 1px solid #f9a8d4 !important;
}
QPushButton#IconButton:hover { 
    background: #f9a8d4; 
    border: 1px solid #f472b6 !important; 
}
QPushButton#IconButton:pressed { 
    background: #f472b6; 
    border: 1px solid #ec4899 !important; 
}

QTableWidget { 
    background: #ffffff; 
    alternate-background-color: #fdf2f8; 
    color: #111827; 
    gridline-color: #fbcfe8; 
    border: 1px solid #f9a8d4;
    border-radius: 4px;
}
QTableWidget::item { padding: 8px; color: #111827; }
QTableWidget::item:selected { background: #fbcfe8; color: #831843; }
QTableWidget::item:hover { background: #fce7f3; }
QHeaderView::section { background: #fce7f3; color: #831843; padding: 6px; border: none; }
QTableCornerButton::section { background: #fce7f3; border: 1px solid #f9a8d4; }

QScrollBar:vertical { background: #ffffff; width: 12px; margin: 0px; border: none; }
QScrollBar::groove:vertical { background: #ffffff; border: none; margin: 0px; border-radius: 6px; }
QScrollBar::handle:vertical { background: #f9a8d4; min-height: 24px; border-radius: 6px; }
QScrollBar::handle:vertical:hover { background: #f472b6; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { background: none; height: 0px; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: #ffffff; }

QScrollBar:horizontal { background: #ffffff; height: 12px; margin: 0px; border: none; }
QScrollBar::groove:horizontal { background: #ffffff; border: none; margin: 0px; border-radius: 6px; }
QScrollBar::handle:horizontal { background: #f9a8d4; min-width: 24px; border-radius: 6px; }
QScrollBar::handle:horizontal:hover { background: #f472b6; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { background: none; width: 0px; }
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: #ffffff; }

QListWidget#Sidebar QScrollBar:vertical { background: #fce7f3; width: 10px; margin: 0px; border: none; }
QListWidget#Sidebar QScrollBar::groove:vertical { background: #fce7f3; border: none; }
QListWidget#Sidebar QScrollBar::handle:vertical { background: #f9a8d4; border-radius: 6px; min-height: 24px; }
QListWidget#Sidebar QScrollBar::handle:vertical:hover { background: #f472b6; }

QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit, QComboBox {
    background: #ffffff; color: #111827; border: 1px solid #f9a8d4; border-radius: 8px; padding: 6px;
    selection-background-color: #fbcfe8; selection-color: #831843;
}

QFrame#SettingsGroup {
    background: #ffffff; border: 1px solid #f9a8d4; border-radius: 12px; padding: 16px; margin: 8px 0;
}

#Toast { background: rgba(252, 231, 243, 0.95); color: #831843; border: 1px solid #f9a8d4; }
"""

def qss_purple() -> str:
    """Tema Roxo - estilo elegante e moderno"""
    return """
* { font-family: 'Segoe UI', Arial; font-size: 14px; color: #1f2937; outline: none; }
QMainWindow { background: #faf5ff; }
#Header { background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #f3e8ff, stop:1 #e9d5ff); border-bottom: 1px solid #d8b4fe; }
#AppTitle { color: #581c87; font-size: 20px; font-weight: 600; }
QLabel#subtitle { color: #6b21a8; }
QLabel { color: #1f2937; }

QListWidget#Sidebar { background: #f3e8ff; color: #581c87; border-right: 1px solid #d8b4fe; }
QListWidget#Sidebar::item { padding: 12px; margin: 6px; border-radius: 10px; }
QListWidget#Sidebar::item:selected { background: #d8b4fe; color: #581c87; }
QListWidget#Sidebar::item:hover { background: #e9d5ff; }

QWidget#RightArea { background: #ffffff; }
QWidget#RightArea QWidget { background: transparent; }
QWidget#RightArea QScrollBar:vertical { background: #ffffff; }
QWidget#RightArea QScrollBar::groove:vertical { background: #ffffff; }

QPushButton { 
    background: #e9d5ff; 
    color: #581c87; 
    padding: 8px 14px; 
    border: 1px solid #d8b4fe !important; 
    border-radius: 10px; 
}
QPushButton:hover { 
    background: #d8b4fe; 
    border: 1px solid #c084fc !important; 
}
QPushButton:pressed { 
    background: #c084fc; 
    border: 1px solid #a855f7 !important; 
}

QPushButton#IconButton {
    padding: 0px;
    min-width: 28px; max-width: 28px;
    min-height: 28px; max-height: 28px;
    border-radius: 14px;
    background: #e9d5ff;
    border: 1px solid #d8b4fe !important;
}
QPushButton#IconButton:hover { 
    background: #d8b4fe; 
    border: 1px solid #c084fc !important; 
}
QPushButton#IconButton:pressed { 
    background: #c084fc; 
    border: 1px solid #a855f7 !important; 
}

QTableWidget { 
    background: #ffffff; 
    alternate-background-color: #faf5ff; 
    color: #111827; 
    gridline-color: #e9d5ff; 
    border: 1px solid #d8b4fe;
    border-radius: 4px;
}
QTableWidget::item { padding: 8px; color: #111827; }
QTableWidget::item:selected { background: #e9d5ff; color: #581c87; }
QTableWidget::item:hover { background: #f3e8ff; }
QHeaderView::section { background: #f3e8ff; color: #581c87; padding: 6px; border: none; }
QTableCornerButton::section { background: #f3e8ff; border: 1px solid #d8b4fe; }

QScrollBar:vertical { background: #ffffff; width: 12px; margin: 0px; border: none; }
QScrollBar::groove:vertical { background: #ffffff; border: none; margin: 0px; border-radius: 6px; }
QScrollBar::handle:vertical { background: #d8b4fe; min-height: 24px; border-radius: 6px; }
QScrollBar::handle:vertical:hover { background: #c084fc; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { background: none; height: 0px; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: #ffffff; }

QScrollBar:horizontal { background: #ffffff; height: 12px; margin: 0px; border: none; }
QScrollBar::groove:horizontal { background: #ffffff; border: none; margin: 0px; border-radius: 6px; }
QScrollBar::handle:horizontal { background: #d8b4fe; min-width: 24px; border-radius: 6px; }
QScrollBar::handle:horizontal:hover { background: #c084fc; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { background: none; width: 0px; }
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: #ffffff; }

QListWidget#Sidebar QScrollBar:vertical { background: #f3e8ff; width: 10px; margin: 0px; border: none; }
QListWidget#Sidebar QScrollBar::groove:vertical { background: #f3e8ff; border: none; }
QListWidget#Sidebar QScrollBar::handle:vertical { background: #d8b4fe; border-radius: 6px; min-height: 24px; }
QListWidget#Sidebar QScrollBar::handle:vertical:hover { background: #c084fc; }

QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit, QComboBox {
    background: #ffffff; color: #111827; border: 1px solid #d8b4fe; border-radius: 8px; padding: 6px;
    selection-background-color: #e9d5ff; selection-color: #581c87;
}

QFrame#SettingsGroup {
    background: #ffffff; border: 1px solid #d8b4fe; border-radius: 12px; padding: 16px; margin: 8px 0;
}

#Toast { background: rgba(243, 232, 255, 0.95); color: #581c87; border: 1px solid #d8b4fe; }
"""

def qss_blue() -> str:
    """Tema Azul - estilo profissional e confiável"""
    return """
* { font-family: 'Segoe UI', Arial; font-size: 14px; color: #1f2937; outline: none; }
QMainWindow { background: #eff6ff; }
#Header { background: qlineargradient(x1:0,y1:0,x2:1,y2:0, stop:0 #dbeafe, stop:1 #bfdbfe); border-bottom: 1px solid #93c5fd; }
#AppTitle { color: #1e3a8a; font-size: 20px; font-weight: 600; }
QLabel#subtitle { color: #1e40af; }
QLabel { color: #1f2937; }

QListWidget#Sidebar { background: #dbeafe; color: #1e3a8a; border-right: 1px solid #93c5fd; }
QListWidget#Sidebar::item { padding: 12px; margin: 6px; border-radius: 10px; }
QListWidget#Sidebar::item:selected { background: #93c5fd; color: #1e3a8a; }
QListWidget#Sidebar::item:hover { background: #bfdbfe; }

QWidget#RightArea { background: #ffffff; }
QWidget#RightArea QWidget { background: transparent; }
QWidget#RightArea QScrollBar:vertical { background: #ffffff; }
QWidget#RightArea QScrollBar::groove:vertical { background: #ffffff; }

QPushButton { 
    background: #bfdbfe; 
    color: #1e3a8a; 
    padding: 8px 14px; 
    border: 1px solid #93c5fd !important; 
    border-radius: 10px; 
}
QPushButton:hover { 
    background: #93c5fd; 
    border: 1px solid #60a5fa !important; 
}
QPushButton:pressed { 
    background: #60a5fa; 
    border: 1px solid #3b82f6 !important; 
}

QPushButton#IconButton {
    padding: 0px;
    min-width: 28px; max-width: 28px;
    min-height: 28px; max-height: 28px;
    border-radius: 14px;
    background: #bfdbfe;
    border: 1px solid #93c5fd !important;
}
QPushButton#IconButton:hover { 
    background: #93c5fd; 
    border: 1px solid #60a5fa !important; 
}
QPushButton#IconButton:pressed { 
    background: #60a5fa; 
    border: 1px solid #3b82f6 !important; 
}

QTableWidget { 
    background: #ffffff; 
    alternate-background-color: #eff6ff; 
    color: #111827; 
    gridline-color: #bfdbfe; 
    border: 1px solid #93c5fd;
    border-radius: 4px;
}
QTableWidget::item { padding: 8px; color: #111827; }
QTableWidget::item:selected { background: #bfdbfe; color: #1e3a8a; }
QTableWidget::item:hover { background: #dbeafe; }
QHeaderView::section { background: #dbeafe; color: #1e3a8a; padding: 6px; border: none; }
QTableCornerButton::section { background: #dbeafe; border: 1px solid #93c5fd; }

QScrollBar:vertical { background: #ffffff; width: 12px; margin: 0px; border: none; }
QScrollBar::groove:vertical { background: #ffffff; border: none; margin: 0px; border-radius: 6px; }
QScrollBar::handle:vertical { background: #93c5fd; min-height: 24px; border-radius: 6px; }
QScrollBar::handle:vertical:hover { background: #60a5fa; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { background: none; height: 0px; }
QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical { background: #ffffff; }

QScrollBar:horizontal { background: #ffffff; height: 12px; margin: 0px; border: none; }
QScrollBar::groove:horizontal { background: #ffffff; border: none; margin: 0px; border-radius: 6px; }
QScrollBar::handle:horizontal { background: #93c5fd; min-width: 24px; border-radius: 6px; }
QScrollBar::handle:horizontal:hover { background: #60a5fa; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { background: none; width: 0px; }
QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal { background: #ffffff; }

QListWidget#Sidebar QScrollBar:vertical { background: #dbeafe; width: 10px; margin: 0px; border: none; }
QListWidget#Sidebar QScrollBar::groove:vertical { background: #dbeafe; border: none; }
QListWidget#Sidebar QScrollBar::handle:vertical { background: #93c5fd; border-radius: 6px; min-height: 24px; }
QListWidget#Sidebar QScrollBar::handle:vertical:hover { background: #60a5fa; }

QLineEdit, QTextEdit, QSpinBox, QDoubleSpinBox, QDateEdit, QComboBox {
    background: #ffffff; color: #111827; border: 1px solid #93c5fd; border-radius: 8px; padding: 6px;
    selection-background-color: #bfdbfe; selection-color: #1e3a8a;
}

QFrame#SettingsGroup {
    background: #ffffff; border: 1px solid #93c5fd; border-radius: 12px; padding: 16px; margin: 8px 0;
}

#Toast { background: rgba(219, 234, 254, 0.95); color: #1e3a8a; border: 1px solid #93c5fd; }
"""

# -----------------------------
# App bootstrap
# -----------------------------
def main() -> None:
    # =====================================================================
    # FORÇAR UTF-8 NO WINDOWS PARA EVITAR ERROS COM EMOJIS
    # =====================================================================
    if sys.platform == 'win32':
        try:
            import io
            # Forçar codificação UTF-8 no stdout/stderr para suportar emojis
            if hasattr(sys.stdout, 'buffer'):
                sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
            if hasattr(sys.stderr, 'buffer'):
                sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
        except Exception as e:
            # Se falhar, continuar sem emojis - não crítico
            print(f"[AVISO] Nao foi possivel configurar UTF-8: {e}")
    
    # =====================================================================
    # =====================================================================
    # APLICAÇÃO LOCAL - SEM VALIDAÇÃO DE LICENÇA
    # =====================================================================
    # =====================================================================
    
    # Auth/Login (usará seus módulos se existirem)
    try:
        from core.services import AuthService
        from ui.dialogs.login_dialog import LoginDialog
        from core.logger import log_event, log_error, log_warning
        from ui.dialogs.custom_messagebox import CustomMessageBox  # apenas para garantir disponibilidade
        from core.config import QSS_POPUP_DARK, QSS_POPUP_LIGHT, load_config, get_database_path
        config = load_config()
        # Aplica o QSS baseado no tema ativo
        theme = config.get("theme", "light")
        qss = QSS_POPUP_DARK if theme == "dark" else QSS_POPUP_LIGHT
        
        # Log de inicialização
        log_event("="*60)
        log_event("SISTEMA CONFEITARIA INICIADO")
        log_event(f"Versão do executável: {sys.executable if getattr(sys, 'frozen', False) else 'Script Python'}")
        log_event(f"Tema aplicado: {theme}")
        log_event("="*60)
    except Exception as ex:
        AuthService = None
        LoginDialog = None
        log_event = lambda *args, **kwargs: None
        log_error = lambda *args, **kwargs: None
        log_warning = lambda *args, **kwargs: None
        qss = ""
        get_database_path = lambda: DB_PATH  # fallback
        print(f"⚠️ Erro ao carregar módulos: {ex}")

    # Cria aplicação Qt
    app = QApplication(sys.argv)
    
    # Usa o caminho do banco configurado ou o padrão
    try:
        db_path = get_database_path()
        print(f"Usando banco de dados: {db_path}")
        log_event(f"📁 Caminho do banco: {db_path}")
    except Exception as e:
        print(f"Erro ao carregar caminho do banco: {e}. Usando padrão.")
        log_warning(f"Erro ao carregar caminho configurado, usando padrão: {e}")
        db_path = get_db_path()  # Usa função em vez de constante
        log_event(f"📁 Usando caminho padrão: {db_path}")
    
    # Função para verificar se arquivo é um banco SQLite válido
    def is_valid_sqlite_file(path: str) -> bool:
        try:
            if not os.path.exists(path):
                return False
            if os.path.getsize(path) == 0:
                return False
            import sqlite3
            conn = sqlite3.connect(path)
            conn.execute("SELECT name FROM sqlite_master WHERE type='table' LIMIT 1")
            conn.close()
            return True
        except Exception:
            return False
    
    # Função para criar backup de arquivo corrompido
    def backup_corrupted_file(path: str) -> bool:
        try:
            if os.path.exists(path):
                backup_path = f"{path}.corrupted.{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                
                # Tenta mover primeiro
                try:
                    import shutil
                    shutil.move(path, backup_path)
                    print(f"Arquivo corrompido movido para: {backup_path}")
                    log_warning(f"⚠️ Banco corrompido movido para: {backup_path}")
                    return True
                except (PermissionError, OSError):
                    # Se não consegue mover, tenta deletar
                    print(f"Não foi possível mover arquivo, tentando deletar: {path}")
                    try:
                        os.remove(path)
                        print(f"Arquivo corrompido removido: {path}")
                        return True
                    except (PermissionError, OSError):
                        # Se não consegue deletar, renomeia
                        try:
                            temp_path = f"{path}.temp_corrupted"
                            os.rename(path, temp_path)
                            print(f"Arquivo corrompido renomeado para: {temp_path}")
                            return True
                        except Exception:
                            print(f"Não foi possível remover arquivo corrompido: {path}")
                            return False
        except Exception as e:
            print(f"Erro ao processar arquivo corrompido: {e}")
        return False
    
    # Tenta conectar ao banco
    try:
        # Verifica se o arquivo é válido antes de tentar conectar
        if not is_valid_sqlite_file(db_path):
            raise sqlite3.DatabaseError(f"Arquivo não é um banco de dados SQLite válido: {db_path}")
        
        db = ExtendedDatabase(db_path)
        print(f"Conectado com sucesso ao banco: {db_path}")
    except Exception as e:
        print(f"Erro ao conectar ao banco {db_path}: {e}")
        
        # Se o banco configurado falhou, tenta o padrão
        default_path = get_db_path()
        if db_path != default_path:
            print(f"Tentando banco padrão: {default_path}")
            try:
                if not is_valid_sqlite_file(default_path):
                    # Se arquivo padrão também está corrompido, remove e cria novo
                    if os.path.exists(default_path):
                        backup_corrupted_file(default_path)
                        print("Arquivo padrão corrompido removido, será criado novo banco")
                
                db = ExtendedDatabase(default_path)
                print(f"Conectado com sucesso ao banco padrão: {default_path}")
            except Exception as e2:
                print(f"Erro ao conectar ao banco padrão: {e2}")
                
                # Último recurso: criar banco temporário
                try:
                    from core.config import get_app_data_directory
                    temp_db_path = os.path.join(get_app_data_directory(), f"confeitaria_temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
                    os.makedirs(os.path.dirname(temp_db_path), exist_ok=True)
                    print(f"Criando banco temporário: {temp_db_path}")
                    db = ExtendedDatabase(temp_db_path)
                    
                    # Mostra mensagem para o usuário
                    from PyQt6.QtWidgets import QMessageBox
                    msg = QMessageBox()
                    msg.setIcon(QMessageBox.Icon.Warning)
                    msg.setWindowTitle("Banco de Dados Corrompido")
                    msg.setText("Os arquivos de banco de dados estavam corrompidos e foram substituídos.")
                    msg.setInformativeText(f"Um novo banco temporário foi criado:\n\n{temp_db_path}\n\nRecomenda-se restaurar um backup ou reconfigurar o banco nas configurações.")
                    msg.setStandardButtons(QMessageBox.StandardButton.Ok)
                    msg.exec()
                    
                except Exception as e3:
                    # Erro fatal - não conseguiu criar nenhum banco
                    try:
                        from PyQt6.QtWidgets import QMessageBox
                        msg = QMessageBox()
                        msg.setIcon(QMessageBox.Icon.Critical)
                        msg.setWindowTitle("Erro Fatal")
                        msg.setText("Não foi possível inicializar nenhum banco de dados.")
                        msg.setDetailedText(f"Erro original: {e}\nErro banco padrão: {e2}\nErro banco temporário: {e3}")
                        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
                        msg.exec()
                    except Exception:
                        print(f"ERRO FATAL: Não foi possível inicializar banco de dados.")
                        print(f"Erro original: {e}")
                        print(f"Erro banco padrão: {e2}")
                        print(f"Erro banco temporário: {e3}")
                    sys.exit(1)
        else:
            # Se já estava tentando o banco padrão e falhou
            print("Banco padrão falhou, tentando criar novo banco limpo")
            default_path = get_db_path()
            try:
                # Remove arquivo corrompido se existir
                if os.path.exists(default_path):
                    success = backup_corrupted_file(default_path)
                    if not success:
                        print("Não foi possível remover arquivo corrompido, tentando criar banco em local alternativo")
                        # Cria banco em local alternativo
                        from core.config import get_app_data_directory
                        alt_db_path = os.path.join(get_app_data_directory(), "confeitaria_recovery.db")
                        os.makedirs(os.path.dirname(alt_db_path), exist_ok=True)
                        print(f"Criando banco alternativo: {alt_db_path}")
                        db = ExtendedDatabase(alt_db_path)
                        print(f"Banco alternativo criado com sucesso: {alt_db_path}")
                    else:
                        print("Criando novo banco no local padrão")
                        db = ExtendedDatabase(default_path)
                        print(f"Novo banco criado: {default_path}")
                else:
                    print("Criando novo banco no local padrão")
                    db = ExtendedDatabase(default_path)
                    print(f"Novo banco criado: {default_path}")
                    
            except Exception as e2:
                print(f"Erro ao criar novo banco no local padrão: {e2}")
                try: log_error(f"Erro ao criar banco padrão: {e2}")
                except Exception: pass
                # Último recurso: banco temporário
                try:
                    from core.config import get_app_data_directory
                    temp_db_path = os.path.join(get_app_data_directory(), f"confeitaria_emergency_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
                    os.makedirs(os.path.dirname(temp_db_path), exist_ok=True)
                    print(f"ÚLTIMO RECURSO: Criando banco de emergência: {temp_db_path}")
                    try: log_warning(f"Criando banco de emergência: {temp_db_path}")
                    except Exception: pass
                    db = ExtendedDatabase(temp_db_path)
                    print(f"Banco de emergência criado: {temp_db_path}")
                    try: log_event(f"✅ Banco de emergência criado")
                    except Exception: pass
                except Exception as e3:
                    print(f"ERRO FATAL: Não foi possível criar nenhum banco: {e3}")
                    try: log_error(f"ERRO FATAL: Não foi possível criar banco: {e3}")
                    except Exception: pass
                    sys.exit(1)

    user = None
    
    # Sempre exige login - usa a tela original do projeto
    if LoginDialog and AuthService:
        log_event("🔐 Iniciando processo de autenticação...")
        auth = AuthService(db)
        while True:
            dlg = LoginDialog()
            if dlg.exec() == QDialog.DialogCode.Accepted:
                username, password = dlg.get_values()
                log_event(f"🔑 Tentativa de login: {username}")
                user = auth.authenticate(username, password)
                if user:
                    log_event(f"✅ Login bem-sucedido: {username} (Perfil: {user.role})")
                    break
                else:
                    log_warning(f"❌ Login falhou para usuário: {username}")
                    show_message(None, "Login falhou", "Usuário ou senha inválidos.", ("OK",))
            else:
                log_event("❌ Login cancelado pelo usuário")
                sys.exit(0)
    else:
        # Se os módulos não carregarem, mostra erro e sai
        log_error("ERRO FATAL: Módulos de autenticação não carregados")
        show_message(None, "Erro", "Não foi possível carregar os módulos de autenticação.", ("OK",))
        sys.exit(1)

    win = MainWindow(user)
    
    # Aplica tema inicial conforme configuração salva
    try:
        from core.config import load_config
        cfg = load_config()
        # Carrega o tema salvo (padrão: light)
        theme = cfg.get('theme', 'light')
    except Exception:
        theme = 'light'
    
    # Aplica o QSS correspondente ao tema
    if theme == 'dark':
        base_qss = qss_dark()
    elif theme == 'pink':
        base_qss = qss_pink()
    elif theme == 'purple':
        base_qss = qss_purple()
    elif theme == 'blue':
        base_qss = qss_blue()
    else:  # 'light' ou qualquer outro valor
        base_qss = qss_light()
    
    # Aplicar apenas o CSS base, sem concatenação adicional
    app.setStyleSheet(base_qss)
    win.show()
    sys.exit(app.exec())

if __name__ == "__main__":
    main()
